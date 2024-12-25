from django.shortcuts import render, redirect
import requests
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.urls import reverse
from django.contrib.auth import get_user_model, login, logout, authenticate
from django.db.models import Q
from rest_framework.authentication import SessionAuthentication
from rest_framework import routers, serializers, viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from drf_yasg.utils import swagger_auto_schema
from rest_framework_bulk import BulkModelViewSet
from django.db.models import Max, Q, F, Sum, FloatField,When,Value ,Avg ,Count
from django.http import StreamingHttpResponse
from .models import *
from .serializers import *
from .permissions import *
from django_cte import With
import calendar
from django.db import models
from django.db import transaction
import json
import openpyxl
from io import StringIO, TextIOWrapper
from django.db.models import OuterRef, Subquery, Max, ExpressionWrapper
# from .validations import *
import traceback
import logging
from django.core.paginator import Paginator, Page

from django.db import connection
from django.apps import apps
import pytz
import time as timer
from datetime import datetime, timedelta, time , date
import os
import csv
import math
from collections import defaultdict
import qrcode
from django.http import JsonResponse
from PIL import Image
from io import BytesIO
import base64
import random
from pyxlsb import open_workbook

def check_ok(input,type):
    if input == True:
        if type == "FG":
            return 1
        else:
            return "ok"
    else:
        return None


def convert_date_to_QRcode(date_string):
    # Split the date string into year, month, and day components
    year, month, day = date_string.split('-')
    
    # Get the last two digits of the year
    short_year = year[-2:]
    
    # Generate a random three-digit number
    random_three_digits = random.randint(100, 999)  # This generates a random number between 100 and 999
    
    # Combine them into the custom format
    formatted_date = f"{random_three_digits} {day}{month}{short_year}"
    
    return formatted_date

def divide_with_remainder(A, B):
    quotient = int(A) // int(B)
    remainder = int(A) % int(B)
    result = f"{quotient}({remainder})"
    return result

def fillterData_WIP(mc,brand,type):
    query = ViewItemmasterproductwip.objects.filter(field_mc=mc,brand=brand,field_type=type).values()
    return query
    
def getDetail_WIP(zca_search):
    query = ViewItemmasterproductwip.objects.filter(field_zca=zca_search).values()[0]
    return query

def getDetail_FG(zca_search):
    query = ViewItemmasterproductfg.objects.filter(zca=zca_search).values()[0]
    return query

def searchInfo_WIP(zca_search):
    query = ViewItemmasterproductwip.objects.filter(field_zca=zca_search).values()[0]
    return query

def searchInfo_FG(zca_search):
    query = ViewItemmasterproductfg.objects.filter(zca=zca_search).values()[0]
    return query

def wood_in_pallet(total_wood, wood_per_pallet, pallet_number):
    # คำนวณจำนวนไม้ที่เหลือหลังจากแบ่งเข้าพาเลตก่อนหน้า
    remaining_wood = total_wood - (wood_per_pallet * (pallet_number - 1))
    
    # ถ้าไม้ที่เหลือมากกว่าหรือเท่ากับจำนวนไม้ต่อพาเลต ให้คืนค่าว่าพาเลตนั้นเต็ม
    # ถ้าไม่เช่นนั้น คืนค่าไม้ที่เหลือในพาเลตนั้น
    return min(remaining_wood, wood_per_pallet) if remaining_wood > 0 else 0

def searchInfo_Operator(employee_id_input):
    try:
        result = CustomUser.objects.get(employee_id=employee_id_input).first_name
    except:
        result = None
    return result







 

class get_FGPlanning(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')

            # date_sql = "2023-10-16"
            # shift_sql = "A"

            data = []

            # planning = PlanProduction.objects.filter(pdplan_date=date_sql, pdplan_shift=shift_sql)

            def divide_with_remainder(A, B):
                quotient = A // B
                remainder = A % B
                result = f"{quotient}({remainder})"
                return result

            items_good = ViewWmsListfillplanproduction.objects.filter(send_date=date_sql, send_shift=shift_sql,product_type="FG").values()
            for i in items_good:
                query_search = ViewItemmasterproductfg.objects.filter(zca=i["zca_on"]).values('zca','name','nameen','pcpallet')[0]
                pcperpallet = int(query_search['pcpallet'])

                try:
                    i["format_date"] = i["product_date"].strftime('%d/%m')
                except:
                    i["format_date"] = ""

                try:
                    i["format_shift"] = i["product_shift"]
                except:
                    i["format_shift"] = ""

                try:
                    i["format_qty"] = divide_with_remainder(int(i["qty_good"]), pcperpallet),
                except:
                    i["format_qty"] = ""

                items_stock = FGPlanningStock.objects.filter(listfillplan_link=i["id"]).first()
                if items_stock:
                    i["location_1_select_1"] = items_stock.location_1_select_1
                    i["location_1_select_2"] = items_stock.location_1_select_2
                    i["location_1_select_3"] = items_stock.location_1_select_3
                    i["location_1_info"] = items_stock.location_1_info
                    i["location_2_select_1"] = items_stock.location_2_select_1
                    i["location_2_select_2"] = items_stock.location_2_select_2
                    i["location_2_select_3"] = items_stock.location_2_select_3
                    i["location_2_info"] = items_stock.location_2_info
                    i["location_3_select_1"] = items_stock.location_3_select_1
                    i["location_3_select_2"] = items_stock.location_3_select_2
                    i["location_3_select_3"] = items_stock.location_3_select_3
                    i["location_3_info"] = items_stock.location_3_info
                    i["location_4_select_1"] = items_stock.location_4_select_1
                    i["location_4_select_2"] = items_stock.location_4_select_2
                    i["location_4_select_3"] = items_stock.location_4_select_3
                    i["location_4_info"] = items_stock.location_4_info
                    i["location_5_select_1"] = items_stock.location_5_select_1
                    i["location_5_select_2"] = items_stock.location_5_select_2
                    i["location_5_select_3"] = items_stock.location_5_select_3
                    i["location_5_info"] = items_stock.location_5_info
                    
                    for location in range(1, 6):
                        if i[f"location_{location}_select_1"] != "":
                            i[f"location_{location}_format"] = f"{i[f'location_{location}_select_1']}{i[f'location_{location}_select_2']}/{i[f'location_{location}_select_3']}={i[f'location_{location}_info']}"
                        else:
                            if i[f"location_{location}_info"] != "":
                                i[f"location_{location}_format"] = i[f"location_{location}_info"]
                            else:
                                i[f"location_{location}_format"] = ""

                    if "" in [i[f"location_{location}_format"] for location in range(1, 6)]:
                        i["location_format"] = " > ".join([i[f"location_{location}_format"] for location in range(1, 6) if i[f"location_{location}_format"] != ""])
                    else:
                        i["location_format"] = " > ".join([i[f"location_{location}_format"] for location in range(1, 6)])

                    # i["location_format"] = " > ".join([i[f"location_{location}_format"] for location in range(1, 6)])


                else:
                    i["location_1_select_1"] = ""
                    i["location_1_select_2"] = ""
                    i["location_1_select_3"] = ""
                    i["location_1_info"] = ""
                    i["location_2_select_1"] = ""
                    i["location_2_select_2"] = ""
                    i["location_2_select_3"] = ""
                    i["location_2_info"] = ""
                    i["location_3_select_1"] = ""
                    i["location_3_select_2"] = ""
                    i["location_3_select_3"] = ""
                    i["location_3_info"] = ""
                    i["location_4_select_1"] = ""
                    i["location_4_select_2"] = ""
                    i["location_4_select_3"] = ""
                    i["location_4_info"] = ""
                    i["location_5_select_1"] = ""
                    i["location_5_select_2"] = ""
                    i["location_5_select_3"] = ""
                    i["location_5_info"] = ""

                data.append(i)

            

            return Response({'success': True, 'data': data})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class post_FGPlanning(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            request_input = request.data
            
            search_row = FGPlanningStock.objects.filter(listfillplan_link=request_input["sql_index"]).first()
            if(search_row):

                search_row.location_1_select_1 = request_input["location_1_select_1"]
                search_row.location_1_select_2 = request_input["location_1_select_2"]
                search_row.location_1_select_3 = request_input["location_1_select_3"]
                search_row.location_1_info = request_input["location_1_info"]

                search_row.location_2_select_1 = request_input["location_2_select_1"]
                search_row.location_2_select_2 = request_input["location_2_select_2"]
                search_row.location_2_select_3 = request_input["location_2_select_3"]
                search_row.location_2_info = request_input["location_2_info"]

                search_row.location_3_select_1 = request_input["location_3_select_1"]
                search_row.location_3_select_2 = request_input["location_3_select_2"]
                search_row.location_3_select_3 = request_input["location_3_select_3"]
                search_row.location_3_info = request_input["location_3_info"]

                search_row.location_4_select_1 = request_input["location_4_select_1"]
                search_row.location_4_select_2 = request_input["location_4_select_2"]
                search_row.location_4_select_3 = request_input["location_4_select_3"]
                search_row.location_4_info = request_input["location_4_info"]

                search_row.location_5_select_1 = request_input["location_5_select_1"]
                search_row.location_5_select_2 = request_input["location_5_select_2"]
                search_row.location_5_select_3 = request_input["location_5_select_3"]
                search_row.location_5_info = request_input["location_5_info"]

                search_row.save()
            else:
                query_listgood = ViewWmsListfillplanproduction.objects.get(id=int(request_input["sql_index"]))
                create_row = FGPlanningStock(
                    machine = query_listgood.machine,
                    zca_on = query_listgood.zca_on,
                    name_th = query_listgood.name_th,
                    name_en = query_listgood.name_en,

                    location_1_select_1 = request_input["location_1_select_1"],
                    location_1_select_2 = request_input["location_1_select_2"],
                    location_1_select_3 = request_input["location_1_select_3"],
                    location_1_info = request_input["location_1_info"],

                    location_2_select_1 = request_input["location_2_select_1"],
                    location_2_select_2 = request_input["location_2_select_2"],
                    location_2_select_3 = request_input["location_2_select_3"],
                    location_2_info = request_input["location_2_info"],

                    location_3_select_1 = request_input["location_3_select_1"],
                    location_3_select_2 = request_input["location_3_select_2"],
                    location_3_select_3 = request_input["location_3_select_3"],
                    location_3_info = request_input["location_3_info"],

                    location_4_select_1 = request_input["location_4_select_1"],
                    location_4_select_2 = request_input["location_4_select_2"],
                    location_4_select_3 = request_input["location_4_select_3"],
                    location_4_info = request_input["location_4_info"],

                    location_5_select_1 = request_input["location_5_select_1"],
                    location_5_select_2 = request_input["location_5_select_2"],
                    location_5_select_3 = request_input["location_5_select_3"],
                    location_5_info = request_input["location_5_info"],

                    listfillplan_link = query_listgood,
                    plan_link = query_listgood.plan_link
                )
                create_row.save()

            return Response({'success': True})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
               
class get_FGPlanningTicket(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')

            # date_sql = "2023-10-16"
            # shift_sql = "A"

            data = []

            # planning = PlanProduction.objects.filter(pdplan_date=date_sql, pdplan_shift=shift_sql)

            def divide_with_remainder(A, B):
                quotient = A // B
                remainder = A % B
                result = f"{quotient}({remainder})"
                return result

            items_good = ListTicketPlanProduction.objects.filter(send_date=date_sql, send_shift=shift_sql,product_type="FG").values()
            for i in items_good:
                query_search = ViewItemmasterproductfg.objects.filter(zca=i["zca_on"]).values('zca','name','nameen','pcpallet')[0]
                pcperpallet = int(query_search['pcpallet'])

                try:
                    i["format_date"] = i["product_date"].strftime('%d/%m')
                except:
                    i["format_date"] = ""

                try:
                    i["format_shift"] = i["product_shift"]
                except:
                    i["format_shift"] = ""

                try:
                    i["format_qty"] = divide_with_remainder(int(i["ticket_qty"]), pcperpallet),
                except:
                    i["format_qty"] = ""

                items_stock = FGPlanningStockTicket.objects.filter(listticketplan_link=i["id"]).first()
                if items_stock:
                    i["location_1_select_1"] = items_stock.location_1_select_1
                    i["location_1_select_2"] = items_stock.location_1_select_2
                    i["location_1_select_3"] = items_stock.location_1_select_3
                    i["location_1_info"] = items_stock.location_1_info
                    i["location_2_select_1"] = items_stock.location_2_select_1
                    i["location_2_select_2"] = items_stock.location_2_select_2
                    i["location_2_select_3"] = items_stock.location_2_select_3
                    i["location_2_info"] = items_stock.location_2_info
                    i["location_3_select_1"] = items_stock.location_3_select_1
                    i["location_3_select_2"] = items_stock.location_3_select_2
                    i["location_3_select_3"] = items_stock.location_3_select_3
                    i["location_3_info"] = items_stock.location_3_info
                    i["location_4_select_1"] = items_stock.location_4_select_1
                    i["location_4_select_2"] = items_stock.location_4_select_2
                    i["location_4_select_3"] = items_stock.location_4_select_3
                    i["location_4_info"] = items_stock.location_4_info
                    i["location_5_select_1"] = items_stock.location_5_select_1
                    i["location_5_select_2"] = items_stock.location_5_select_2
                    i["location_5_select_3"] = items_stock.location_5_select_3
                    i["location_5_info"] = items_stock.location_5_info
                    
                    for location in range(1, 6):
                        if i[f"location_{location}_select_1"] != "":
                            i[f"location_{location}_format"] = f"{i[f'location_{location}_select_1']}{i[f'location_{location}_select_2']}/{i[f'location_{location}_select_3']}={i[f'location_{location}_info']}"
                        else:
                            if i[f"location_{location}_info"] != "":
                                i[f"location_{location}_format"] = i[f"location_{location}_info"]
                            else:
                                i[f"location_{location}_format"] = ""

                    if "" in [i[f"location_{location}_format"] for location in range(1, 6)]:
                        i["location_format"] = " > ".join([i[f"location_{location}_format"] for location in range(1, 6) if i[f"location_{location}_format"] != ""])
                    else:
                        i["location_format"] = " > ".join([i[f"location_{location}_format"] for location in range(1, 6)])

                    # i["location_format"] = " > ".join([i[f"location_{location}_format"] for location in range(1, 6)])


                else:
                    i["location_1_select_1"] = ""
                    i["location_1_select_2"] = ""
                    i["location_1_select_3"] = ""
                    i["location_1_info"] = ""
                    i["location_2_select_1"] = ""
                    i["location_2_select_2"] = ""
                    i["location_2_select_3"] = ""
                    i["location_2_info"] = ""
                    i["location_3_select_1"] = ""
                    i["location_3_select_2"] = ""
                    i["location_3_select_3"] = ""
                    i["location_3_info"] = ""
                    i["location_4_select_1"] = ""
                    i["location_4_select_2"] = ""
                    i["location_4_select_3"] = ""
                    i["location_4_info"] = ""
                    i["location_5_select_1"] = ""
                    i["location_5_select_2"] = ""
                    i["location_5_select_3"] = ""
                    i["location_5_info"] = ""

                data.append(i)

            

            return Response({'success': True, 'data': data})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class post_FGPlanningTicket(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            request_input = request.data
            
            search_row = FGPlanningStockTicket.objects.filter(listticketplan_link=request_input["sql_index"]).first()
            if(search_row):

                search_row.location_1_select_1 = request_input["location_1_select_1"]
                search_row.location_1_select_2 = request_input["location_1_select_2"]
                search_row.location_1_select_3 = request_input["location_1_select_3"]
                search_row.location_1_info = request_input["location_1_info"]

                search_row.location_2_select_1 = request_input["location_2_select_1"]
                search_row.location_2_select_2 = request_input["location_2_select_2"]
                search_row.location_2_select_3 = request_input["location_2_select_3"]
                search_row.location_2_info = request_input["location_2_info"]

                search_row.location_3_select_1 = request_input["location_3_select_1"]
                search_row.location_3_select_2 = request_input["location_3_select_2"]
                search_row.location_3_select_3 = request_input["location_3_select_3"]
                search_row.location_3_info = request_input["location_3_info"]

                search_row.location_4_select_1 = request_input["location_4_select_1"]
                search_row.location_4_select_2 = request_input["location_4_select_2"]
                search_row.location_4_select_3 = request_input["location_4_select_3"]
                search_row.location_4_info = request_input["location_4_info"]

                search_row.location_5_select_1 = request_input["location_5_select_1"]
                search_row.location_5_select_2 = request_input["location_5_select_2"]
                search_row.location_5_select_3 = request_input["location_5_select_3"]
                search_row.location_5_info = request_input["location_5_info"]

                search_row.save()
            else:
                query_listgood = ListTicketPlanProduction.objects.get(id=int(request_input["sql_index"]))
                create_row = FGPlanningStockTicket(
                    machine = query_listgood.machine,
                    zca_on = query_listgood.zca_on,
                    name_th = query_listgood.name_th,
                    name_en = query_listgood.name_en,

                    location_1_select_1 = request_input["location_1_select_1"],
                    location_1_select_2 = request_input["location_1_select_2"],
                    location_1_select_3 = request_input["location_1_select_3"],
                    location_1_info = request_input["location_1_info"],

                    location_2_select_1 = request_input["location_2_select_1"],
                    location_2_select_2 = request_input["location_2_select_2"],
                    location_2_select_3 = request_input["location_2_select_3"],
                    location_2_info = request_input["location_2_info"],

                    location_3_select_1 = request_input["location_3_select_1"],
                    location_3_select_2 = request_input["location_3_select_2"],
                    location_3_select_3 = request_input["location_3_select_3"],
                    location_3_info = request_input["location_3_info"],

                    location_4_select_1 = request_input["location_4_select_1"],
                    location_4_select_2 = request_input["location_4_select_2"],
                    location_4_select_3 = request_input["location_4_select_3"],
                    location_4_info = request_input["location_4_info"],

                    location_5_select_1 = request_input["location_5_select_1"],
                    location_5_select_2 = request_input["location_5_select_2"],
                    location_5_select_3 = request_input["location_5_select_3"],
                    location_5_info = request_input["location_5_info"],

                    listticketplan_link = query_listgood,
                    plan_link = query_listgood.plan_link
                )
                create_row.save()

            return Response({'success': True})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class get_workplan_forklift(APIView):
    permission_classes = [PlannerPerm | AdminPerm | ManagerPerm | PISPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        # try:
            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')
            data = {}
            queryset = Forklift_Worklist.objects.filter(
                Q(receive_date=date_sql, receive_shift=shift_sql) |
                Q(send_date=date_sql, send_shift=shift_sql)
            ).order_by('id').values()

            result_warehouse = defaultdict(lambda: {})
            query_warehouse = Warehouse.objects.order_by('-id').values()
            for item in query_warehouse:
                result_warehouse[item["id"]]["name"] = item["name"]

            data["work"] = []
            
            list_fill = []
            list_withdraw = []
            list_transfer = []
            for pallet in queryset:
                if pallet["maplistfillplan_link_id"] != None and pallet["maplistfillplan_link_id"] not in list_fill:
                    list_fill.append(pallet["maplistfillplan_link_id"])
                    query_data = Forklift_Worklist.objects.filter(maplistfillplan_link=int(pallet["maplistfillplan_link_id"]))
                    pallet["total_pallet"] = len(query_data)
                    pallet["total_qty"] = 0
                    for i in query_data:
                        pallet["total_qty"] += i.qty


                    data["work"].append(pallet)

                elif pallet["maplistwithdrawplan_link_id"] != None and pallet["maplistwithdrawplan_link_id"] not in list_withdraw:
                    list_withdraw.append(pallet["maplistwithdrawplan_link_id"])

                    query_data = Forklift_Worklist.objects.filter(maplistwithdrawplan_link=int(pallet["maplistwithdrawplan_link_id"]))
                    pallet["total_pallet"] = len(query_data)
                    pallet["total_qty"] = 0
                    for i in query_data:
                        pallet["total_qty"] += i.qty
                    
                    data["work"].append(pallet)

                elif pallet["maplisttransferplan_link_id"] != None and pallet["maplisttransferplan_link_id"] not in list_transfer:
                    list_transfer.append(pallet["maplisttransferplan_link_id"])

                    query_data = Forklift_Worklist.objects.filter(maplisttransferplan_link=int(pallet["maplisttransferplan_link_id"]))
                    pallet["name_th"] = ",".join(query_data.order_by('name_th').values_list('name_th', flat=True).distinct())
                    pallet["zca_on"] = ",".join(query_data.order_by('zca_on').values_list('zca_on', flat=True).distinct())
                    pallet["machine"] = "-"
                    pallet["total_pallet"] = len(query_data)
                    pallet["total_qty"] = 0
                    for i in query_data:
                        pallet["total_qty"] += i.qty
                    
                    data["work"].append(pallet)
            
            return Response({'success': True, 'data': data})
    
class post_workplan_forklift(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        # try:
            request_input = request.data
            if request_input["work_type_select"] == "withdraw":
                select_row_query = Forklift_Worklist.objects.filter(maplistwithdrawplan_link_id=int(request_input["id_select"]))
            else:
                select_row_query = Forklift_Worklist.objects.filter(maplistfillplan_link_id=int(request_input["id_select"]))
            for select_row in select_row_query:
                if request_input["status"] == 0:
                    select_row.forklift_success = False
                    select_row.forklift_force = True
                    select_row.forklift_force_operator = request.user.employee_id
                elif request_input["status"] == 1:
                    select_row.forklift_success = True
                    select_row.forklift_force = True
                    select_row.forklift_force_operator = request.user.employee_id
                    if ("fill" in select_row.type_transport):
                        ViewWmsMapManagement_select = ViewWmsMapManagement.objects.get(forklift_link=int(select_row.id))
                        ViewWmsMapManagement_select.map_approve = 1
                        ViewWmsMapManagement_select.save()
                    elif ("withdraw" in select_row.type_transport):
                        ViewWmsMapManagement_select = ViewWmsMapManagement.objects.get(forklift_link=int(select_row.id))
                        ViewWmsMapManagement_select.map_approve = 0
                        ViewWmsMapManagement_select.save()
                    elif ("return" in select_row.type_transport):
                        ViewWmsMapManagement_select = ViewWmsMapManagement.objects.get(forklift_link=int(select_row.id))
                        ViewWmsMapManagement_select.map_approve = 1
                        ViewWmsMapManagement_select.save()
                    elif ("ticket" in select_row.type_transport):
                        ViewWmsMapManagement_select = ViewWmsMapManagement.objects.get(forklift_link=int(select_row.id))
                        ViewWmsMapManagement_select.map_approve = 1
                        ViewWmsMapManagement_select.save()

                select_row.save()

            return Response({'success': True})

class get_worklist_forklift(APIView):
    permission_classes = [PlannerPerm | AdminPerm | ManagerPerm | PISPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
            def calculate_time_difference(start_time, end_time):
                """คำนวณความต่างของเวลาในหน่วยวินาที."""
                time_diff = end_time - start_time
                return time_diff.total_seconds()
        # try:
            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')
            sevenday_sql = request.query_params.get('sevenday', 'false').lower() == 'true'

            # Parse the start date
            start_date = datetime.strptime(date_sql, "%Y-%m-%d")

            # Initialize the query filter
            query_filter = Q()

            # Add conditions based on the sevenday parameter
            if sevenday_sql:
                # Calculate the end date if sevenday is True
                end_date = start_date - timedelta(days=7)
                query_filter |= Q(receive_date__range=[end_date, start_date])
                query_filter |= Q(send_date__range=[end_date, start_date])
            else:
                # Apply the original filter if sevenday is False
                query_filter |= Q(receive_date=date_sql, receive_shift=shift_sql)
                query_filter |= Q(send_date=date_sql, send_shift=shift_sql)

            # Apply the filter to the queryset
            queryset = Forklift_Worklist.objects.filter(query_filter).order_by('forklift_success').values()

            data = {}
            result_warehouse = defaultdict(lambda: {})
            query_warehouse = Warehouse.objects.order_by('-id').values()
            for item in query_warehouse:
                result_warehouse[item["id"]]["name"] = item["name"]
            data["distance"] = 0
            data["work"] = []
            data["chart_pie"] = {'all': 0, 'inprogress': 0, 'success': 0}
            data["pallet_bar_allwork"] = {"4M": [0, 0, 0], "3M": [0, 0, 0], "2.4M": [0, 0, 0], "1.2M": [0, 0, 0]}
            data["pallet_bar_successwork"] = {"4M": [0, 0, 0], "3M": [0, 0, 0], "2.4M": [0, 0, 0], "1.2M": [0, 0, 0]}

            data["para_velocity"] = ParameterSave.objects.filter(key="para_velocity").last().value
            data["para_effciency"] = ParameterSave.objects.filter(key="para_effciency").last().value
            data["para_rentalcost"] = ParameterSave.objects.filter(key="para_rentalcost").last().value
            # [เบิก,เก็บ,โยก]
            
            data["forklift_report"] = []

            for pallet in queryset:
                    
                if "fill" in pallet["type_transport"]:
                    pallet["from"] = pallet["machine"]

                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["to"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"])
                    
                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][1] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][1] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][1] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][1] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][1] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][1] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][1] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][1] += 1

                    

                elif "withdraw" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["from"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"]) + "-R" + str(ViewWmsMapManagement_select["row"])+ "-L" + str(ViewWmsMapManagement_select["level"])
                    
                    pallet["to"] = pallet["machine"]

                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][0] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][0] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][0] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][0] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][0] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][0] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][0] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][0] += 1
                
                elif "return" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["from"] = pallet["machine"]
                    
                    pallet["to"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"])
                    

                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][0] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][0] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][0] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][0] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][0] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][0] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][0] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][0] += 1

                elif "ticket" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["from"] = pallet["machine"]
                    
                    pallet["to"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"])

                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][1] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][1] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][1] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][1] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][1] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][1] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][1] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][1] += 1
                elif "transfer" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select_from = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    ViewWmsMapManagement_select_to = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[1]
                    
                    if int(ViewWmsMapManagement_select_from["warehouse_id"]) in result_warehouse:
                        pallet["warehouse_from"] = result_warehouse[int(ViewWmsMapManagement_select_from["warehouse_id"])]["name"]
                    else:
                        pass

                    if int(ViewWmsMapManagement_select_to["warehouse_id"]) in result_warehouse:
                        pallet["warehouse_to"] = result_warehouse[int(ViewWmsMapManagement_select_to["warehouse_id"])]["name"]
                    else:
                        pass

                    pallet["from"] = pallet["warehouse_from"] + "-SZ" + str(ViewWmsMapManagement_select_from["zone"]) + "-C" + str(ViewWmsMapManagement_select_from["column"]) + "-R" + str(ViewWmsMapManagement_select_from["row"])+ "-L" + str(ViewWmsMapManagement_select_from["level"])
                    
                    pallet["to"] = pallet["warehouse_to"] + "-SZ" + str(ViewWmsMapManagement_select_to["zone"]) + "-C" + str(ViewWmsMapManagement_select_to["column"])
                    
                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][2] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][2] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][2] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][2] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][2] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][2] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][2] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][2] += 1
                
                # Priority
                if pallet["forklift_success"] == False:
                    priority_query = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    priority_query_2 = ViewWmsMapManagement.objects.filter(warehouse=priority_query["warehouse_id"],column=priority_query["column"],map_approve__in=[2,3,4,5]).values()

                    temp_search = []
                    for i in priority_query_2:
                        if i["map_approve"] == 2:
                            type = "transfer"
                            id_list = i["maplisttransferplan_link_id"]
                        elif i["map_approve"] == 3:
                            type = "transfer"
                            id_list = i["maplisttransferplan_link_id"]
                        elif i["map_approve"] == 4:
                            type = "withdraw"
                            id_list = i["maplistwithdrawplan_link_id"]
                        elif i["map_approve"] == 5:
                            type = "fill"
                            id_list = i["maplistfillplan_link_id"]


                        if [type,id_list] not in temp_search:
                            temp_search.append([type,id_list])

                    if priority_query["map_approve"] == 2:
                        type = "transfer"
                        id_list = priority_query["maplisttransferplan_link_id"]
                    elif priority_query["map_approve"] == 3:
                        type = "transfer"
                        id_list = priority_query["maplisttransferplan_link_id"]
                    elif priority_query["map_approve"] == 4:
                        type = "withdraw"
                        id_list = priority_query["maplistwithdrawplan_link_id"]
                    elif priority_query["map_approve"] == 5:
                        type = "fill"
                        id_list = priority_query["maplistfillplan_link_id"]
                    pallet["priority"] = temp_search.index([type,id_list])

                if pallet["forklift_success"] == True:
                    data["chart_pie"]["success"] += 1
                else:
                    data["chart_pie"]["inprogress"] += 1

                data["chart_pie"]["all"] += 1

                if "Lab" not in pallet["machine"]:
                    data["distance"] += getDistance(str(pallet["from"]),str(pallet["to"]))

                # Forklift Stat
                if pallet["forklift_success"] == True and pallet["forklift_force"] == False:
                    time_difference_seconds = calculate_time_difference( pallet["forklift_scan_check_time"], pallet["forklift_scan_finish_time"])

                    index_found = None
                    for index, report in enumerate(data["forklift_report"]):
                        if report["employee_id"] == pallet["forklift_scan_finish_operator"]:
                            index_found = index
                            break

                    if index_found is not None:
                        data["forklift_report"][index_found]["work_success"] += 1
                        data["forklift_report"][index_found]["total_dis"] += getDistance(pallet["from"],pallet["to"])
                        data["forklift_report"][index_found]["total_time"] += time_difference_seconds
                    else:
                        data["forklift_report"].append(
                            {
                                "employee_id":pallet["forklift_scan_finish_operator"],
                                "name":CustomUser.objects.get(employee_id=pallet["forklift_scan_finish_operator"]).first_name,
                                "efficiency": 0,
                                "work_success": 1,
                                "total_dis": getDistance(pallet["from"],pallet["to"]),
                                "total_time": time_difference_seconds
                            }
                        )

                data["work"].append(pallet)
            
            for forklift in data["forklift_report"]:
                standard =  (forklift["total_dis"]/(float(data["para_effciency"])*float(data["para_velocity"])*1000/3600))
                actual = float(forklift["total_time"])
                forklift["efficiency"] = (forklift["total_dis"]/(float(data["para_effciency"])/100*float(data["para_velocity"])*1000/3600)) / (float(forklift["total_time"])) *100
                forklift["efficiency"] = round(forklift["efficiency"], 2)
            return Response({'success': True, 'data': data})
        # except Exception as e:
        #     print("ERROR >>>",e)
        #     return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class get_DailyWorklist_forklift(APIView):
    permission_classes = [PlannerPerm | AdminPerm | ManagerPerm | PISPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
            def calculate_time_difference(start_time, end_time):
                """คำนวณความต่างของเวลาในหน่วยวินาที."""
                time_diff = end_time - start_time
                return time_diff.total_seconds()
        # try:
            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')
            sevenday_sql = request.query_params.get('sevenday', 'false').lower() == 'true'
            # ตรวจสอบช่วงเวลาและกำหนดกะ
            if shift_sql == 'C':
                start, end = 0, 8
            elif shift_sql == 'A':
                start, end = 8, 16
            elif shift_sql == 'B':
                start, end = 16, 24

            # Parse the start date
            start_date = datetime.strptime(date_sql, "%Y-%m-%d")

            # Initialize the query filter
            query_filter = Q()

            # Add conditions based on the sevenday parameter
            if sevenday_sql:
                # Calculate the end date if sevenday is True
                end_date = start_date - timedelta(days=7)
                query_filter |= Q(receive_date__range=[end_date, start_date])
                query_filter |= Q(send_date__range=[end_date, start_date])
            else:
                # Apply the original filter if sevenday is False
                query_filter |= Q(receive_date=date_sql, receive_shift=shift_sql)
                query_filter |= Q(send_date=date_sql, send_shift=shift_sql)

            # Apply the filter to the queryset
            queryset = Forklift_Worklist.objects.filter(query_filter).order_by('forklift_success').values()
            print("query set ==>",queryset.count())

            data = {}
            result_warehouse = defaultdict(lambda: {})
            query_warehouse = Warehouse.objects.order_by('-id').values()
            for item in query_warehouse:
                result_warehouse[item["id"]]["name"] = item["name"]
            data["distance"] = 0
            data["work"] = []
            data["chart_pie"] = {'all': 0, 'inprogress': 0, 'success': 0,'Dailyprogress':0}
            data["pallet_bar_allwork"] = {"4M": [0, 0, 0], "3M": [0, 0, 0], "2.4M": [0, 0, 0], "1.2M": [0, 0, 0]}
            data["pallet_bar_successwork"] = {"4M": [0, 0, 0], "3M": [0, 0, 0], "2.4M": [0, 0, 0], "1.2M": [0, 0, 0]}

            data["para_velocity"] = ParameterSave.objects.filter(key="para_velocity").last().value
            data["para_effciency"] = ParameterSave.objects.filter(key="para_effciency").last().value
            data["para_rentalcost"] = ParameterSave.objects.filter(key="para_rentalcost").last().value
            # [เบิก,เก็บ,โยก]
            
            data["forklift_report"] = []

            for pallet in queryset:
                    
                if "fill" in pallet["type_transport"]:
                    pallet["from"] = pallet["machine"]

                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["to"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"])
                    
                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][1] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][1] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][1] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][1] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][1] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][1] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][1] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][1] += 1

                    

                elif "withdraw" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["from"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"]) + "-R" + str(ViewWmsMapManagement_select["row"])+ "-L" + str(ViewWmsMapManagement_select["level"])
                    
                    pallet["to"] = pallet["machine"]

                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][0] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][0] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][0] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][0] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][0] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][0] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][0] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][0] += 1
                
                elif "return" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["from"] = pallet["machine"]
                    
                    pallet["to"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"])
                    

                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][0] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][0] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][0] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][0] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][0] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][0] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][0] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][0] += 1

                elif "ticket" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    if int(ViewWmsMapManagement_select["warehouse_id"]) in result_warehouse:
                        pallet["warehouse"] = result_warehouse[int(ViewWmsMapManagement_select["warehouse_id"])]["name"]
                    else:
                        pass
                    pallet["from"] = pallet["machine"]
                    
                    pallet["to"] = pallet["warehouse"] + "-SZ" + str(ViewWmsMapManagement_select["zone"]) + "-C" + str(ViewWmsMapManagement_select["column"])

                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][1] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][1] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][1] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][1] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][1] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][1] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][1] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][1] += 1
                elif "transfer" in pallet["type_transport"]:
                    
                    ViewWmsMapManagement_select_from = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    ViewWmsMapManagement_select_to = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[1]
                    
                    if int(ViewWmsMapManagement_select_from["warehouse_id"]) in result_warehouse:
                        pallet["warehouse_from"] = result_warehouse[int(ViewWmsMapManagement_select_from["warehouse_id"])]["name"]
                    else:
                        pass

                    if int(ViewWmsMapManagement_select_to["warehouse_id"]) in result_warehouse:
                        pallet["warehouse_to"] = result_warehouse[int(ViewWmsMapManagement_select_to["warehouse_id"])]["name"]
                    else:
                        pass

                    pallet["from"] = pallet["warehouse_from"] + "-SZ" + str(ViewWmsMapManagement_select_from["zone"]) + "-C" + str(ViewWmsMapManagement_select_from["column"]) + "-R" + str(ViewWmsMapManagement_select_from["row"])+ "-L" + str(ViewWmsMapManagement_select_from["level"])
                    
                    pallet["to"] = pallet["warehouse_to"] + "-SZ" + str(ViewWmsMapManagement_select_to["zone"]) + "-C" + str(ViewWmsMapManagement_select_to["column"])
                    
                    if pallet["product_length"] == 400:
                        data["pallet_bar_allwork"]["4M"][2] += 1
                    elif pallet["product_length"] == 300:
                        data["pallet_bar_allwork"]["3M"][2] += 1
                    elif pallet["product_length"] == 240:
                        data["pallet_bar_allwork"]["2.4M"][2] += 1
                    elif pallet["product_length"] == 120:
                        data["pallet_bar_allwork"]["1.2M"][2] += 1

                    if ( pallet["product_length"] == 400 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["4M"][2] += 1
                    elif ( pallet["product_length"] == 300 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["3M"][2] += 1
                    elif ( pallet["product_length"] == 240 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["2.4M"][2] += 1
                    elif ( pallet["product_length"] == 120 ) and ( pallet["forklift_success"] == True ):
                        data["pallet_bar_successwork"]["1.2M"][2] += 1
                
                # Priority
                if pallet["forklift_success"] == False:
                    priority_query = ViewWmsMapManagement.objects.filter(forklift_link=int(pallet["id"])).values()[0]
                    priority_query_2 = ViewWmsMapManagement.objects.filter(warehouse=priority_query["warehouse_id"],column=priority_query["column"],map_approve__in=[2,3,4,5]).values()

                    temp_search = []
                    for i in priority_query_2:
                        if i["map_approve"] == 2:
                            type = "transfer"
                            id_list = i["maplisttransferplan_link_id"]
                        elif i["map_approve"] == 3:
                            type = "transfer"
                            id_list = i["maplisttransferplan_link_id"]
                        elif i["map_approve"] == 4:
                            type = "withdraw"
                            id_list = i["maplistwithdrawplan_link_id"]
                        elif i["map_approve"] == 5:
                            type = "fill"
                            id_list = i["maplistfillplan_link_id"]


                        if [type,id_list] not in temp_search:
                            temp_search.append([type,id_list])

                    if priority_query["map_approve"] == 2:
                        type = "transfer"
                        id_list = priority_query["maplisttransferplan_link_id"]
                    elif priority_query["map_approve"] == 3:
                        type = "transfer"
                        id_list = priority_query["maplisttransferplan_link_id"]
                    elif priority_query["map_approve"] == 4:
                        type = "withdraw"
                        id_list = priority_query["maplistwithdrawplan_link_id"]
                    elif priority_query["map_approve"] == 5:
                        type = "fill"
                        id_list = priority_query["maplistfillplan_link_id"]
                    pallet["priority"] = temp_search.index([type,id_list])

                daily_start_time = start_date + timedelta(hours=start)
                daily_end_time = start_date + timedelta(hours=end)
                print(pallet["forklift_scan_finish_time"])
                if pallet["forklift_scan_finish_time"] is not None:
                    if daily_start_time <= pallet["forklift_scan_finish_time"] < daily_end_time:
                        data["chart_pie"]["Dailyprogress"] += 1
                    else:
                        data["chart_pie"]["success"] += 1
                else:
                    # Handle the case where forklift_scan_finish_time is None
                    data["chart_pie"]["inprogress"] += 1

                data["chart_pie"]["all"] += 1

                if "Lab" not in pallet["machine"]:
                    data["distance"] += getDistance(str(pallet["from"]),str(pallet["to"]))

                # Forklift Stat
                if pallet["forklift_success"] == True and pallet["forklift_force"] == False and daily_start_time <= pallet["forklift_scan_finish_time"] < daily_end_time:
                    time_difference_seconds = calculate_time_difference( pallet["forklift_scan_check_time"], pallet["forklift_scan_finish_time"])

                    index_found = None
                    for index, report in enumerate(data["forklift_report"]):
                        if report["employee_id"] == pallet["forklift_scan_finish_operator"]:
                            index_found = index
                            break

                    if index_found is not None:
                        data["forklift_report"][index_found]["work_success"] += 1
                        data["forklift_report"][index_found]["total_dis"] += getDistance(pallet["from"],pallet["to"])
                        data["forklift_report"][index_found]["total_time"] += time_difference_seconds
                    else:
                        data["forklift_report"].append(
                            {
                                "employee_id":pallet["forklift_scan_finish_operator"],
                                "name":CustomUser.objects.get(employee_id=pallet["forklift_scan_finish_operator"]).first_name,
                                "efficiency": 0,
                                "work_success": 1,
                                "total_dis": getDistance(pallet["from"],pallet["to"]),
                                "total_time": time_difference_seconds
                            }
                        )

                data["work"].append(pallet)
            
            for forklift in data["forklift_report"]:
                standard =  (forklift["total_dis"]/(float(data["para_effciency"])*float(data["para_velocity"])*1000/3600))
                actual = float(forklift["total_time"])
                forklift["efficiency"] = (forklift["total_dis"]/(float(data["para_effciency"])/100*float(data["para_velocity"])*1000/3600)) / (float(forklift["total_time"])) *100
                forklift["efficiency"] = round(forklift["efficiency"], 2)
            return Response({'success': True, 'data': data})
        # except Exception as e:
        #     print("ERROR >>>",e)
        #     return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class post_worklist_forklift(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        # try:
            request_input = request.data
            select_row = Forklift_Worklist.objects.get(id=int(request_input["id_select"]))

            if request_input["status"] == 0:
                select_row.forklift_success = False
                select_row.forklift_force = True
                select_row.forklift_force_operator = request.user.employee_id
            elif request_input["status"] == 1:
                if select_row.type_transport in ("fill", "return", "ticket", "labreturn", "badfill"):
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.get(forklift_link=int(select_row.id))
                    pallet_map_list = ViewWmsMapManagement.objects.filter(maplistfillplan_link=select_row.maplistfillplan_link,warehouse=ViewWmsMapManagement_select.warehouse,column=ViewWmsMapManagement_select.column,map_approve=5).order_by('column','row','level','sub_column').first()  # Changed to .first()
                    if pallet_map_list:
                        pallet_map_list.pallet_no = select_row.pallet_no
                        pallet_map_list.product_date = select_row.product_date
                        pallet_map_list.product_shift = select_row.product_shift
                        pallet_map_list.qty = select_row.qty
                        # Simplify your conditional logic as needed here.
                        if pallet_map_list.map_approve in [5, 3]:
                            pallet_map_list.map_approve = 1
                        elif pallet_map_list.map_approve in [4, 2]:
                            pallet_map_list.map_approve = 0
                        else:
                            pallet_map_list.map_approve = 1

                        pallet_map_list.save()      

                elif ("withdraw" == select_row.type_transport):
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.get(forklift_link=int(select_row.id))
                    ViewWmsMapManagement_select.map_approve = 0
                    ViewWmsMapManagement_select.save()
                elif ("transfer" == select_row.type_transport):

                    ViewWmsMapManagement_select_col = ViewWmsMapManagement.objects.filter(forklift_link=int(select_row.id)).last()
                    ViewWmsMapManagement_select = ViewWmsMapManagement.objects.filter(maplisttransferplan_link=select_row.maplisttransferplan_link, column=ViewWmsMapManagement_select_col.column, map_approve=3).order_by('column','row','level','sub_column').first()  # Changed to .first()
                    maplist_select = MapListTransferPallet.objects.filter(
                                                                            maplisttransferplan_link=ViewWmsMapManagement_select.maplisttransferplan_link,
                                                                            zca_on=ViewWmsMapManagement_select.zca_on,
                                                                            pallet_no=ViewWmsMapManagement_select.pallet_no
                                                                        )
                    
                    maplist_select2 = MapListTransferPallet.objects.filter(
                                                                            maplisttransferplan_link=ViewWmsMapManagement_select.maplisttransferplan_link,
                                                                            zca_on=ViewWmsMapManagement_select.zca_on,
                                                                            pallet_no=select_row.pallet_no
                                                                        )

                    source_maplist = maplist_select2[0]
                    source = ViewWmsMapManagement.objects.get(maplisttransferpallet_link=source_maplist)
                    source.map_approve = 0
                    source.pallet_no = select_row.pallet_no
                    source.save()

                    destination_maplist = maplist_select[1]
                    destination = ViewWmsMapManagement.objects.get(maplisttransferpallet_link=destination_maplist)
                    destination.map_approve = 1
                    destination.pallet_no = select_row.pallet_no
                    destination.save()


                select_row.forklift_success = True
                select_row.forklift_force = True
                select_row.forklift_force_time = datetime.now()
                select_row.forklift_force_operator = request.user.employee_id

            select_row.save()

            return Response({'success': True})
    
class post_ParameterSave(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        for key_read in request_input:
            savedata = ParameterSave(key=key_read,value=request_input[key_read],operator=request.user.employee_id)
            savedata.save()
        return Response({'success': True})
        

class get_LabManage(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        date_filter = request.query_params.getlist('date_search[]')
        mc_filter = request.query_params.getlist('machine[]')
        zca_filter = request.query_params.get('zca')
        # try:
        data = []
        
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        
        queryset = MapListFillPlan.objects.filter(work_type="fill")
        if len(mc_filter) > 0:
            queryset = queryset.filter(zca_on=zca_filter)
        if len(mc_filter) > 0:
            queryset = queryset.filter(machine__in=mc_filter)
        if len(date_filter) == 2:
            queryset = queryset.filter(receive_date__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), receive_date__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
        elif len(date_filter) == 1:
            queryset = queryset.filter(receive_date=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

        queryset = queryset.order_by('-id').values()
            
        for row in queryset:
            row["detail"] = MapListFillPallet.objects.filter(maplistfillplan_link=row["id"]).values()
            row["total_pallet"] = len(row["detail"])
            try:

                try:
                    queryset_lab = Tiger_GoodsReceive.objects.filter(idmainfromwms=int(row["listfillplan_link_id"]), typez="good").values().first()
                except Exception as e:
                    traceback.print_exc()  # This will print the full stack trace in the server logs
                    queryset_lab = {"statuslab":None}

                if queryset_lab["statuslab"] == None:
                    row["read_status_lab"] = "รอผลตรวจแลป"
                    for pallet in row["detail"]:
                        pallet["read_status_lab_pallet"] = "รอผลตรวจแลป"

                elif queryset_lab["statuslab"] == "pass":
                    for pallet in row["detail"]:
                        pallet["read_status_lab_pallet"] = "ผ่าน"

                elif queryset_lab["statuslab"] == "lock":
                    queryset_lock = Tiger_LockPallet.objects.get(id_goodsreceive_to_tiger=queryset_lab["id"]).values()

                    numbers = range(1, len(row["detail"]+1))
                    strings_pallets = [f"numberpallet{num:02}" for num in numbers]

                    for index,pallet in enumerate(row["detail"]):
                        if queryset_lock[strings_pallets[index]] == "ไม่ผ่าน":
                            pallet["read_status_lab_pallet"] = False
                        else:
                            pallet["read_status_lab_pallet"] = "ผ่าน"

                # row["detail"] = [row["detail"][i:i+5] for i in range(0, len(row["detail"]), 5)]
            except Exception as e:
                print("Error Error:",e)
                pass

        data = queryset


        latest_mapid_subquery = ViewWmsMapManagement.objects.filter(
            mapid=OuterRef('mapid'), 
            level=OuterRef('level'), 
            sub_column=OuterRef('sub_column')
        ).exclude(map_approve = 2).order_by('-created_at').values('created_at')[:1]
                        
        queryset_map = ViewWmsMapManagement.objects.annotate(latest_created_at = Subquery(latest_mapid_subquery)).filter(map_approve=1 ,created_at=F('latest_created_at')  ).values("zca_on","name_th","machine","qty","product_length","lab","lock","damaged")
        
        total_pallet = 0
        total_pass = 0
        total_lock = 0
        total_waiting = 0
        total_damaged = 0
        data_product = []

        data_temp_machine = defaultdict(lambda: defaultdict({}))
        mc_product = ["HS3", "HS4", "HS5", "HS6", "HS7", "HS8", "HS9",
            "CT1", "CT2", "CT3", "CT4",
            "XY1", "CM5", "CM6", "CM7", "CM8",
            "AS1", "PK1", "PK2", "PK3", "PK4", "PK5", "PK6",
            "DET", "MS1", "OC1", "OC2", "DP1", "DP2", "OS1",
            "PL1", "RT1", "RT2", "SD1", "SEG"]
        for mc in mc_product:
            data_temp_machine[mc] = {"waiting":0,"pass":0,"lock":0,"damaged":0}

        for pallet in queryset_map:
            search = False
            for item_detail in data_product:
                if item_detail["zca_on"] == pallet["zca_on"] and item_detail["machine"] == pallet["machine"]:
                    search = True
                    item_detail["pallet_total"] +=1
                    if pallet["lab"] == 1:
                        item_detail["pallet_pass"] += 1
                    elif pallet["lab"] == 2:
                        item_detail["pallet_lock"] += 1
                    else:
                        item_detail["pallet_waiting"] += 1

                    if pallet["damaged"]:
                        item_detail["pallet_damaged"] += 1

                    break
            
            if not search:
                pallet["pallet_total"] = 1
                pallet["pallet_waiting"] = 0
                pallet["pallet_pass"] = 0
                pallet["pallet_lock"] = 0
                pallet["pallet_damaged"] = 0

                if pallet["lab"] == 1:
                    pallet["pallet_pass"] += 1
                elif pallet["lab"] == 2:
                    pallet["pallet_lock"] += 1
                else:
                    pallet["pallet_waiting"] += 1

                if pallet["damaged"]:
                    pallet["pallet_damaged"] += 1

                data_product.append(pallet)

            if pallet["machine"] in data_temp_machine.keys():
                if pallet["lab"] == 1:
                    data_temp_machine[pallet["machine"]]["pass"] += 1
                elif pallet["lab"] == 2:
                    data_temp_machine[pallet["machine"]]["lock"] += 1
                else:
                    data_temp_machine[pallet["machine"]]["waiting"] += 1

                if pallet["damaged"]:
                    data_temp_machine[pallet["machine"]]["damaged"] += 1

            total_pallet += 1
            if pallet["lab"] == 1:
                total_pass += 1
            elif pallet["lab"] == 2:
                total_lock += 1
            else:
                total_waiting += 1

            if pallet["damaged"]:
                total_damaged += 1

        data_machine_waiting = []
        data_machine_pass = []
        data_machine_lock = []
        data_machine_damaged = []

        for mc in data_temp_machine.keys():
            data_machine_waiting.append(data_temp_machine[mc]["waiting"])
            data_machine_pass.append(data_temp_machine[mc]["pass"])
            data_machine_lock.append(data_temp_machine[mc]["lock"])
            data_machine_damaged.append(data_temp_machine[mc]["damaged"])

        dropdownlist_zca = ViewItemmasterproductwip.objects.values("field_zca","field_name").distinct()
        datalist_zca = [{"label":"ทั้งหมด", "value":"ALL"}]
        for i in dropdownlist_zca:
            datalist_zca.append({"label":i["field_zca"] + " " + i["field_name"], "value":i["field_zca"]})

        return Response({
            'success': True, 
            'data': data, 
            'data_product': data_product, 
            'data_machine_waiting':data_machine_waiting,
            'data_machine_pass':data_machine_pass,
            'data_machine_lock':data_machine_lock,
            'data_machine_damaged':data_machine_damaged,
            'total_pallet':total_pallet,
            'total_waiting':total_waiting,
            'total_pass':total_pass,
            'total_lock':total_lock,
            'total_damaged':total_damaged,
            'dropdownlist_zca':datalist_zca
        })
        # except Exception as e:
        #     print("ERROR >>>",e)
        #     return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class get_approvereturn(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')

            mc = ListReturnPlanProduction.objects.filter(return_keyin="success",return_date=date_sql, return_shift=shift_sql).values_list('return_machine', flat=True).distinct()

            data = {}

            for mc in mc:
                if mc not in data.keys():
                    data[mc] = {}
                    data[mc]["items"] = []

                data[mc]["items"].extend(ListReturnPlanProduction.objects.filter(return_keyin="success",return_machine=mc, return_date=date_sql, return_shift=shift_sql).values())

            for i in data:
                for j in data[i]["items"]:
                    list_select_pallet = ListReturnPalletProduction.objects.filter(returnplan_link=int(j["id"])).values("id","zca_on","product_date","product_shift","pallet_no","qty","pcsperpallet","return_type")
                    j["operator_keyin_name"] = searchInfo_Operator(j["operator_keyin"])
                    j["operator_approve_name"] = searchInfo_Operator(j["operator_approve"])
                    # Initialize a dictionary to hold the grouped data
                    grouped_data = defaultdict(lambda: defaultdict(list))

                    # Iterate over each item in the data list
                    for item in list_select_pallet:
                        # Use the 'receive_date' and 'receive_shift' to group the items
                        grouped_data[item['product_date']][item['product_shift']].append(item)

                    # Now, transform the grouped data into the desired list format
                    grouped_list = []
                    for date, shifts in grouped_data.items():
                        for shift, pallets in shifts.items():
                            grouped_list.append({
                                "product_date": date,
                                "product_shift": shift,
                                "total_pallets": len(pallets),
                                "pallets": pallets,
                            })

                    list_good = []
                    list_bad = []
                    for k in list_select_pallet:
                        if k["return_type"] == 1:
                            list_good.append(k)
                        elif k["return_type"] == 2:
                            list_bad.append(k)
                    query_search = ViewItemmasterproductwip.objects.filter(field_zca=j["zca_on"]).values('field_zca','field_name','field_nameeng','field_mc','pcsperpallet')[0]
                    j["pcperpallet"] = int(query_search['pcsperpallet'])
                    
                    j["total-good"] = len(list_good)
                    j["total-bad"] = len(list_bad)

                    j["pallet_list"] = grouped_list

                    j["note_good"] = []
                    j["note_bad"] = []
                    select_note = ListReturnPlanNoteProduction.objects.filter(returnplan_link=int(j["id"])).values()
                    for note in select_note:
                        if note["type"] == "good":
                            j["note_good"].append({"value":note["message"],"label":note["message"]})
                        elif note["type"] == "bad":
                            j["note_bad"].append({"value":note["message"],"label":note["message"]})
                    
                        
            for i in data:
                list_all = 0
                list_wait = 0
                list_approve = 0
                for j in data[i]["items"]:
                    list_all += 1
                    if j["return_approve"] == "success":
                        list_approve += 1
                    else:
                        list_wait += 1
                data[i]["list_all"] = list_all
                data[i]["list_wait"] = list_wait
                data[i]["list_approve"] = list_approve

            # DontsendData_mc = DontsendData.objects.filter(date=date_sql, shift=shift_sql).values_list('machine', flat=True).distinct()
            # dontsend = []
            # for mc in DontsendData_mc:
            #     dontsend.append(mc)

            return Response({'success': True, 'data': data})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class post_approvereturn(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        # try:
            request_input = request.data
            select_row = ListReturnPlanProduction.objects.get(id=int(request_input["sqlindex"]))
            if select_row.return_keyin != "success":
                    return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})
            if select_row.return_approve:
                    return Response({"success": False, "message":"ข้อมูลนี้ถูก Approve ก่อนแล้ว"})

            select_row.return_approve = request_input["status"]
            select_row.note_planner = request_input["note_planner"]
            select_row.operator_approve = request.user.employee_id
            select_row.save()
            
            WIPSelect = getDetail_WIP(select_row.zca_on)

            select_pallet = ListReturnPalletProduction.objects.filter(returnplan_link=int(select_row.id))
            
            select_pallet_good = select_pallet.filter(return_type="good")
            total_good = 0
            date_good = ""
            shift_good = ""
            machine_good = ""
            if len(select_pallet_good) > 0:
                PlanMapSQL = MapListFillPlan(
                    work_type= "return",
                    machine = select_row.product_machine,
                    zca_on = select_row.zca_on,
                    name_th = select_row.name_th,
                    name_en = select_row.name_en,
                    product_type = select_row.product_type,
                    # product_date = select_row.product_date,
                    # product_shift = select_row.product_shift,
                    qty_good = select_row.qty_good,

                    receive_date = select_row.return_date,
                    receive_shift = select_row.return_shift,

                    pcsperpallet = WIPSelect["pcsperpallet"],
                    product_length = WIPSelect["field_lengthpallet"],
                    kgpcs = WIPSelect["field_kgpcs"],
                    plan_link = select_row.plan_link,
                )
                PlanMapSQL.save()
                
                for pallet in select_pallet_good:
                    MapSQL = MapListFillPallet(
                        work_type= "return",
                        machine = pallet.product_machine,
                        zca_on = pallet.zca_on,
                        name_th = pallet.name_th,
                        name_en = pallet.name_en,
                        product_type = pallet.product_type,
                        product_date = pallet.product_date,
                        product_shift = pallet.product_shift,
                        #ticket_type = "good",
                        qty = pallet.qty,
                        pallet_no = pallet.pallet_no,
                        receive_date = select_row.return_date,
                        receive_shift = select_row.return_shift,
                        plan_link = pallet.plan_link,
                        listfillplan_link = pallet.listfillplan_link,
                        maplistfillplan_link = PlanMapSQL,
                    )

                    total_good += int(pallet.qty)
                    date_good = pallet.product_date
                    shift_good = pallet.product_shift
                    machine_good = pallet.product_machine
                    MapSQL.pcsperpallet = WIPSelect["pcsperpallet"]
                    MapSQL.product_length = WIPSelect["field_lengthpallet"]
                    MapSQL.kgpcs = WIPSelect["field_kgpcs"]
                    MapSQL.save()

            select_pallet_bad = select_pallet.filter(return_type="bad")
            total_bad = 0
            date_bad = ""
            shift_bad = ""
            machine_bad = ""
            for pallet in select_pallet_bad:
                total_bad += pallet.qty
                date_bad = pallet.product_date
                shift_bad = pallet.product_shift
                machine_bad = pallet.product_machine

            
            # select_pallet_bad = select_pallet.filter(return_type="bad")

            # for pallet in select_pallet_bad:
            #     pallet_bad = ListLabBadUnlockPalletProduction(
            #         product_machine = select_row.product_machine,
            #         zca_on = select_row.zca_on,
            #         name_th = select_row.name_th,
            #         product_type = pallet.product_type,
            #         product_date = pallet.product_date,
            #         product_shift = pallet.product_shift,
            #         receive_date = select_row.return_date,
            #         receive_shift = select_row.return_shift,
                    
            #         qty = pallet.qty,
            #         pallet_no = pallet.pallet_no,
            #         plan_link = select_row.plan_link,
            #         returnplan_link = select_row,
            #         returnpallet_link = pallet,
            #     )
            #     pallet_bad.save()


            # SAP Tiger
            select_note = ListReturnPlanNoteProduction.objects.filter(returnplan_link=select_row).values()
            note_good = []
            note_bad = []
            for note in select_note:
                if note["type"] == "good":
                    note_good.append(note["message"])
                elif note["type"] == "bad":
                    note_bad.append(note["message"])
            
            if(total_good > 0):
                SAP_tiger = Tiger_GoodsReturn(
                    idmainfromwms = select_row.id,
                    usermachine = searchInfo_Operator(select_row.operator_keyin),
                    machine = select_row.return_machine,
                    datereturn = select_row.return_date,
                    shiftreturn = select_row.return_shift,
                    matno = select_row.zca_on,
                    matname = select_row.name_th,
                    qtyreturn = total_good,
                    pallet = divide_with_remainder(int(total_good), WIPSelect["pcsperpallet"]),
                    ton = float(total_good)*WIPSelect["field_kgpcs"]*0.001,
                    batch = str(int(date_good.year)%10) + str(date_good.strftime('%m')) + str(date_good.strftime('%d')) + machine_good + shift_good,
                    typereturn = "good",
                    notereturn = " ,".join(note_good),
                    approver = request.user.employee_id,
                    datetimewmssend = datetime.now(),
                )
                SAP_tiger.save()

            if(total_bad > 0):
                SAP_tiger = Tiger_GoodsReturn(
                    idmainfromwms = select_row.id,
                    usermachine = searchInfo_Operator(select_row.operator_keyin),
                    machine = select_row.return_machine,
                    datereturn = select_row.return_date,
                    shiftreturn = select_row.return_shift,
                    matno = select_row.zca_on,
                    matname = select_row.name_th,
                    qtyreturn = total_bad,
                    pallet = divide_with_remainder(int(total_bad), WIPSelect["pcsperpallet"]),
                    ton = float(total_bad)*WIPSelect["field_kgpcs"]*0.001,
                    batch = str(int(date_bad.year)%10) + str(date_bad.strftime('%m')) + str(date_bad.strftime('%d')) + machine_bad + shift_bad,
                    typereturn = "bad",
                    notereturn = " ,".join(note_bad),
                    approver = request.user.employee_id,
                    datetimewmssend = datetime.now(),
                )
                SAP_tiger.save()
                



            return Response({'success': True})
        # except Exception as e:
        #     print("ERROR >>>",e)
        #     return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class get_approveticket(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')

            mc = ListTicketPlanProduction.objects.filter(send_date=date_sql, send_shift=shift_sql).values_list('machine', flat=True).distinct()

            data = {}

            for mc in mc:
                if mc not in data.keys():
                    data[mc] = {}
                    data[mc]["items"] = []

                data[mc]["items"].extend(ListTicketPlanProduction.objects.filter(send_date=date_sql, send_shift=shift_sql).values())

            for i in data:
                for j in data[i]["items"]:
                    list_select_pallet = ListFillTicketPalletProduction.objects.filter(fillticketreturnplan_link=int(j["id"]),ticket_return_status__in=[1,0]).values()
                    j["pallet"] = list_select_pallet

                    j["operator_keyin_name"] = searchInfo_Operator(j["operator_keyin"])
                    j["operator_approve_name"] = searchInfo_Operator(j["operator_approve"])

            for i in data:
                list_all = 0
                list_wait = 0
                list_approve = 0
                for j in data[i]["items"]:
                    list_all += 1
                    if j["approve_fill"] == "success":
                        list_approve += 1
                    else:
                        list_wait += 1
                data[i]["list_all"] = list_all
                data[i]["list_wait"] = list_wait
                data[i]["list_approve"] = list_approve

            # DontsendData_mc = DontsendData.objects.filter(date=date_sql, shift=shift_sql).values_list('machine', flat=True).distinct()
            # dontsend = []
            # for mc in DontsendData_mc:
            #     dontsend.append(mc)

            return Response({'success': True, 'data': data})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class post_approveticket(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            request_input = request.data
            try:
                select_row = ListTicketPlanProduction.objects.get(id=int(request_input["id_select"]))
            except:
                return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})

            if select_row.fill_success != "success":
                    return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})
            if select_row.approve_fill:
                    return Response({"success": False, "message":"ข้อมูลนี้ถูก Approve ก่อนแล้ว"})

            select_row.approve_fill = "success"
            select_row.note_planner = request_input["note_planner"]
            select_row.operator_approve = request.user.employee_id
            
            select_row.save()

            if request_input["status"] == "success" and select_row.product_type == "WIP":
                PlanMapSQL = MapListFillPlan(
                    work_type= "ticket",
                    machine = select_row.machine,
                    zca_on = select_row.zca_on,
                    name_th = select_row.name_th,
                    name_en = select_row.name_en,
                    product_type = select_row.product_type,
                    product_date = select_row.product_date,
                    product_shift = select_row.product_shift,
                    qty_good = select_row.ticket_qty,
                    ticket_type = "good",

                    receive_date = select_row.send_date,
                    receive_shift = select_row.send_shift,

                    pcsperpallet = select_row.pcsperpallet,
                    product_length = select_row.product_length,
                    kgpcs = select_row.kgpcs,

                    plan_link = select_row.plan_link,
                    listfillplan_link = select_row.fillplan_link,
                    listticketplan_link = select_row,
                )
                PlanMapSQL.save()
                listpallet =  ListFillTicketPalletProduction.objects.filter(fillticketreturnplan_link=select_row,ticket_return_status=1)
                for pallet in listpallet:
                    MapSQL = MapListFillPallet(
                        work_type= "ticket",
                        machine = select_row.machine,
                        zca_on = select_row.zca_on,
                        name_th = select_row.name_th,
                        name_en = select_row.name_en,
                        product_type = select_row.product_type,
                        product_date = select_row.product_date,
                        product_shift = select_row.product_shift,
                        ticket_type = "good",
                        qty = pallet.qty_ticket,
                        pallet_no = pallet.pallet_no,
                        receive_date = select_row.send_date,
                        receive_shift = select_row.send_shift,

                        pcsperpallet = select_row.pcsperpallet,
                        product_length = select_row.product_length,
                        kgpcs = select_row.kgpcs,

                        plan_link = select_row.plan_link,
                        listfillplan_link = select_row.fillplan_link,
                        listticketplan_link = select_row,
                        maplistfillplan_link = PlanMapSQL
                    )
                    MapSQL.save()
                    
            return Response({'success': True})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class get_StatusFillTiger(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        date_filter = request.query_params.getlist('date_search[]')
        mc_filter = request.query_params.getlist('machine[]')
        # try:
        data = []
        
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        queryset = ViewWmsListfillplanproduction.objects
        local_timezone = pytz.timezone('Asia/Bangkok')  # Use your local timezone
        converted_dates = [datetime.fromisoformat(date.replace('Z', '+00:00')).astimezone(local_timezone) for date in date_filter]
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        if len(mc_filter) > 0:
            queryset = queryset.filter(machine__in=mc_filter)
        if len(converted_dates) == 2:
            queryset = queryset.filter(product_date__gte=converted_dates[0], product_date__lte=converted_dates[1])
        elif len(converted_dates) == 1:
            queryset = queryset.filter(product_date=converted_dates[0])

        queryset = queryset.order_by('-id').values()

        for i in queryset:
            i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
            i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])
            try:
                query_mapplan = MapListFillPlan.objects.filter(listfillplan_link=i["id"]).values()[0]
                query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=query_mapplan["id"]).values()
                if len(query_map) > 0:
                    i["approve_map"] = "success"
                else:
                    i["approve_map"] = None
            except:
                i["approve_map"] = None


        data = queryset

        grouped_data = {}

        for item in data:
            # สร้าง key สำหรับจัดกลุ่ม
            key = (item['zca_on'], item['product_date'], item['machine'])
            
            # ถ้า key นี้ยังไม่มีใน grouped_data ให้สร้างขึ้นมา
            if key not in grouped_data:
                grouped_data[key] = {
                    "zca_on": item['zca_on'],
                    "name_th": item['name_th'],
                    "product_date": item['product_date'],
                    "machine": item['machine'],
                    "total_qty": 0,
                    "total_loss": 0,
                    "total_lab": 0,
                    "list": [],
                }
            
            # เพิ่ม item นี้เข้าไปใน list ของกลุ่มนั้นๆ
            grouped_data[key]["list"].append(item)
            
            # เพิ่มค่า qty ลงใน total_qty ของกลุ่มนั้นๆ
            if item['qty_good'] != None :
                grouped_data[key]["total_qty"] += item['qty_good']
            if item['qty_loss'] != None :
                grouped_data[key]["total_loss"] += item['qty_loss']
            if item['qty_lab'] != None :
                grouped_data[key]["total_lab"] += item['qty_lab']


        result = list(grouped_data.values())

        tiger = []

        for i in result:
            for j in i["list"]:
                tiger.append(j)
            tiger.append({
                    "footer": True,
                    "zca_on": i['zca_on'],
                    "name_th": i['name_th'],
                    "product_date": i['product_date'],
                    "machine": i['machine'],
                    "qty_good": i["total_qty"],
                    "qty_loss": i["total_loss"],
                    "qty_lab": i["total_lab"],
                })


        return Response({'success': True, 'data': tiger,'result':result})

class get_StatusFill(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        date_filter = request.query_params.getlist('date_search[]')
        mc_filter = request.query_params.getlist('machine[]')
        # try:
        data = []
        queryset = ViewWmsListfillplanproduction.objects
        local_timezone = pytz.timezone('Asia/Bangkok')  # Use your local timezone
        try:
            converted_dates = [datetime.fromisoformat(date.replace('Z', '+00:00')).astimezone(local_timezone) for date in date_filter]
        except:
            converted_dates = [date.replace('Z', '+00:00') for date in date_filter]
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        if len(mc_filter) > 0:
            queryset = queryset.filter(machine__in=mc_filter)
        if len(converted_dates) == 2:
            queryset = queryset.filter(product_date__gte=converted_dates[0], product_date__lte=converted_dates[1])
        elif len(converted_dates) == 1:
            queryset = queryset.filter(product_date=converted_dates[0])
            
        queryset = queryset.order_by('-id').values()

        for i in queryset:
            i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
            i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])
            try:
                query_mapplan = MapListFillPlan.objects.filter(listfillplan_link=i["id"]).values()[0]
                query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=query_mapplan["id"]).values()
                if len(query_map) > 0:
                    i["approve_map"] = "success"
                else:
                    i["approve_map"] = None
            except:
                i["approve_map"] = None


        data = queryset
        return Response({'success': True, 'data': data})
    
class get_StatusWithdraw(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        date_filter = request.query_params.getlist('date_search[]')
        mc_filter = request.query_params.getlist('machine[]')
        # try:
        data = []
        
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        
        queryset = ListWithdrawPlanProduction.objects.exclude(delete_add="Delete")
        local_timezone = pytz.timezone('Asia/Bangkok')  # Use your local timezone
        converted_dates = [datetime.fromisoformat(date.replace('Z', '+00:00')).astimezone(local_timezone) for date in date_filter]
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        if len(mc_filter) > 0:
            queryset = queryset.filter(machine__in=mc_filter)
        if len(converted_dates) == 2:
            queryset = queryset.filter(receive_date__gte=converted_dates[0], receive_date__lte=converted_dates[1])
        elif len(converted_dates) == 1:
            queryset = queryset.filter(receive_date=converted_dates[0])

        queryset = queryset.order_by('-id').values()

        for i in queryset:
            i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
            i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])

            try:
                qty_send_query = list(MapListWithdrawPallet.objects.filter(listwithdraw_link=int(i["id"])).values("qty"))
                total_sum = 0
                for item_qty in qty_send_query:
                    total_sum += item_qty["qty"]
                i["qty_send"]= total_sum
            except:
                pass

            try:
                query_mapplan = MapListWithdrawPlan.objects.filter(listwithdraw_link=i["id"]).values()[0]
                query_map = ViewWmsMapManagement.objects.filter(maplistwithdrawplan_link=query_mapplan["id"]).values()
                if len(query_map) > 0:
                    i["approve_map"] = "success"
                else:
                    i["approve_map"] = None
            except:
                i["approve_map"] = None

        data = queryset
        return Response({'success': True, 'data': data})
    
class get_StatusReturn(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        date_filter = request.query_params.getlist('date_search[]')
        mc_filter = request.query_params.getlist('machine[]')

        # try:
        data = []
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        
        queryset = ListReturnPlanProduction.objects
        local_timezone = pytz.timezone('Asia/Bangkok')  # Use your local timezone
        converted_dates = [datetime.fromisoformat(date.replace('Z', '+00:00')).astimezone(local_timezone) for date in date_filter]
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        if len(mc_filter) > 0:
            queryset = queryset.filter(machine__in=mc_filter)
        if len(converted_dates) == 2:
            queryset = queryset.filter(return_date__gte=converted_dates[0], return_date__lte=converted_dates[1])
        elif len(converted_dates) == 1:
            queryset = queryset.filter(return_date=converted_dates[0])

        queryset = queryset.order_by('-id').values()

        for i in queryset:
            i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
            i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])


        data = queryset
        return Response({'success': True, 'data': data})
    
class get_StatusTicket(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        date_filter = request.query_params.getlist('date_search[]')
        mc_filter = request.query_params.getlist('machine[]')
        # try:
        data = []
        
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        queryset = ListTicketPlanProduction.objects
        local_timezone = pytz.timezone('Asia/Bangkok')  # Use your local timezone
        converted_dates = [datetime.fromisoformat(date.replace('Z', '+00:00')).astimezone(local_timezone) for date in date_filter]
        # queryset = ListGoodFillPlanProduction.objects.filter(approve_fill="success").order_by('-id').values()
        if len(mc_filter) > 0:
            queryset = queryset.filter(machine__in=mc_filter)
        if len(converted_dates) == 2:
            queryset = queryset.filter(product_date__gte=converted_dates[0], product_date__lte=converted_dates[1])
        elif len(converted_dates) == 1:
            queryset = queryset.filter(product_date=converted_dates[0])

        queryset = queryset.order_by('-id').values()

        for i in queryset:
            i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
            i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])

            try:
                query_mapplan = MapListFillPlan.objects.filter(listticketplan_link=i["id"]).values()[0]
                query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=query_mapplan["id"]).values()
                if len(query_map) > 0:
                    i["approve_map"] = "success"
                else:
                    i["approve_map"] = None
            except:
                i["approve_map"] = None

        data = queryset
        return Response({'success': True, 'data': data})

class get_Dashboard(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        # try:

            date_filter = request.query_params.getlist('date_search[]')
            # try:
            data = {
                        "filllist_stat":{"success":0,"remain":0},
                        "filllist":[],
                        "withdrawlist_stat":{"success":0,"remain":0},
                        "withdrawlist":[],
                        "returnlist_stat":{"success":0,"remain":0},
                        "returnlist":[],
                        "ticketlist_stat":{"success":0,"remain":0},
                        "ticketlist":[],

                        "labwithdrawlist_stat":{"success":0,"remain":0},
                        "labwithdrawlist":[],
                        "labreturnlist_stat":{"success":0,"remain":0},
                        "labreturnlist":[],
                        "labunlockbadlist_stat":{"success":0,"remain":0},
                        "labunlockbadlist":[],

                        "map_filllist_stat":{"success":0,"remain":0},
                        "map_filllist":[],
                        "map_withdrawlist_stat":{"success":0,"remain":0},
                        "map_withdrawlist":[],
                        "map_returnlist_stat":{"success":0,"remain":0},
                        "map_returnlist":[],
                        "map_ticketlist_stat":{"success":0,"remain":0},
                        "map_ticketlist":[],

                        "map_labwithdrawlist_stat":{"success":0,"remain":0},
                        "map_labwithdrawlist":[],
                        "map_labreturnlist_stat":{"success":0,"remain":0},
                        "map_labreturnlist":[],
                        "map_labunlockbadlist_stat":{"success":0,"remain":0},
                        "map_labunlockbadlist":[],
                    }
            
            queryset = ViewWmsListfillplanproduction.objects.filter(fill_success="success")
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            filltemp = []
            for row in queryset:
                if row["approve_fill"] != "success":
                    data["filllist_stat"]["remain"] += 1
                    filltemp.append(row)
                else:
                    data["filllist_stat"]["success"] += 1


            grouped_data = defaultdict(lambda: defaultdict(list))
            # Iterate over each item in the data list
            for item in filltemp:
                # Use the 'receive_date' and 'receive_shift' to group the items
                grouped_data[item['send_date']][item['send_shift']].append(item)

            # Now, transform the grouped data into the desired list format
            grouped_list = []
            for date, shifts in grouped_data.items():
                for shift, lists in shifts.items():
                    grouped_list.append({
                        "date": date,
                        "shift": shift,
                        "total_lists": len(lists),
                        "lists": lists,
                    })
            data["filllist"] = grouped_list

            queryset = ListWithdrawPlanProduction.objects.filter(withdraw_keyin="success").exclude(machine="Lab")
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            withdrawtemp = []
            for row in queryset:
                if row["approve_withdraw"] != "success":
                    data["withdrawlist_stat"]["remain"] += 1
                    withdrawtemp.append(row)
                else:
                    data["withdrawlist_stat"]["success"] += 1
            
            grouped_data = defaultdict(lambda: defaultdict(list))
            # Iterate over each item in the data list
            for item in withdrawtemp:
                # Use the 'receive_date' and 'receive_shift' to group the items
                grouped_data[item['receive_date']][item['receive_shift']].append(item)

            # Now, transform the grouped data into the desired list format
            grouped_list = []
            for date, shifts in grouped_data.items():
                for shift, lists in shifts.items():
                    grouped_list.append({
                        "date": date,
                        "shift": shift,
                        "total_lists": len(lists),
                        "lists": lists,
                    })
            data["withdrawlist"] = grouped_list

            queryset = ListReturnPlanProduction.objects.filter(return_keyin="success")
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            returntemp = []
            for row in queryset:
                if row["return_approve"] != "success":
                    data["returnlist_stat"]["remain"] += 1
                    returntemp.append(row)
                else:
                    data["returnlist_stat"]["success"] += 1

            grouped_data = defaultdict(lambda: defaultdict(list))
            # Iterate over each item in the data list
            for item in returntemp:
                # Use the 'receive_date' and 'receive_shift' to group the items
                grouped_data[item['return_date']][item['return_shift']].append(item)

            # Now, transform the grouped data into the desired list format
            grouped_list = []
            for date, shifts in grouped_data.items():
                for shift, lists in shifts.items():
                    grouped_list.append({
                        "date": date,
                        "shift": shift,
                        "total_lists": len(lists),
                        "lists": lists,
                    })
            data["returnlist"] = grouped_list



            queryset = ListTicketPlanProduction.objects.filter(fill_success="success")
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            tickettemp = []
            for row in queryset:
                if row["approve_fill"] != "success":
                    data["ticketlist_stat"]["remain"] += 1
                    tickettemp.append(row)
                else:
                    data["ticketlist_stat"]["success"] += 1

            grouped_data = defaultdict(lambda: defaultdict(list))
            # Iterate over each item in the data list
            for item in tickettemp:
                # Use the 'receive_date' and 'receive_shift' to group the items
                grouped_data[item['send_date']][item['send_shift']].append(item)

            # Now, transform the grouped data into the desired list format
            grouped_list = []
            for date, shifts in grouped_data.items():
                for shift, lists in shifts.items():
                    grouped_list.append({
                        "date": date,
                        "shift": shift,
                        "total_lists": len(lists),
                        "lists": lists,
                    })
            data["ticketlist"] = grouped_list


            # LAB Withdraw
            queryset = ListWithdrawPlanProduction.objects.filter(machine="Lab",withdraw_keyin="success")
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            withdrawtemp = []
            for row in queryset:
                if row["approve_withdraw"] != "success":
                    data["labwithdrawlist_stat"]["remain"] += 1
                    withdrawtemp.append(row)
                else:
                    data["labwithdrawlist_stat"]["success"] += 1
            
            grouped_data = defaultdict(lambda: defaultdict(list))
            # Iterate over each item in the data list
            for item in withdrawtemp:
                # Use the 'receive_date' and 'receive_shift' to group the items
                grouped_data[item['receive_date']][item['receive_shift']].append(item)

            # Now, transform the grouped data into the desired list format
            grouped_list = []
            for date, shifts in grouped_data.items():
                for shift, lists in shifts.items():
                    grouped_list.append({
                        "date": date,
                        "shift": shift,
                        "total_lists": len(lists),
                        "lists": lists,
                    })
            data["labwithdrawlist"] = grouped_list

            #LAB Return
            queryset = ListLabReturnPlanProduction.objects.filter(return_keyin="success")
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            returntemp = []
            for row in queryset:
                if row["return_approve"] != "success":
                    data["labreturnlist_stat"]["remain"] += 1
                    returntemp.append(row)
                else:
                    data["labreturnlist_stat"]["success"] += 1

            grouped_data = defaultdict(lambda: defaultdict(list))
            # Iterate over each item in the data list
            for item in returntemp:
                # Use the 'receive_date' and 'receive_shift' to group the items
                grouped_data[item['return_date']][item['return_shift']].append(item)

            # Now, transform the grouped data into the desired list format
            grouped_list = []
            for date, shifts in grouped_data.items():
                for shift, lists in shifts.items():
                    grouped_list.append({
                        "date": date,
                        "shift": shift,
                        "total_lists": len(lists),
                        "lists": lists,
                    })
            data["labreturnlist"] = grouped_list


            #LAB Unlock Bad
            queryset = ListLabBadUnlockPlanProduction.objects.filter(return_keyin="success")
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            returntemp = []
            for row in queryset:
                if row["return_approve"] != "success":
                    data["labunlockbadlist_stat"]["remain"] += 1
                    returntemp.append(row)
                else:
                    data["labunlockbadlist_stat"]["success"] += 1

            grouped_data = defaultdict(lambda: defaultdict(list))
            # Iterate over each item in the data list
            for item in returntemp:
                # Use the 'receive_date' and 'receive_shift' to group the items
                grouped_data[item['return_date']][item['return_shift']].append(item)

            # Now, transform the grouped data into the desired list format
            grouped_list = []
            for date, shifts in grouped_data.items():
                for shift, lists in shifts.items():
                    grouped_list.append({
                        "date": date,
                        "shift": shift,
                        "total_lists": len(lists),
                        "lists": lists,
                    })
            data["labunlockbadlist"] = grouped_list



            queryset = MapListFillPlan.objects
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            map_filltemp = []
            map_returntemp = []
            map_tickettemp = []
            map_labreturntemp = []
            map_unlockbadtemp = []
            for row in queryset:
                if row["work_type"] == "fill":
                    try:
                        query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=row["id"]).values()
                        if len(query_map) > 0:
                            data["map_filllist_stat"]["success"] += 1
                            break
                        else:
                            data["map_filllist_stat"]["remain"] += 1
                    except:
                        data["map_filllist_stat"]["remain"] += 1

                    # data["map_filllist_stat"]["remain"] += 1
                    map_filltemp.append(row)
                elif row["work_type"] == "return":
                    try:
                        query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=row["id"]).values()
                        if len(query_map) > 0:
                            data["map_returnlist_stat"]["success"] += 1
                            break
                        else:
                            data["map_returnlist_stat"]["remain"] += 1
                    except:
                        data["map_returnlist_stat"]["remain"] += 1
                    # data["map_returnlist_stat"]["remain"] += 1
                    map_returntemp.append(row)
                elif row["work_type"] == "ticket":
                    try:
                        query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=row["id"]).values()
                        if len(query_map) > 0:
                            data["map_ticketlist_stat"]["success"] += 1
                            break
                        else:
                            data["map_ticketlist_stat"]["remain"] += 1
                    except:
                        data["map_ticketlist_stat"]["remain"] += 1
                    # data["map_ticketlist_stat"]["remain"] += 1
                    map_tickettemp.append(row)
                elif row["work_type"] == "labreturn":
                    try:
                        query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=row["id"]).values()
                        if len(query_map) > 0:
                            data["map_labreturnlist_stat"]["success"] += 1
                            break
                        else:
                            data["map_labreturnlist_stat"]["remain"] += 1
                    except:
                        data["map_labreturnlist_stat"]["remain"] += 1
                    # data["map_ticketlist_stat"]["remain"] += 1
                    map_labreturntemp.append(row)
                elif row["work_type"] == "badfill":
                    try:
                        query_map = ViewWmsMapManagement.objects.filter(maplistfillplan_link=row["id"]).values()
                        if len(query_map) > 0:
                            data["map_labunlockbadlist_stat"]["success"] += 1
                            break
                        else:
                            data["map_labunlockbadlist_stat"]["remain"] += 1
                    except:
                        data["map_ticketlist_stat"]["remain"] += 1
                    # data["map_ticketlist_stat"]["remain"] += 1x
                    map_unlockbadtemp.append(row)
            def groupData(data_input,field_date,field_shift):
                grouped_data = defaultdict(lambda: defaultdict(list))
                # Iterate over each item in the data list
                for item in data_input:
                    # Use the 'receive_date' and 'receive_shift' to group the items
                    grouped_data[item[field_date]][item[field_shift]].append(item)

                # Now, transform the grouped data into the desired list format
                grouped_list = []
                for date, shifts in grouped_data.items():
                    for shift, lists in shifts.items():
                        grouped_list.append({
                            "date": date,
                            "shift": shift,
                            "total_lists": len(lists),
                            "lists": lists,
                        })
                return grouped_list

            data["map_filllist"] = groupData(map_filltemp,"receive_date","receive_shift")
            data["map_returnlist"] = groupData(map_returntemp,"receive_date","receive_shift")
            data["map_ticketlist"] = groupData(map_tickettemp,"receive_date","receive_shift")
            data["map_labreturnlist"] = groupData(map_labreturntemp,"receive_date","receive_shift")
            data["map_labunlockbadlist"] = groupData(map_unlockbadtemp,"receive_date","receive_shift")


            #
            queryset = MapListWithdrawPlan.objects.filter(withdraw_success=False)
            if len(date_filter) == 2:
                queryset = queryset.filter(created_at__gte=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')), created_at__lte=datetime.fromisoformat(date_filter[1].replace('Z', '+00:00')))
            elif len(date_filter) == 1:
                queryset = queryset.filter(created_at=datetime.fromisoformat(date_filter[0].replace('Z', '+00:00')))

            queryset = queryset.order_by('-id').values()
            map_withdrawtemp = []
            map_labwithdrawtemp = []

            for row in queryset:
                if row["machine"] == "Lab":
                    data["map_labwithdrawlist_stat"]["remain"] += 1
                    map_labwithdrawtemp.append(row)
                else:
                    data["map_withdrawlist_stat"]["remain"] += 1
                    map_withdrawtemp.append(row)

            def groupData(data_input,field_date,field_shift):
                grouped_data = defaultdict(lambda: defaultdict(list))
                # Iterate over each item in the data list
                for item in data_input:
                    # Use the 'receive_date' and 'receive_shift' to group the items
                    grouped_data[item[field_date]][item[field_shift]].append(item)

                # Now, transform the grouped data into the desired list format
                grouped_list = []
                for date, shifts in grouped_data.items():
                    for shift, lists in shifts.items():
                        grouped_list.append({
                            "date": date,
                            "shift": shift,
                            "total_lists": len(lists),
                            "lists": lists,
                        })
                return grouped_list

            data["map_withdrawlist"] = groupData(map_withdrawtemp,"send_date","send_shift")
            data["map_labwithdrawlist"] = groupData(map_labwithdrawtemp,"send_date","send_shift")
            

            
            return Response({'success': True, 'data': data})
    
class get_approvelabwithdraw(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')

            mc = ListWithdrawPlanProduction.objects.filter(withdraw_keyin="success",machine="Lab", receive_date=date_sql, receive_shift=shift_sql).values_list('machine', flat=True).distinct()

            data = {}

            for mc in mc:
                if mc not in data.keys():
                    data[mc] = {}
                    data[mc]["items"] = []

                data[mc]["items"].extend(ListWithdrawPlanProduction.objects.filter(machine=mc,withdraw_keyin="success", receive_date=date_sql, receive_shift=shift_sql).values())

            for i in data:
                for j in data[i]["items"]:
                    try:
                        query_search = ViewItemmasterproductwip.objects.filter(field_zca=j["zca_on"]).values('field_zca','field_name','field_nameeng','field_mc','pcsperpallet')[0]
                        j["pcperpallet"] = int(query_search['pcsperpallet'])
                    except:
                        query_search = ViewItemmasterproductfg.objects.filter(zca=j["zca_on"]).values('zca','name','nameen','pcpallet')[0]
                        j["pcperpallet"] = int(query_search['pcpallet'])

                    try:
                        select_balance = Tiger_StockBalance.objects.get(zca=j["zca_on"])
                        j["instock_good"] = select_balance.urstock
                        j["instock_lab"] = select_balance.block
                    except Exception as e:
                        print("Error>>>",e)
                        j["instock_good"] = 0
                        j["instock_lab"] = 0
                    j["operator_keyin_name"] = searchInfo_Operator(j["operator_keyin"])
                    j["operator_approve_name"] = searchInfo_Operator(j["operator_approve"])
                        
            for i in data:
                list_all = 0
                list_wait = 0
                list_approve = 0
                for j in data[i]["items"]:
                    list_all += 1
                    if j["approve_withdraw"] == "success":
                        list_approve += 1
                    else:
                        list_wait += 1
                data[i]["list_all"] = list_all
                data[i]["list_wait"] = list_wait
                data[i]["list_approve"] = list_approve

            DontsendData_mc = DontsendData.objects.filter(date=date_sql, shift=shift_sql).values_list('machine', flat=True).distinct()
            dontsend = []
            for mc in DontsendData_mc:
                dontsend.append(mc)

            return Response({'success': True, 'data': data,'dontsend':dontsend})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class post_approvelabwithdraw(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            request_input = request.data
            select_row = ListWithdrawPlanProduction.objects.get(id=int(request_input["sqlindex"]))
            if select_row.withdraw_keyin != "success":
                    return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})
            if select_row.approve_withdraw:
                    return Response({"success": False, "message":"ข้อมูลนี้ถูก Approve ก่อนแล้ว"})
            select_row.approve_withdraw = "success"
            # select_row.qtysend = request_input["qtysend"]
            # select_row.note_planner = request_input["note_planner"]
            select_row.operator_approve = request.user.employee_id

            WIPSelect = getDetail_WIP(select_row.zca_on)

            def divide_with_remainder(A, B):
                quotient = A // B
                remainder = A % B
                result = f"{quotient}({remainder})"
                return result
            
            def get_totalpallet(A, B):
                quotient = A // B
                if int(A % B) > 0:
                    quotient += 1
                return quotient

            MapSQL = MapListWithdrawPlan(
                machine = select_row.machine,
                zca_on = select_row.zca_on,
                name_th = select_row.name_th,
                name_en = select_row.name_en,
                product_type = "WIP",
                qty = select_row.qty,

                qty_total_pallet = get_totalpallet(int(select_row.qty),  int(WIPSelect["pcsperpallet"])),
                qty_format = divide_with_remainder(int(select_row.qty),  int(WIPSelect["pcsperpallet"])),
                
                pcsperpallet = WIPSelect["pcsperpallet"],
                product_length = WIPSelect["field_lengthpallet"],
                kgpcs = WIPSelect["field_kgpcs"],

                send_date = select_row.receive_date,
                send_shift = select_row.receive_shift,
                listwithdraw_link = select_row,
                plan_link = select_row.plan_link
            )
            MapSQL.save()

            select_row.save()



            return Response({'success': True})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class get_approvelabreturn(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')

            mc = ListLabReturnPlanProduction.objects.filter(return_date=date_sql, return_shift=shift_sql).values_list('product_machine', flat=True).distinct()

            data = {}

            for mc in mc:
                if mc not in data.keys():
                    data[mc] = {}
                    data[mc]["items"] = []

                data[mc]["items"].extend(ListLabReturnPlanProduction.objects.filter(return_date=date_sql, return_shift=shift_sql).values())

            for i in data:
                for j in data[i]["items"]:
                    list_select_pallet = ListLabReturnPalletProduction.objects.filter(labreturnplan_link=int(j["id"])).values("id","zca_on","product_date","product_shift","pallet_no","qty","pcsperpallet","return_type")
                    j["operator_keyin_name"] = searchInfo_Operator(j["operator_keyin"])
                    j["operator_approve_name"] = searchInfo_Operator(j["operator_approve"])
                    # Initialize a dictionary to hold the grouped data
                    grouped_data = defaultdict(lambda: defaultdict(list))

                    # Iterate over each item in the data list
                    for item in list_select_pallet:
                        # Use the 'receive_date' and 'receive_shift' to group the items
                        grouped_data[item['product_date']][item['product_shift']].append(item)

                    # Now, transform the grouped data into the desired list format
                    grouped_list = []
                    for date, shifts in grouped_data.items():
                        for shift, pallets in shifts.items():
                            grouped_list.append({
                                "product_date": date,
                                "product_shift": shift,
                                "total_pallets": len(pallets),
                                "pallets": pallets,
                            })

                    list_good = []
                    list_bad = []
                    for k in list_select_pallet:
                        if k["return_type"] == 1:
                            list_good.append(k)
                        elif k["return_type"] == 2:
                            list_bad.append(k)
                    query_search = ViewItemmasterproductwip.objects.filter(field_zca=j["zca_on"]).values('field_zca','field_name','field_nameeng','field_mc','pcsperpallet')[0]
                    j["pcperpallet"] = int(query_search['pcsperpallet'])
                    
                    j["total-good"] = len(list_good)
                    j["total-bad"] = len(list_bad)

                    j["pallet_list"] = grouped_list

                    j["note_good"] = []
                    j["note_bad"] = []

            for i in data:
                list_all = 0
                list_wait = 0
                list_approve = 0
                for j in data[i]["items"]:
                    list_all += 1
                    if j["return_approve"] == "success":
                        list_approve += 1
                    else:
                        list_wait += 1
                data[i]["list_all"] = list_all
                data[i]["list_wait"] = list_wait
                data[i]["list_approve"] = list_approve

            # DontsendData_mc = DontsendData.objects.filter(date=date_sql, shift=shift_sql).values_list('machine', flat=True).distinct()
            # dontsend = []
            # for mc in DontsendData_mc:
            #     dontsend.append(mc)

            return Response({'success': True, 'data': data})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class post_approvelabreturn(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            request_input = request.data
            try:
                select_row = ListLabReturnPlanProduction.objects.get(id=int(request_input["sqlindex"]))
            except:
                return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})

            if select_row.return_keyin != "success":
                    return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})
            if select_row.return_approve:
                    return Response({"success": False, "message":"ข้อมูลนี้ถูก Approve ก่อนแล้ว"})

            # select_row.approve_fill = "success"
            # select_row.note_planner = request_input["note_planner"]
            # select_row.operator_approve = request.user.employee_id
            
            # select_row.save()

            # if request_input["status"] == "success":
            #     PlanMapSQL = MapListFillPlan(
            #         work_type= "labreturn",
            #         machine = select_row.machine,
            #         zca_on = select_row.zca_on,
            #         name_th = select_row.name_th,
            #         name_en = select_row.name_en,
            #         product_type = select_row.product_type,
            #         product_date = select_row.product_date,
            #         product_shift = select_row.product_shift,
            #         qty_good = select_row.ticket_qty,
            #         ticket_type = "good",

            #         receive_date = select_row.send_date,
            #         receive_shift = select_row.send_shift,

            #         pcsperpallet = select_row.pcsperpallet,
            #         product_length = select_row.product_length,
            #         kgpcs = select_row.kgpcs,

            #         plan_link = select_row.plan_link,
            #         listfillplan_link = select_row.fillplan_link,
            #         listticketplan_link = select_row,
            #     )
            #     PlanMapSQL.save()
            #     listpallet =  ListLabReturnPalletProduction.objects.filter(fillticketreturnplan_link=select_row,ticket_return_status=1)
            #     for pallet in listpallet:
            #         MapSQL = MapListFillPallet(
            #             work_type= "labreturn",
            #             machine = select_row.machine,
            #             zca_on = select_row.zca_on,
            #             name_th = select_row.name_th,
            #             name_en = select_row.name_en,
            #             product_type = select_row.product_type,
            #             product_date = select_row.product_date,
            #             product_shift = select_row.product_shift,
            #             ticket_type = "good",
            #             qty = pallet.qty_ticket,
            #             pallet_no = pallet.pallet_no,
            #             receive_date = select_row.send_date,
            #             receive_shift = select_row.send_shift,

            #             pcsperpallet = select_row.pcsperpallet,
            #             product_length = select_row.product_length,
            #             kgpcs = select_row.kgpcs,

            #             plan_link = select_row.plan_link,
            #             listfillplan_link = select_row.fillplan_link,
            #             listticketplan_link = select_row,
            #             maplistfillplan_link = PlanMapSQL
            #         )
            #         MapSQL.save()

            select_row.return_approve = request_input["status"]
            select_row.note_planner = request_input["note_planner"]
            select_row.operator_approve = request.user.employee_id
            select_row.save()
            
            WIPSelect = getDetail_WIP(select_row.zca_on)

            select_pallet = ListLabReturnPalletProduction.objects.filter(labreturnplan_link=int(select_row.id))
            select_pallet_good = select_pallet.filter(return_type="good")

            if len(select_pallet_good) > 0:
                PlanMapSQL = MapListFillPlan(
                    work_type= "labreturn",
                    machine = select_row.product_machine,
                    zca_on = select_row.zca_on,
                    name_th = select_row.name_th,
                    name_en = select_row.name_en,
                    product_type = select_row.product_type,
                    # product_date = select_row.product_date,
                    # product_shift = select_row.product_shift,
                    qty_good = select_row.qty_good,

                    receive_date = select_row.return_date,
                    receive_shift = select_row.return_shift,

                    pcsperpallet = WIPSelect["pcsperpallet"],
                    product_length = WIPSelect["field_lengthpallet"],
                    kgpcs = WIPSelect["field_kgpcs"],
                    plan_link = select_row.plan_link,
                )
                PlanMapSQL.save()

            for pallet in select_pallet_good:
                MapSQL = MapListFillPallet(
                    work_type= "labreturn",
                    machine = pallet.product_machine,
                    zca_on = pallet.zca_on,
                    name_th = pallet.name_th,
                    name_en = pallet.name_en,
                    product_type = pallet.product_type,
                    product_date = pallet.product_date,
                    product_shift = pallet.product_shift,
                    #ticket_type = "good",
                    qty = pallet.qty,
                    pallet_no = pallet.pallet_no,
                    receive_date = select_row.return_date,
                    receive_shift = select_row.return_shift,
                    plan_link = pallet.plan_link,
                    listfillplan_link = pallet.listfillplan_link,
                    maplistfillplan_link = PlanMapSQL,
                )

                MapSQL.pcsperpallet = WIPSelect["pcsperpallet"]
                MapSQL.product_length = WIPSelect["field_lengthpallet"]
                MapSQL.kgpcs = WIPSelect["field_kgpcs"]
                MapSQL.save()

                pallet.maplistfillpallet_link = MapSQL
                pallet.save()

            select_row.maplistfillplan_link = PlanMapSQL
            select_row.save()
                    
            return Response({'success': True})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class get_approvelabunlockbad(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')

            mc = ListLabBadUnlockPlanProduction.objects.filter(return_date=date_sql, return_shift=shift_sql).values_list('product_machine', flat=True).distinct()

            data = {}

            for mc in mc:
                if mc not in data.keys():
                    data[mc] = {}
                    data[mc]["items"] = []

                data[mc]["items"].extend(ListLabBadUnlockPlanProduction.objects.filter(return_date=date_sql, return_shift=shift_sql).values())

            for i in data:
                for j in data[i]["items"]:
                    list_select_pallet = ListLabBadUnlockPalletProduction.objects.filter(labunlockbadplan_link=int(j["id"])).values("id","zca_on","product_date","product_shift","pallet_no","qty","pcsperpallet","return_type")
                    j["operator_keyin_name"] = searchInfo_Operator(j["operator_keyin"])
                    j["operator_approve_name"] = searchInfo_Operator(j["operator_approve"])
                    # Initialize a dictionary to hold the grouped data
                    grouped_data = defaultdict(lambda: defaultdict(list))

                    # Iterate over each item in the data list
                    for item in list_select_pallet:
                        # Use the 'receive_date' and 'receive_shift' to group the items
                        grouped_data[item['product_date']][item['product_shift']].append(item)

                    # Now, transform the grouped data into the desired list format
                    grouped_list = []
                    for date, shifts in grouped_data.items():
                        for shift, pallets in shifts.items():
                            grouped_list.append({
                                "product_date": date,
                                "product_shift": shift,
                                "total_pallets": len(pallets),
                                "pallets": pallets,
                            })

                    list_good = []
                    list_bad = []
                    for k in list_select_pallet:
                        if k["return_type"] == 1:
                            list_good.append(k)
                        elif k["return_type"] == 2:
                            list_bad.append(k)
                    query_search = ViewItemmasterproductwip.objects.filter(field_zca=j["zca_on"]).values('field_zca','field_name','field_nameeng','field_mc','pcsperpallet')[0]
                    j["pcperpallet"] = int(query_search['pcsperpallet'])
                    
                    j["total-good"] = len(list_good)
                    j["total-bad"] = len(list_bad)

                    j["pallet_list"] = grouped_list

                    j["note_good"] = []
                    j["note_bad"] = []

            for i in data:
                list_all = 0
                list_wait = 0
                list_approve = 0
                for j in data[i]["items"]:
                    list_all += 1
                    if j["return_approve"] == "success":
                        list_approve += 1
                    else:
                        list_wait += 1
                data[i]["list_all"] = list_all
                data[i]["list_wait"] = list_wait
                data[i]["list_approve"] = list_approve

            # DontsendData_mc = DontsendData.objects.filter(date=date_sql, shift=shift_sql).values_list('machine', flat=True).distinct()
            # dontsend = []
            # for mc in DontsendData_mc:
            #     dontsend.append(mc)

            return Response({'success': True, 'data': data})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class post_approvelabunlockbad(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            request_input = request.data
            try:
                select_row = ListLabBadUnlockPlanProduction.objects.get(id=int(request_input["sqlindex"]))
            except:
                return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})

            if select_row.return_keyin != "success":
                    return Response({"success": False, "message":"ข้อมูลนี้มีการแก้ไขจากผลิต"})
            if select_row.return_approve:
                    return Response({"success": False, "message":"ข้อมูลนี้ถูก Approve ก่อนแล้ว"})

            # select_row.approve_fill = "success"
            # select_row.note_planner = request_input["note_planner"]
            # select_row.operator_approve = request.user.employee_id
            
            # select_row.save()

            # if request_input["status"] == "success":
            #     PlanMapSQL = MapListFillPlan(
            #         work_type= "badfill",
            #         machine = select_row.machine,
            #         zca_on = select_row.zca_on,
            #         name_th = select_row.name_th,
            #         name_en = select_row.name_en,
            #         product_type = select_row.product_type,
            #         product_date = select_row.product_date,
            #         product_shift = select_row.product_shift,
            #         qty_good = select_row.ticket_qty,
            #         ticket_type = "good",

            #         receive_date = select_row.send_date,
            #         receive_shift = select_row.send_shift,

            #         pcsperpallet = select_row.pcsperpallet,
            #         product_length = select_row.product_length,
            #         kgpcs = select_row.kgpcs,

            #         plan_link = select_row.plan_link,
            #         listfillplan_link = select_row.fillplan_link,
            #         listticketplan_link = select_row,
            #     )
            #     PlanMapSQL.save()
            #     listpallet =  ListLabBadUnlockPalletProduction.objects.filter(fillticketreturnplan_link=select_row,ticket_return_status=1)
            #     for pallet in listpallet:
            #         MapSQL = MapListFillPallet(
            #             work_type= "badfill",
            #             machine = select_row.machine,
            #             zca_on = select_row.zca_on,
            #             name_th = select_row.name_th,
            #             name_en = select_row.name_en,
            #             product_type = select_row.product_type,
            #             product_date = select_row.product_date,
            #             product_shift = select_row.product_shift,
            #             ticket_type = "good",
            #             qty = pallet.qty_ticket,
            #             pallet_no = pallet.pallet_no,
            #             receive_date = select_row.send_date,
            #             receive_shift = select_row.send_shift,

            #             pcsperpallet = select_row.pcsperpallet,
            #             product_length = select_row.product_length,
            #             kgpcs = select_row.kgpcs,

            #             plan_link = select_row.plan_link,
            #             listfillplan_link = select_row.fillplan_link,
            #             listticketplan_link = select_row,
            #             maplistfillplan_link = PlanMapSQL
            #         )
            #         MapSQL.save()
            select_row.return_approve = request_input["status"]
            select_row.note_planner = request_input["note_planner"]
            select_row.operator_approve = request.user.employee_id
            select_row.save()
            
            WIPSelect = getDetail_WIP(select_row.zca_on)

            select_pallet = ListLabBadUnlockPalletProduction.objects.filter(labunlockbadplan_link=int(select_row.id))
            select_pallet_good = select_pallet.filter(return_type="good")

            if len(select_pallet_good) > 0:
                PlanMapSQL = MapListFillPlan(
                    work_type= "badfill",
                    machine = select_row.product_machine,
                    zca_on = select_row.zca_on,
                    name_th = select_row.name_th,
                    name_en = select_row.name_en,
                    product_type = select_row.product_type,
                    # product_date = select_row.product_date,
                    # product_shift = select_row.product_shift,
                    qty_good = select_row.qty_good,

                    receive_date = select_row.return_date,
                    receive_shift = select_row.return_shift,

                    pcsperpallet = WIPSelect["pcsperpallet"],
                    product_length = WIPSelect["field_lengthpallet"],
                    kgpcs = WIPSelect["field_kgpcs"],
                    plan_link = select_row.plan_link,
                )
                PlanMapSQL.save()

                select_row.maplistfillplan_link = PlanMapSQL
                select_row.save()

            for pallet in select_pallet_good:
                MapSQL = MapListFillPallet(
                    work_type= "badfill",
                    machine = pallet.product_machine,
                    zca_on = pallet.zca_on,
                    name_th = pallet.name_th,
                    name_en = pallet.name_en,
                    product_type = pallet.product_type,
                    product_date = pallet.product_date,
                    product_shift = pallet.product_shift,
                    #ticket_type = "good",
                    qty = pallet.qty,
                    pallet_no = pallet.pallet_no,
                    receive_date = select_row.return_date,
                    receive_shift = select_row.return_shift,
                    plan_link = pallet.plan_link,
                    maplistfillplan_link = PlanMapSQL,
                )

                MapSQL.pcsperpallet = WIPSelect["pcsperpallet"]
                MapSQL.product_length = WIPSelect["field_lengthpallet"]
                MapSQL.kgpcs = WIPSelect["field_kgpcs"]
                MapSQL.save()
                    
            return Response({'success': True})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class get_StockBalance(APIView):
    permission_classes = [PlannerPerm | AdminPerm | ManagerPerm | PISPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        # try:
            group_filter = request.query_params.get('group')
            subgroup_filter = request.query_params.get('subgroup')
            latest_mapid_subquery = ViewWmsMapManagement.objects.filter(
                mapid=OuterRef('mapid'), 
                level=OuterRef('level'), 
                sub_column=OuterRef('sub_column')
            ).exclude(map_approve = 2).order_by('-created_at').values('created_at')[:1]
                        
            queryset = ViewWmsMapManagement.objects.annotate(latest_created_at = Subquery(latest_mapid_subquery)).filter(map_approve=1 ,created_at=F('latest_created_at')  ).values("zca_on","name_th","machine","qty","product_length","map_approve")

            for i in queryset:
                query_selct = getDetail_WIP(i["zca_on"])
                i["group"] = query_selct["field_prodgroup"]
                i["subgroup"] = query_selct["field_prodname"]
                i["kgpcs"] = query_selct["field_kgpcs"]

            result = {}
            for item in queryset:
                group = item['group']
                subgroup = item['subgroup']
                qty_kg = float(item['qty']) * float(item['kgpcs'])

                # รวมค่าน้ำหนักสำหรับแต่ละกลุ่มและย่อยกลุ่ม
                if group not in result:
                    result[group] = {}

                if subgroup not in result[group]:
                    result[group][subgroup] = 0

                result[group][subgroup] += (qty_kg)

            # จัดรูปแบบผลลัพธ์ตามที่ต้องการ
            # สร้าง dictionary สำหรับการเก็บข้อมูล
            group_data = {}

            for group, subgroups in result.items():
                for subgroup, qty_kg in subgroups.items():
                    if group not in group_data:
                        group_data[group] = {"name": group, "data": []}
                    
                    group_data[group]["data"].append({"x": subgroup, "y": round(qty_kg / 1000, 2)})

            # แปลง dictionary เป็น list
            formatted_result = list(group_data.values())

            # Step 2: Initialize a dictionary for counts and weights
            counts_and_weights = {}

            # Step 3: Iterate and calculate
            for item in queryset:
                key = (item['zca_on'], item["name_th"], item["group"], item["subgroup"], item["product_length"])
                
                qty = float(item['qty'])
                weight_kg = qty * float(item['kgpcs'])  # Calculate the weight for this item

                ready_qty = 0
                ready_weight_kg = 0
                if item['map_approve'] == 1:
                    ready_qty += qty
                    ready_weight_kg += qty * float(item['kgpcs'])

                if key in counts_and_weights:
                    counts_and_weights[key]['total_qty'] += qty
                    counts_and_weights[key]['total_weight_kg'] += weight_kg
                    counts_and_weights[key]['ready_qty'] += ready_qty
                    counts_and_weights[key]['ready_weight_kg'] += ready_weight_kg
                else:
                    counts_and_weights[key] = {'total_qty': qty, 'total_weight_kg': weight_kg, 'ready_qty': ready_qty, 'ready_weight_kg': ready_weight_kg }

            # Step 4: Convert to a list
            output_list = []
            for key, values in counts_and_weights.items():
                total_weight_tons = values['total_weight_kg'] / 1000  # Convert kg to tons
                ready_weight_tons = values['ready_weight_kg'] / 1000  # Convert kg to tons
                output_list.append({
                    'zca_on': key[0], 
                    'name_th': key[1],
                    'group': key[2], 
                    'subgroup': key[3],
                    'product_type': "WIP",
                    'product_length': round(key[4] / 100, 1), 
                    'Bifrost_total_qty': values['total_qty'],
                    'Bifrost_total_tons': round(total_weight_tons, 2),
                    'Bifrost_ready_qty': values['ready_qty'],
                    'Bifrost_ready_tons': round(ready_weight_tons, 2),
                    'Bifrost_lab_qty': 0,
                    'Bifrost_lab_tons': round(0, 2),
                    'SAP_total_qty': 0,
                    'SAP_total_tons': 0,
                    'SAP_ready_qty': 0,
                    'SAP_ready_tons': 0,
                    'SAP_lab_qty': 0,
                    'SAP_lab_tons': 0,
                })

            # SAP StockBalance From Tiger
            SAP_getlist = Tiger_StockBalance.objects.values()
            for i in SAP_getlist:
                try:
                    i["qi"] = int(i["qi"])
                except:
                    i["qi"] = 0
                try:
                    i["block"] = int(i["block"])
                except:
                    i["block"] = 0
                try:
                    i["urstock"] = int(i["urstock"])
                except:
                    i["urstock"] = 0
                if(i["zca"][3] == "W"):
                    try:
                        query_selct = getDetail_WIP(i["zca"])
                    except:
                        query_selct["field_name"] = "Unknown"
                        query_selct["field_prodgroup"] = "Unknown"
                        query_selct["field_prodname"] = "Unknown"
                        query_selct["field_kgpcs"] = 0
                        query_selct["product_length"] = 0

                    i["zca_on"] = i["zca"]
                    i["name_th"] = query_selct["field_name"]
                    i["group"] = query_selct["field_prodgroup"]
                    i["subgroup"] = query_selct["field_prodname"]
                    i["kgpcs"] = query_selct["field_kgpcs"]
                    i["product_length"] = round(int(query_selct["field_lengthpallet"]) / 100, 1)
                    i["product_type"] = "WIP"
                    

                    i["SAP_total_qty"] = i["urstock"]+i["qi"]
                    try:
                        i["SAP_total_tons"] = round((i["urstock"]+i["qi"])*float(i["kgpcs"])/1000, 2)
                    except:
                        i["SAP_total_tons"] = 0

                    i["SAP_ready_qty"] = i["urstock"]
                    try:
                        i["SAP_ready_tons"] = round((i["urstock"])*float(i["kgpcs"])/1000, 2)
                    except:
                        i["SAP_ready_tons"] = 0

                    i["SAP_lab_qty"] = i["qi"]
                    try:
                        i["SAP_lab_tons"] = round((i["qi"])*float(i["kgpcs"])/1000, 2)
                    except:
                        i["SAP_lab_tons"] = 0

                    have_in_bifrost = False
                    for j in output_list:
                        if j["zca_on"] == i["zca"]:
                            j["SAP_total_qty"] = i["urstock"]+i["qi"]
                            try:
                                j["SAP_total_tons"] = round((i["urstock"]+i["qi"])*float(i["kgpcs"])/1000, 2)
                            except:
                                j["SAP_total_tons"] = 0

                            j["SAP_ready_qty"] = i["urstock"]
                            try:
                                j["SAP_ready_tons"] = round((i["urstock"])*float(i["kgpcs"])/1000, 2)
                            except:
                                j["SAP_ready_tons"] = 0

                            j["SAP_lab_qty"] = i["qi"]
                            try:
                                j["SAP_lab_tons"] = round((i["qi"])*float(i["kgpcs"])/1000, 2)
                            except:
                                j["SAP_lab_tons"] = 0

                            have_in_bifrost = True
                            break
                    if have_in_bifrost == False:
                        i["Bifrost_total_qty"] = 0
                        i["Bifrost_total_tons"] = 0
                        i["Bifrost_ready_qty"] = 0
                        i["Bifrost_ready_tons"] = 0
                        i["Bifrost_lab_qty"] = 0
                        i["Bifrost_lab_tons"] = 0
                        output_list.append(i)
                

            dropdownlist = ViewItemmasterproductwip.objects.exclude(field_prodgroup="NULL").values("field_prodgroup","field_prodname").distinct()
            formatted_data = {}
            for item in dropdownlist:
                group = item['field_prodgroup']
                prodname = item['field_prodname']
                
                if group in formatted_data:
                    formatted_data[group].append(prodname)
                else:
                    formatted_data[group] = [prodname]

            # แปลงเป็นรูปแบบที่ต้องการ
            output = [{"group": key, "subgroup": value} for key, value in formatted_data.items()]
            
            if group_filter != "ALL":
                if subgroup_filter != "ALL":
                    dropdownlist_zca = ViewItemmasterproductwip.objects.filter(field_prodgroup = group_filter,field_prodname = subgroup_filter).values("field_zca","field_name").distinct()
                else:
                    dropdownlist_zca = ViewItemmasterproductwip.objects.filter(field_prodgroup = group_filter).values("field_zca","field_name").distinct()
            else:
                dropdownlist_zca = ViewItemmasterproductwip.objects.values("field_zca","field_name").distinct()

            data = [{"label":"ทั้งหมด", "value":"ALL"}]
            for i in dropdownlist_zca:
                data.append({"label":i["field_zca"] + " " + i["field_name"], "value":i["field_zca"]})

            return Response({'success': True, 'graphdata': formatted_result, 'tabledata': output_list,'dropdownlist': output,'dropdownlist_zca': data})
    
class get_nonmoving(APIView):
    permission_classes = [PlannerPerm | AdminPerm | ManagerPerm | PISPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
            zca_filter = request.query_params.get('zca')
            group_filter = request.query_params.get('group')
            subgroup_filter = request.query_params.get('subgroup')
            size_filter = request.query_params.get('size')


            queryset = ViewWmsMapManagement.objects.filter(map_approve=1).values("zca_on","name_th","machine","qty","product_length","pcsperpallet","product_date","product_shift")

            for i in queryset:
                query_selct = getDetail_WIP(i["zca_on"])
                i["group"] = query_selct["field_prodgroup"]
                i["subgroup"] = query_selct["field_prodname"]
                i["kgpcs"] = query_selct["field_kgpcs"]
            
            filtered_queryset = queryset


            # Apply ZCA filter
            if zca_filter and zca_filter != 'ALL':
                filtered_queryset = [item for item in filtered_queryset if item["zca_on"] == zca_filter]

            if size_filter and size_filter != 'ALL':
                filtered_queryset = [item for item in filtered_queryset if item["product_length"] == float(size_filter)*100]

            # Apply Group and Subgroup filters
            # Note: You need to ensure that the 'group' and 'subgroup' keys are available in your queryset items
            if group_filter and group_filter != 'ALL':
                filtered_queryset = [item for item in filtered_queryset if item.get("group") == group_filter]

            if subgroup_filter and subgroup_filter != 'ALL':
                filtered_queryset = [item for item in filtered_queryset if item.get("subgroup") == subgroup_filter]


            data_all = {}
            for pallet in filtered_queryset:

                if not(pallet["zca_on"] in data_all):
                    data_all[pallet["zca_on"]] = { "name":pallet["name_th"],"kgpcs":pallet["kgpcs"],"scale_0":0,"scale_31":0,"scale_61":0,"scale_100":0 }
                days_old = (datetime.now().date() - pallet["product_date"]).days
                
                if days_old > 100:
                    data_all[pallet["zca_on"]]["scale_100"] += int(i["qty"])
                elif days_old > 60:
                    data_all[pallet["zca_on"]]["scale_61"] += int(i["qty"])
                elif days_old > 30:
                    data_all[pallet["zca_on"]]["scale_31"] += int(i["qty"])
                else:
                    data_all[pallet["zca_on"]]["scale_0"] += int(i["qty"])


            name = []
            s1 = []
            s2 = []
            s3 = []
            s4 = []
            kgpcs = []
            for k in data_all:
                name.append(data_all[k]["name"])
                s1.append(data_all[k]["scale_0"])
                s2.append(data_all[k]["scale_31"])
                s3.append(data_all[k]["scale_61"])
                s4.append(data_all[k]["scale_100"])
                kgpcs.append(data_all[k]["kgpcs"])

            dropdownlist = ViewItemmasterproductwip.objects.exclude(field_prodgroup="NULL").values("field_prodgroup","field_prodname").distinct()
            formatted_data = {}
            for item in dropdownlist:
                group = item['field_prodgroup']
                prodname = item['field_prodname']
                
                if group in formatted_data:
                    formatted_data[group].append(prodname)
                else:
                    formatted_data[group] = [prodname]

            # แปลงเป็นรูปแบบที่ต้องการ
            output = [{"group": key, "subgroup": value} for key, value in formatted_data.items()]
            if group_filter != "ALL":
                if subgroup_filter != "ALL":

                    dropdownlist_zca = ViewItemmasterproductwip.objects.filter(field_prodgroup = group_filter,field_prodname = subgroup_filter).values("field_zca","field_name").distinct()
                else:
                    dropdownlist_zca = ViewItemmasterproductwip.objects.filter(field_prodgroup = group_filter).values("field_zca","field_name").distinct()
            else:
                dropdownlist_zca = ViewItemmasterproductwip.objects.values("field_zca","field_name").distinct()

            data = [{"label":"ทั้งหมด", "value":"ALL"}]
            for i in dropdownlist_zca:
                data.append({"label":i["field_zca"] + " " + i["field_name"], "value":i["field_zca"]})

            return Response({'success': True, 'data_all':data_all, "graphdata":{"name":name,"kgpcs":kgpcs , "s1":s1, "s2":s2, "s3":s3, "s4":s4},'dropdownlist': output,'dropdownlist_zca': data})

class get_PaperPlanPDF(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')
            type_transport_sql = request.query_params.get('type_transport')

            # date_sql = "2023-10-16"
            # shift_sql = "A"

            data = []

            # planning = PlanProduction.objects.filter(pdplan_date=date_sql, pdplan_shift=shift_sql)

            def divide_with_remainder(A, B):
                quotient = A // B
                remainder = A % B
                result = f"{quotient}({remainder})"
                return result

            items_good = Forklift_Worklist.objects.filter(send_date=date_sql, send_shift=shift_sql,product_type="FG").values()


            queryset_all = Forklift_Worklist.objects.filter(
                Q(receive_date=date_sql, receive_shift=shift_sql) |
                Q(send_date=date_sql, send_shift=shift_sql)
                ,type_transport=type_transport_sql
            ).order_by("-id").values()

            for queryset in queryset_all:

                try:
                    queryset["format_date"] = queryset["product_date"].strftime('%d/%m')
                except:
                    queryset["format_date"] = ""

                try:
                    queryset["format_shift"] = queryset["product_shift"]
                except:
                    queryset["format_shift"] = ""
                pallet_to = ViewWmsMapManagement.objects.filter(forklift_link=int(queryset["id"])).values()[0]
                

                result_warehouse = defaultdict(lambda: {})
                query_warehouse = Warehouse.objects.order_by('-id').values()
                for item in query_warehouse:
                    result_warehouse[item["id"]]["name"] = item["name"]

                if int(pallet_to["warehouse_id"]) in result_warehouse:
                    queryset["warehouse_name"] = result_warehouse[int(pallet_to["warehouse_id"])]["name"]
                else:
                    queryset["warehouse_name"] = "Unknown ID " + pallet_to["warehouse_id"]

                if queryset["type_transport"] == "fill":
                    queryset["from"] = queryset["machine"]
                    queryset["to"] = queryset["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"]) 
                elif queryset["type_transport"] == "return":
                    queryset["from"] = queryset["machine"]
                    queryset["to"] = queryset["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"]) 
                elif queryset["type_transport"] == "ticket":

                    pallet_to = ViewWmsMapManagement.objects.filter(forklift_link=int(queryset["id"])).values()[0]
                    map_list_fill_plan_obj = MapListFillPlan.objects.filter(id=pallet_to["maplistfillplan_link_id"]).values()[0]
                    pallet_from = ListTicketPlanProduction.objects.get(id=int(map_list_fill_plan_obj["listticketplan_link_id"]))

                    queryset["note"] = pallet_from.note_production
                    queryset["from"] = queryset["machine"]
                    queryset["to"] = queryset["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"]) 
                
                elif queryset["type_transport"] == "labreturn":

                    pallet_to = ViewWmsMapManagement.objects.filter(forklift_link=int(queryset["id"])).values()[0]
                    map_list_fill_plan_obj = MapListFillPlan.objects.filter(id=pallet_to["maplistfillplan_link_id"]).values()[0]
                    pallet_from = ListLabReturnPlanProduction.objects.get(maplistfillplan_link=int(map_list_fill_plan_obj["id"]))

                    queryset["note"] = pallet_from.note_production
                    queryset["from"] = queryset["machine"]
                    queryset["to"] = queryset["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"]) 

                elif queryset["type_transport"] == "badfill":

                    pallet_to = ViewWmsMapManagement.objects.filter(forklift_link=int(queryset["id"])).values()[0]
                    map_list_fill_plan_obj = MapListFillPlan.objects.filter(id=pallet_to["maplistfillplan_link_id"]).values()[0]
                    pallet_from = ListLabBadUnlockPlanProduction.objects.get(maplistfillplan_link=int(map_list_fill_plan_obj["id"]))

                    queryset["note"] = pallet_from.note_production
                    queryset["from"] = queryset["machine"]
                    queryset["to"] = queryset["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"]) 

                elif queryset["type_transport"] == "withdraw":
                    pallet_maplist = MapListWithdrawPlan.objects.filter(id=int(queryset["maplistwithdrawplan_link_id"])).values()[0]
                    if pallet_maplist["machine"] == "Lab":
                        queryset["from"] = queryset["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"])
                        pallet_to = ViewWmsMapManagement.objects.filter(forklift_link=int(queryset["id"])).values()[0]
                        map_list_withdraw_plan_obj = MapListWithdrawPlan.objects.filter(id=pallet_to["maplistwithdrawplan_link_id"]).values()[0]
                        pallet_from = ListWithdrawPlanProduction.objects.get(id=int(map_list_withdraw_plan_obj["listwithdraw_link_id"]))
                        queryset["note"] = pallet_from.note_production
                        queryset["to"] = "LAB-" + pallet_from.note_production
                    else:
                        queryset["from"] = queryset["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"]) + "-R" + str(pallet_to["row"])+ "-L" + str(pallet_to["level"])
                        queryset["to"] = queryset["machine"]

                elif queryset["type_transport"] == "transfer":
                    pallet_from = ViewWmsMapManagement.objects.filter(forklift_link=int(queryset["id"])).values()[0]
                    pallet_to = ViewWmsMapManagement.objects.filter(forklift_link=int(queryset["id"])).values()[1]

                    if int(pallet_from["warehouse_id"]) in result_warehouse:
                        pallet_from["warehouse_name"] = result_warehouse[int(pallet_from["warehouse_id"])]["name"]
                    else:
                        pallet_from["warehouse_name"] = "Unknown ID " + pallet_from["warehouse_id"]

                    if int(pallet_to["warehouse_id"]) in result_warehouse:
                        pallet_to["warehouse_name"] = result_warehouse[int(pallet_to["warehouse_id"])]["name"]
                    else:
                        pallet_to["warehouse_name"] = "Unknown ID " + pallet_to["warehouse_id"]
                
                    queryset["from"] = pallet_from["warehouse_name"] + "-SZ" + str(pallet_from["zone"]) + "-C" + str(pallet_from["column"]) + "-R" + str(pallet_from["row"])+ "-L" + str(pallet_from["level"])
                    queryset["to"] = pallet_to["warehouse_name"] + "-SZ" + str(pallet_to["zone"]) + "-C" + str(pallet_to["column"]) + "-R" + str(pallet_to["row"])+ "-L" + str(pallet_to["level"])

            

            return Response({'success': True, 'data': queryset_all})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class get_WithdrawPDF(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')
            machine_sql = request.query_params.get('machine')


            query = ListWithdrawPlanProduction.objects.filter(machine=machine_sql,withdraw_keyin="success", receive_date=date_sql, receive_shift=shift_sql).values()
            
            try:
                pdf_costcenter = Info_MCCostcenter.objects.filter(machine=machine_sql).values()[0]["costcenter_machine"]
            except:
                pdf_costcenter = "---"


            pdf_matdoc = "---"
            pdf_approve = "---"

            for i in query:

                try:
                    query_search = ViewItemmasterproductwip.objects.filter(field_zca=i["zca_on"]).values('field_zca','field_name','field_nameeng','field_mc','pcsperpallet')[0]
                    i["pcperpallet"] = int(query_search['pcsperpallet'])
                except:
                    query_search = ViewItemmasterproductfg.objects.filter(zca=i["zca_on"]).values('zca','name','nameen','pcpallet')[0]
                    i["pcperpallet"] = int(query_search['pcpallet'])

                i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
                i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])
                if i["approve_withdraw"] == "success":
                    try:
                        maplist = MapListWithdrawPlan.objects.get(listwithdraw_link=i["id"])
                        i["pdf_qty"] = maplist.qty
                        pdf_approve = i["operator_approve_name"]
                    except:
                        i["pdf_qty"] = "ERROR"

                    try:
                        pdf_matdoc = Tiger_GoodsIssue.objects.filter(idmainfromwms=i["id"]).values()[0]["matdocgr"]
                    except:
                        pass
                else:
                    i["pdf_qty"] = ""

            return Response({'success': True, 'data': query, 'pdf_approve':pdf_approve, 'pdf_costcenter':pdf_costcenter, 'pdf_matdoc':pdf_matdoc})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class get_ReturnGoodPDF(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:

            id_sql = request.query_params.get('id')


            # query = ListWithdrawPlanProduction.objects.filter(machine=machine_sql,withdraw_keyin="success", receive_date=date_sql, receive_shift=shift_sql).values()
            query = ListReturnPlanProduction.objects.filter(id=id_sql).values()

            try:
                pdf_costcenter = Info_MCCostcenter.objects.filter(machine=query[0]["return_machine"]).values()[0]["costcenter_machine"]
            except:
                pdf_costcenter = "---"

            pdf_matdoc = "---"
            pdf_approve = ""

            for i in query:
                try:
                    query_search = ViewItemmasterproductwip.objects.filter(field_zca=i["zca_on"]).values('field_zca','field_name','field_nameeng','field_mc','pcsperpallet')[0]
                    i["pcperpallet"] = int(query_search['pcsperpallet'])
                except:
                    query_search = ViewItemmasterproductfg.objects.filter(zca=i["zca_on"]).values('zca','name','nameen','pcpallet')[0]
                    i["pcperpallet"] = int(query_search['pcpallet'])

                i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
                i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])

                list_select_pallet = ListReturnPalletProduction.objects.filter(returnplan_link=int(i["id"]),return_type="good")
                
                i["batch_list"] = []
                for item in list_select_pallet:
                    batchstr = str(int(item.product_date.year)%10) + str(item.product_date.strftime('%m')) + str(item.product_date.strftime('%d')) + item.product_machine + item.product_shift
                    # Use the 'receive_date' and 'receive_shift' to group the items
                    if batchstr not in i["batch_list"]:
                        i["batch_list"].append(batchstr)

                i["note_good_list"] = []
                select_note = ListReturnPlanNoteProduction.objects.filter(returnplan_link=int(i["id"]),type="good").values()
                for note in select_note:
                    i["note_good_list"].append(note["message"])
                i["note_good"] = ",".join(i["note_good_list"])

                try:
                    pdf_matdoc = Tiger_GoodsReturn.objects.filter(idmainfromwms=str(i["id"]),typereturn="good").values()[0]["matdocreturn"]
                except:
                    pass


            return Response({'success': True, 'data': query, 'pdf_approve':pdf_approve, 'pdf_costcenter':pdf_costcenter, 'pdf_matdoc':pdf_matdoc})
        except Exception as e:
            print("ERROR >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class get_ReturnBadPDF(APIView):
    permission_classes = [PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):

            id_sql = request.query_params.get('id')


            # query = ListWithdrawPlanProduction.objects.filter(machine=machine_sql,withdraw_keyin="success", receive_date=date_sql, receive_shift=shift_sql).values()
            query = ListReturnPlanProduction.objects.filter(id=id_sql).values()

            try:
                pdf_costcenter = Info_MCCostcenter.objects.filter(machine=query[0]["return_machine"]).values()[0]["costcenter_machine"]
            except:
                pdf_costcenter = "---"
                
            pdf_matdoc = "---"
            pdf_approve = ""

            for i in query:
                try:
                    query_search = ViewItemmasterproductwip.objects.filter(field_zca=i["zca_on"]).values('field_zca','field_name','field_nameeng','field_mc','pcsperpallet')[0]
                    i["pcperpallet"] = int(query_search['pcsperpallet'])
                except:
                    query_search = ViewItemmasterproductfg.objects.filter(zca=i["zca_on"]).values('zca','name','nameen','pcpallet')[0]
                    i["pcperpallet"] = int(query_search['pcpallet'])

                i["operator_keyin_name"] = searchInfo_Operator(i["operator_keyin"])
                i["operator_approve_name"] = searchInfo_Operator(i["operator_approve"])

                list_select_pallet = ListReturnPalletProduction.objects.filter(returnplan_link=int(i["id"]),return_type="bad")
                
                i["batch_list"] = []
                for item in list_select_pallet:
                    batchstr = str(int(item.product_date.year)%10) + str(item.product_date.strftime('%m')) + str(item.product_date.strftime('%d')) + item.product_machine + item.product_shift
                    # Use the 'receive_date' and 'receive_shift' to group the items
                    if batchstr not in i["batch_list"]:
                        i["batch_list"].append(batchstr)

                i["note_bad_list"] = []
                select_note = ListReturnPlanNoteProduction.objects.filter(returnplan_link=int(i["id"]),type="bad").values()
                for note in select_note:
                    i["note_bad_list"].append(note["message"])
                i["note_bad"] = ",".join(i["note_bad_list"])

                try:
                    pdf_matdoc = Tiger_GoodsReturn.objects.filter(idmainfromwms=str(i["id"]),typereturn="bad").values()[0]["matdocreturn"]
                except:
                    pass

            return Response({'success': True, 'data': query, 'pdf_approve':pdf_approve, 'pdf_costcenter':pdf_costcenter, 'pdf_matdoc':pdf_matdoc})


# ================  ================  Material  ================  ================  ================ ================ ================ ================


class get_material(APIView):
    permission_classes = [ ScmPerm | PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):

        data_all = Masterwipstk.objects.values()[:100]

        return Response({'success': True, 'data':data_all})

class post_material(APIView):
    permission_classes = [ ScmPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            print(data)
            hscode = data.get('hscode')
            zca = data.get('zca')
            name_th = data.get('name_th')
            name_en = data.get('name_en')
            weight_p_stk = data.get('weight_p_stk')
            sqr_p_stk = data.get('sqr_p_stk')
            stk_p_shift = data.get('stk_p_shift')
            stk_p_hr = data.get('stk_p_hr')
            ton_p_shift = data.get('ton_p_shift')
            ton_p_hr = data.get('ton_p_hr')
            badge = data.get('badge')
            hs = data.get('hs')
            compressed = data.get('compressed')


            # update_or_create
            material, created = Masterwipstk.objects.update_or_create(
                zca=zca,
                hs=hs,
                defaults={
                    'zca': zca,
                    'name_th': name_th,
                    'name_en': name_en,
                    'weight_p_stk': weight_p_stk,
                    'sqr_p_stk': sqr_p_stk,
                    'stk_p_shift': stk_p_shift,
                    'stk_p_hr': stk_p_hr,
                    'ton_p_shift': ton_p_shift,
                    'ton_p_hr': ton_p_hr,
                    'badge': badge,
                    'hs': hs,
                    'compressed': compressed,
                }
            )

            # Return a success message
            if created:
                message = 'Material created successfully'
            else:
                message = 'Material updated successfully'

            return Response({'success': True, 'message': message}, status=status.HTTP_200_OK)
        
        except Masterwipstk.DoesNotExist:
            return Response({'success': False, 'message': 'Material not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class edit_material(APIView):
    permission_classes = [ScmPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            data = request.data
        
            material_id = data.get('id')

            material = Masterwipstk.objects.get(id=material_id)

            # Update the fields
            print(data.get('hscode', material.hscode))

            material.hscode = data.get('hscode', material.hscode)
            material.hs = data.get('hs', material.hs)
            material.zca = data.get('zca', material.zca)
            material.name_th = data.get('name_th', material.name_th)
            material.name_en = data.get('name_en', material.name_en)
            material.weight_p_stk = data.get('weight_p_stk', material.weight_p_stk)
            material.sqr_p_stk = data.get('sqr_p_stk', material.sqr_p_stk)
            material.stk_p_shift = data.get('stk_p_shift', material.stk_p_shift)
            material.stk_p_hr = data.get('stk_p_hr', material.stk_p_hr)
            material.ton_p_shift = data.get('ton_p_shift', material.ton_p_shift)
            material.ton_p_hr = data.get('ton_p_hr', material.ton_p_hr)
            material.badge = data.get('badge', material.badge)
            material.compressed = data.get('compressed', material.compressed)


            material.save()

            return Response({'success': True, 'message': 'Material updated successfully'}, status=status.HTTP_200_OK)
        
        except Masterwipstk.DoesNotExist:
            return Response({'success': False, 'message': 'Material not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)



class delete_material(APIView):
    permission_classes = [PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        Masterwipstk.objects.get(id=request_input.get("id_select")).delete()
        return Response({'success': True})



# ================  ================  WIP!  ================  ================  ================ ================ ================ ================

class check_zca(APIView):
    permission_classes = [ ScmPerm| PlannerPerm|PlantCoPerm| AdminPerm]

    def post(self, request, *args, **kwargs):
        zca = request.data.get('zca', '')
        zca_type = request.data.get('type', '')  # Capture the type from the request
        print(zca,)
        print(zca_type)
        if zca_type == "WIP":
            # For WIP, ZCA must start with "ZCAW"
            if not zca.startswith("ZCAW"):
                return Response({'exists': False, 'message': 'Invalid ZCA format for WIP. It must start with "ZCAW".'}, status=200)
            
            # Check in ViewItemmasterproductwip table
            exists = ViewItemmasterproductwip.objects.filter(field_zca=zca).exists()
            if exists:
                return Response({'exists': True, 'message': 'ZCA already exists in ViewItemmasterproductwip'}, status=200)
            else:
                return Response({'exists': False, 'message': 'ZCA does not exist in ViewItemmasterproductwip'}, status=200)

        elif zca_type == "FG":
            # For FG, ZCA must start with "ZCA" but not "ZCAW"
            if not zca.startswith("ZCA") or zca.startswith("ZCAW"):
                return Response({'exists': False, 'message': 'Invalid ZCA format for FG. It must start with "ZCA" and not "ZCAW".'}, status=200)
            
            # Check in ViewItemmasterproductfg table
            exists = ViewItemmasterproductfg.objects.filter(zca=zca).exists()
            if exists:
                return Response({'exists': True, 'message': 'ZCA already exists in ViewItemmasterproductfg'}, status=200)
            else:
                return Response({'exists': False, 'message': 'ZCA does not exist in ViewItemmasterproductfg'}, status=200)

        else:
            return Response({'exists': False, 'message': 'Invalid type. Please enter a valid type (FG or WIP).'}, status=200)
        
class get_editzcawip(APIView):
    permission_classes = [ ScmPerm| ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]
    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):

        data_all = ViewItemmasterproductwip.objects.filter().order_by("-field_id").values()
        for data in data_all:
            # change date format
            if(data['updated_at'] != None):
                data['updated_at_new'] = data['updated_at'].strftime('%d-%m-%Y')
            if(data['created_at'] != None):
                data['created_at_new'] = data['created_at'].strftime('%d-%m-%Y')
            # get operator name
            if(data['operator_keyin'] != None):
                data['operator_keyin_new'] = searchInfo_Operator(data['operator_keyin'])
            if(data['operator_edit'] != None):
                data['operator_edit_new'] = searchInfo_Operator(data['operator_edit'])

        return Response({'success': True, 'data':data_all})

class post_editzcawip_2(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        machines = request_input.get("mc")
        query = ViewItemmasterproductwip.objects.filter(field_zca=request_input.get("zca"))
        machine_list = list(query.values_list('field_mc', flat=True))

        delete_list=[]

        for mc in machine_list:
            if(mc not in machines):
                delete_list.append(mc)

        for mc in machines:
            # update
            if(mc in machine_list and mc not in delete_list):

                update_wip = ViewItemmasterproductwip.objects.get(field_zca=request_input.get("zca"), field_mc=mc)
                update_wip.field_mc = mc
                update_wip.field_nameShort = request_input.get("nameShort")
                update_wip.field_name = request_input.get("namethai")
                update_wip.field_nameeng = request_input.get("nameeng")
                update_wip.field_type = request_input.get("type")
                update_wip.brand = request_input.get("brand")
                update_wip.field_prodgroup = request_input.get("prodgroup")
                update_wip.field_prodname = request_input.get("prodname")
                update_wip.field_size = request_input.get("size")
                update_wip.field_length = request_input.get("length")
                update_wip.field_lengthpallet = request_input.get("lengthpallet")
                update_wip.field_pcspallet = request_input.get("pcsperpallet")
                update_wip.field_layer = request_input.get("layer")
                update_wip.field_kgpcs = request_input.get("kgpcs")
                update_wip.pcsperpallet = request_input.get("pcsperpallet")
                update_wip.type1 = request_input.get("type1")
                update_wip.tickness = request_input.get("tickness")
                update_wip.ct1 = check_ok(request_input.get("CT1"), "WIP")
                update_wip.ct2 = check_ok(request_input.get("CT2"), "WIP")
                update_wip.ct3 = check_ok(request_input.get("CT3"), "WIP")
                update_wip.ct4 = check_ok(request_input.get("CT4"), "WIP")
                update_wip.xy1 = check_ok(request_input.get("XY1"), "WIP")
                update_wip.cm5 = check_ok(request_input.get("CM5"), "WIP")
                update_wip.cm6 = check_ok(request_input.get("CM6"), "WIP")
                update_wip.cm7 = check_ok(request_input.get("CM7"), "WIP")
                update_wip.cm8 = check_ok(request_input.get("CM8"), "WIP")
                update_wip.as1 = check_ok(request_input.get("AS1"), "WIP")
                update_wip.pk1 = check_ok(request_input.get("PK1"), "WIP")
                update_wip.pk2 = check_ok(request_input.get("PK2"), "WIP")
                update_wip.pk3 = check_ok(request_input.get("PK3"), "WIP")
                update_wip.pk4 = check_ok(request_input.get("PK4"), "WIP")
                update_wip.pk5 = check_ok(request_input.get("PK5"), "WIP")
                update_wip.dp1 = check_ok(request_input.get("DP1"), "WIP")
                update_wip.det = check_ok(request_input.get("DET"), "WIP")
                update_wip.ms1 = check_ok(request_input.get("MS1"), "WIP")
                update_wip.oc1 = check_ok(request_input.get("OC1"), "WIP")
                update_wip.oc2 = check_ok(request_input.get("OC2"), "WIP")
                update_wip.os1 = check_ok(request_input.get("OS1"), "WIP")
                update_wip.pl1 = check_ok(request_input.get("PL1"), "WIP")
                update_wip.rt1 = check_ok(request_input.get("RT1"), "WIP")
                update_wip.rt2 = check_ok(request_input.get("RT2"), "WIP")
                update_wip.sd1 = check_ok(request_input.get("SD1"), "WIP")
                update_wip.seg = check_ok(request_input.get("SEG"), "WIP")
                update_wip.dp2 = check_ok(request_input.get("DP2"), "WIP")
                update_wip.pk6 = check_ok(request_input.get("PK6"), "WIP")
                update_wip.operator_edit= request.user.employee_id
                update_wip.save()

            elif(mc not in machine_list and mc not in delete_list):
                # add (not in machine_list and delete_list)
                print('add machine>>>', mc)
                save_wip = ViewItemmasterproductwip(
                    field_zca=request_input.get("zca"),
                    field_mc=mc,
                    field_nameShort=request_input.get("nameShort"),
                    field_name=request_input.get("namethai"),
                    field_nameeng=request_input.get("nameeng"),
                    field_type=request_input.get("type"),
                    brand=request_input.get("brand"),
                    field_prodgroup=request_input.get("prodgroup"),
                    field_prodname=request_input.get("prodname"),
                    field_size=request_input.get("size"),
                    field_length=request_input.get("length"),
                    field_lengthpallet=request_input.get("lengthpallet"),
                    field_pcspallet=request_input.get("pcsperpallet"),
                    field_layer=request_input.get("layer"),
                    field_kgpcs=request_input.get("kgpcs"),
                    pcsperpallet=request_input.get("pcsperpallet"),
                    type1=request_input.get("type1"),
                    tickness=request_input.get("tickness"),
                    ct1=check_ok(request_input.get("CT1"),"WIP"),
                    ct2=check_ok(request_input.get("CT2"),"WIP"),
                    ct3=check_ok(request_input.get("CT3"),"WIP"),
                    ct4=check_ok(request_input.get("CT4"),"WIP"),
                    xy1=check_ok(request_input.get("XY1"),"WIP"),
                    cm5=check_ok(request_input.get("CM5"),"WIP"),
                    cm6=check_ok(request_input.get("CM6"),"WIP"),
                    cm7=check_ok(request_input.get("CM7"),"WIP"),
                    cm8=check_ok(request_input.get("CM8"),"WIP"),
                    as1=check_ok(request_input.get("AS1"),"WIP"),
                    pk1=check_ok(request_input.get("PK1"),"WIP"),
                    pk2=check_ok(request_input.get("PK2"),"WIP"),
                    pk3=check_ok(request_input.get("PK3"),"WIP"),
                    pk4=check_ok(request_input.get("PK4"),"WIP"),
                    pk5=check_ok(request_input.get("PK5"),"WIP"),
                    dp1=check_ok(request_input.get("DP1"),"WIP"),
                    det=check_ok(request_input.get("DET"),"WIP"),
                    ms1=check_ok(request_input.get("MS1"),"WIP"),
                    oc1=check_ok(request_input.get("OC1"),"WIP"),
                    oc2=check_ok(request_input.get("OC2"),"WIP"),
                    os1=check_ok(request_input.get("OS1"),"WIP"),
                    pl1=check_ok(request_input.get("PL1"),"WIP"),
                    rt1=check_ok(request_input.get("RT1"),"WIP"),
                    rt2=check_ok(request_input.get("RT2"),"WIP"),
                    sd1=check_ok(request_input.get("SD1"),"WIP"),
                    seg=check_ok(request_input.get("SEG"),"WIP"),
                    dp2=check_ok(request_input.get("DP2"),"WIP"),
                    pk6=check_ok(request_input.get("PK6"),"WIP"),
                    operator_keyin= request.user.employee_id
                )
                save_wip.save()

        # delete
        if delete_list:
            query.filter(field_mc__in=delete_list).delete()

        # for mc in delete_list:
        #     if(mc in delete_list):
        #         print('delete machine>>>', mc)

        #         delete_wip = query.get(field_mc=mc)
        #         delete_wip

        #         ViewItemmasterproductwip.objects.get(field_id=i['field_id']).delete()
        
        return Response({'success': True})


    
class post_editzcawip(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data

        for mc in request_input.get("mc"):
            save_wip = ViewItemmasterproductwip(
                field_zca=request_input.get("zca"),
                field_mc=mc,
                field_name=request_input.get("namethai"),
                field_nameShort=request_input.get("nameShort"),
                field_nameeng=request_input.get("nameeng"),
                field_type=request_input.get("type"),
                brand=request_input.get("brand"),
                field_prodgroup=request_input.get("prodgroup"),
                field_prodname=request_input.get("prodname"),
                field_size=request_input.get("size"),
                field_length=request_input.get("length"),
                field_lengthpallet=request_input.get("lengthpallet"),
                field_pcspallet=request_input.get("pcsperpallet"),
                field_layer=request_input.get("layer"),
                field_kgpcs=request_input.get("kgpcs"),
                pcsperpallet=request_input.get("pcsperpallet"),
                type1=request_input.get("type1"),
                tickness=request_input.get("tickness"),
                ct1=check_ok(request_input.get("CT1"),"WIP"),
                ct2=check_ok(request_input.get("CT2"),"WIP"),
                ct3=check_ok(request_input.get("CT3"),"WIP"),
                ct4=check_ok(request_input.get("CT4"),"WIP"),
                xy1=check_ok(request_input.get("XY1"),"WIP"),
                cm5=check_ok(request_input.get("CM5"),"WIP"),
                cm6=check_ok(request_input.get("CM6"),"WIP"),
                cm7=check_ok(request_input.get("CM7"),"WIP"),
                cm8=check_ok(request_input.get("CM8"),"WIP"),
                as1=check_ok(request_input.get("AS1"),"WIP"),
                pk1=check_ok(request_input.get("PK1"),"WIP"),
                pk2=check_ok(request_input.get("PK2"),"WIP"),
                pk3=check_ok(request_input.get("PK3"),"WIP"),
                pk4=check_ok(request_input.get("PK4"),"WIP"),
                pk5=check_ok(request_input.get("PK5"),"WIP"),
                dp1=check_ok(request_input.get("DP1"),"WIP"),
                det=check_ok(request_input.get("DET"),"WIP"),
                ms1=check_ok(request_input.get("MS1"),"WIP"),
                oc1=check_ok(request_input.get("OC1"),"WIP"),
                oc2=check_ok(request_input.get("OC2"),"WIP"),
                os1=check_ok(request_input.get("OS1"),"WIP"),
                pl1=check_ok(request_input.get("PL1"),"WIP"),
                rt1=check_ok(request_input.get("RT1"),"WIP"),
                rt2=check_ok(request_input.get("RT2"),"WIP"),
                sd1=check_ok(request_input.get("SD1"),"WIP"),
                seg=check_ok(request_input.get("SEG"),"WIP"),
                dp2=check_ok(request_input.get("DP2"),"WIP"),
                pk6=check_ok(request_input.get("PK6"),"WIP"),
                operator_keyin= request.user.employee_id
            )
            save_wip.save()
        
        return Response({'success': True})
    
class post_deletezcawip(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        ViewItemmasterproductwip.objects.get(field_id=request_input.get("id_select")).delete()
        return Response({'success': True})

# ================  ================  FG!  ================  ================  ================ ================ ================ ================


class get_editzcafg(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        try:
            data_all = ViewItemmasterproductfg.objects.order_by("-id").values()

            for data in data_all:
                # change date format
                if(data['updated_at'] != None):
                    data['updated_at_new'] = data['updated_at'].strftime('%d-%m-%Y')
                if(data['created_at'] != None):
                    data['created_at_new'] = data['created_at'].strftime('%d-%m-%Y')
                # get operator name
                if(data['operator_keyin'] != None):
                    data['operator_keyin_new'] = searchInfo_Operator(data['operator_keyin'])
                if(data['operator_edit'] != None):
                    data['operator_edit_new'] = searchInfo_Operator(data['operator_edit'])

            data = {'success': True, 'data':data_all}
            return Response(data)
        except Exception as e:
            # logger.error("Error >>> %s", e)
            print("Error >>>",e)
            return Response({'success': False, 'error':str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class post_editzcafg_2(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data

        update_fg = ViewItemmasterproductfg.objects.get(zca=request_input.get("zca"))
        update_fg.zca = request_input.get("zca")
        update_fg.name = request_input.get("namethai")
        update_fg.nameshort = request_input.get("nameshort")
        update_fg.namethai = request_input.get("namethai")
        update_fg.nameen = request_input.get("nameeng")
        update_fg.type = request_input.get("type")
        update_fg.brand = request_input.get("brand")
        update_fg.zcacustomer = request_input.get("zcacustomer")
        update_fg.format = request_input.get("format","TH")
        update_fg.tis = request_input.get("tis",0)
        update_fg.size = request_input.get("size")
        update_fg.sizemm = request_input.get("sizemm")
        update_fg.pcpallet = request_input.get("pcsperpallet")
        update_fg.kg = request_input.get("kgpcs")
        update_fg.om = "PC"
        update_fg.hs3_tl=check_ok(request_input.get("HS3"),"FG")
        update_fg.hs4_tl=check_ok(request_input.get("HS4"),"FG")
        update_fg.hs5_tl=check_ok(request_input.get("HS5"),"FG")
        update_fg.hs6_tl=check_ok(request_input.get("HS6"),"FG")
        update_fg.hs7_tl=check_ok(request_input.get("HS7"),"FG")
        update_fg.hs8_tl=check_ok(request_input.get("HS8"),"FG")
        update_fg.hs9_tl=check_ok(request_input.get("HS9"),"FG")
        update_fg.ct1_tl=check_ok(request_input.get("CT1"),"FG")
        update_fg.ct2_tl=check_ok(request_input.get("CT2"),"FG")
        update_fg.ct3_tl=check_ok(request_input.get("CT3"),"FG")
        update_fg.ct4_tl=check_ok(request_input.get("CT4"),"FG")
        update_fg.cm5_tl=check_ok(request_input.get("CM5"),"FG")
        update_fg.cm6_tl=check_ok(request_input.get("CM6"),"FG")
        update_fg.cm7_tl=check_ok(request_input.get("CM7"),"FG")
        update_fg.cm8_tl=check_ok(request_input.get("CM8"),"FG")
        update_fg.dp1_tl=check_ok(request_input.get("DP1"),"FG")
        update_fg.dp2_tl=check_ok(request_input.get("DP2"),"FG")
        update_fg.det_tl=check_ok(request_input.get("DET"),"FG")
        update_fg.ms1_tl=check_ok(request_input.get("MS1"),"FG")
        update_fg.oc1_tl=check_ok(request_input.get("OC1"),"FG")
        update_fg.oc2_tl=check_ok(request_input.get("OC2"),"FG")
        update_fg.os1_tl=check_ok(request_input.get("OS1"),"FG")
        update_fg.pk1_tl=check_ok(request_input.get("PK1"),"FG")
        update_fg.pk2_tl=check_ok(request_input.get("PK2"),"FG")
        update_fg.pk3_tl=check_ok(request_input.get("PK3"),"FG")
        update_fg.pk4_tl=check_ok(request_input.get("PK4"),"FG")
        update_fg.pk5_tl=check_ok(request_input.get("PK5"),"FG")
        update_fg.pk6_tl=check_ok(request_input.get("PK6"),"FG")
        update_fg.pl1_tl=check_ok(request_input.get("PL1"),"FG")
        update_fg.rt1_tl=check_ok(request_input.get("RT1"),"FG")
        update_fg.rt2_tl=check_ok(request_input.get("RT2"),"FG")
        update_fg.sd1_tl=check_ok(request_input.get("SD1"),"FG")
        update_fg.seg_tl=check_ok(request_input.get("SEG"),"FG")
        update_fg.operator_edit= request.user.employee_id
        update_fg.save()
        
        return Response({'success': True})

class post_editzcafg(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data

        save_fg = ViewItemmasterproductfg(
            zca=request_input.get("zca"),
            name=request_input.get("namethai"),
            namethai=request_input.get("namethai"),
            nameshort=request_input.get("nameshort"),
            nameen=request_input.get("nameeng"),
            type=request_input.get("type"),
            brand=request_input.get("brand"),
            zcacustomer=request_input.get("zcacustomer"),
            format=request_input.get("format","TH"),
            tis=request_input.get("tis",0),
            size=request_input.get("size"),
            sizemm=request_input.get("sizemm"),
            pcpallet=request_input.get("pcsperpallet"),
            kg=request_input.get("kgpcs"),
            om="PC",

            hs3_tl=check_ok(request_input.get("HS3"),"FG"),
            hs4_tl=check_ok(request_input.get("HS4"),"FG"),
            hs5_tl=check_ok(request_input.get("HS5"),"FG"),
            hs6_tl=check_ok(request_input.get("HS6"),"FG"),
            hs7_tl=check_ok(request_input.get("HS7"),"FG"),
            hs8_tl=check_ok(request_input.get("HS8"),"FG"),
            hs9_tl=check_ok(request_input.get("HS9"),"FG"),
            ct1_tl=check_ok(request_input.get("CT1"),"FG"),
            ct2_tl=check_ok(request_input.get("CT2"),"FG"),
            ct3_tl=check_ok(request_input.get("CT3"),"FG"),
            ct4_tl=check_ok(request_input.get("CT4"),"FG"),
            cm5_tl=check_ok(request_input.get("CM5"),"FG"),
            cm6_tl=check_ok(request_input.get("CM6"),"FG"),
            cm7_tl=check_ok(request_input.get("CM7"),"FG"),
            cm8_tl=check_ok(request_input.get("CM8"),"FG"),
            dp1_tl=check_ok(request_input.get("DP1"),"FG"),
            dp2_tl=check_ok(request_input.get("DP2"),"FG"),
            det_tl=check_ok(request_input.get("DET"),"FG"),
            ms1_tl=check_ok(request_input.get("MS1"),"FG"),
            oc1_tl=check_ok(request_input.get("OC1"),"FG"),
            oc2_tl=check_ok(request_input.get("OC2"),"FG"),
            os1_tl=check_ok(request_input.get("OS1"),"FG"),
            pk1_tl=check_ok(request_input.get("PK1"),"FG"),
            pk2_tl=check_ok(request_input.get("PK2"),"FG"),
            pk3_tl=check_ok(request_input.get("PK3"),"FG"),
            pk4_tl=check_ok(request_input.get("PK4"),"FG"),
            pk5_tl=check_ok(request_input.get("PK5"),"FG"),
            pk6_tl=check_ok(request_input.get("PK6"),"FG"),
            pl1_tl=check_ok(request_input.get("PL1"),"FG"),
            rt1_tl=check_ok(request_input.get("RT1"),"FG"),
            rt2_tl=check_ok(request_input.get("RT2"),"FG"),
            sd1_tl=check_ok(request_input.get("SD1"),"FG"),
            seg_tl=check_ok(request_input.get("SEG"),"FG"),
            operator_edit= request.user.employee_id,
        )
        save_fg.save()
        
        return Response({'success': True})
    
class post_deletezcafg(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        ViewItemmasterproductfg.objects.get(id=request_input.get("id_select")).delete()
        return Response({'success': True})


# ================  ================  ProcessLock  ================  ================ ================ ================ ================



class post_deleteplan(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        plan_id = request_input.get('planid')
        print(plan_id)
        if not plan_id:
            return Response({'success': False, 'message': 'Plan ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # Delete all ProcessLock records with the given planid
                ProcessLock.objects.filter(planid=plan_id).delete()

            return Response({'success': True}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
class get_dupe_zca_processlock(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        # Get the 'zca' from the query parameters
        zca = request.query_params.get('zca', None)

        # If 'zca' is provided, search for it in the ProcessLock model
        if zca:
            # Filter the ProcessLock model by 'field_zca'
            data_all = ProcessLock.objects.filter(field_zca=zca).order_by("-planid").values()

            if data_all.exists():
                return Response({'success': True, 'exists': True, 'data': list(data_all)})

            return Response({'success': True, 'exists': False, 'message': f'ZCA {zca} does not exist in ProcessLock'})

        # If 'zca' is not provided in the request
        return Response({'success': False, 'message': 'ZCA parameter is missing'}, status=400)

class get_editprocesslock(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):

        data_all = ProcessLock.objects.order_by("-planid").values()

        return Response({'success': True, 'data':data_all})

class update_processlock(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            process_list = request.data.get('process', [])

            for process in process_list:
                field_id = process.get('field_id')

                if field_id:
                    # Update the existing ProcessLock
                    try:
                        process_lock = ProcessLock.objects.get(field_id=field_id)
                        serializer = ProcessLockSerializer(process_lock, data=process, partial=True)
                        if serializer.is_valid(raise_exception=True):
                            serializer.save()
                    except ProcessLock.DoesNotExist:
                        return Response({'success': False, 'message': f'ProcessLock with id {field_id} does not exist'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                else:
                    # Create a new ProcessLock
                    serializer = ProcessLockSerializer(data=process)
                    if serializer.is_valid(raise_exception=True):
                        serializer.save()

            return Response({'success': True, 'message': 'ProcessLock records updated/created successfully'})
        
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ToggleDisableView(APIView):
    permission_classes = [ScmPerm|ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema(request_body=None)
    def post(self, request, *args, **kwargs):
        try:
            # Access the parsed data directly from request.data
            id_select = request.data.get('id_select')
            disable = request.data.get('disable')

            if id_select is None or disable is None:
                return Response({'success': False, 'message': 'Invalid data provided'}, status=400)

            # Find the ProcessLock object by id_select
            process_lock = ProcessLock.objects.get(field_id=id_select)
            process_lock.disable = disable  # This will be saved as 0 or 1 due to the save method
            process_lock.save()

            return Response({'success': True, 'message': 'Status updated successfully'})
        except ProcessLock.DoesNotExist:
            return Response({'success': False, 'message': 'ProcessLock not found'}, status=404)
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=500)


class get_zca_data(APIView):
    permission_classes = [ScmPerm|ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
        zca = request.query_params.get('zca')
        print(f"Received request for ZCA: {zca}")

        if not zca:
            return Response({'error': 'ZCA parameter is missing'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if 'ZCAW' in zca:
                process_locks = ViewItemmasterproductwip.objects.filter(field_zca=zca)
                if process_locks.exists():
                    process_lock = process_locks.first()  # Get the first matching object
                    data = {
                        'name': process_lock.field_name,
                        'machine': process_lock.field_mc,
                        'type': process_lock.field_type
                    }

                    return Response(data, status=status.HTTP_200_OK)
                else:
                    return Response({'error': 'ZCA not found'}, status=status.HTTP_404_NOT_FOUND)
            elif 'ZCA' in zca:
                process_locks = ViewItemmasterproductfg.objects.filter(zca=zca)
                mc = Masterwipstk.objects.filter(zca=zca)
                if process_locks.exists() and mc.exists():
                    process_lock = process_locks.first()  # Get the first matching object
                    mc_type=mc.first()
                    data = {
                        'name': process_lock.name,
                        'machine':mc_type.hs,
                        'type': process_lock.type
                    }
                    return Response(data, status=status.HTTP_200_OK)
                else:
                    return Response({'error': 'ZCA not found'}, status=status.HTTP_404_NOT_FOUND)
            else:
                data = {
                    'name': '-',
                    'machine': '-',
                    'type': '-'
                }
                return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': 'An error occurred'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class post_editprocesslock(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data

        # Generate a new non-duplicate Plan ID
        latest_process_lock = ProcessLock.objects.order_by('-planid').first()
        if latest_process_lock:
            latest_planid = latest_process_lock.planid
        else:
            latest_planid = 0
        
        new_planid = latest_planid + 1

        try:
            for i in request_input.get("process"):
                save_process = ProcessLock(
                    field_mc=i["Machine"],
                    field_typeno=0,
                    field_type=i["Type"],
                    field_zca=i["zca"],
                    field_name=i["name"],
                    field_source=i["source"],
                    order=i["order"],
                    field_destination=i["destination"],
                    planid=new_planid  # Use the new non-duplicate Plan ID
                )
                save_process.save()
            return Response({'message': 'Process locks created successfully', 'planid': new_planid}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class post_deleteprocesslock(APIView):
    permission_classes = [ScmPerm|PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        
        field_id = request_input.get("id_select")
        
        if field_id:
            try:
                process_lock = ProcessLock.objects.get(field_id=field_id)
                process_lock.delete()
            except ProcessLock.DoesNotExist:
                pass  

        
        return Response({'success': True})


# ================  ================    ================  ================ ================ ================ ================



class export_csv(APIView):
    permission_classes = [ScmPerm|ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]
    def post(self, request, *args, **kwargs):
        data = request.data['params']['data']


        # Create a buffer
        buffer = StringIO()
        writer = csv.writer(buffer)
        # Write the headers (Assuming all dictionaries have the same keys)
        if data:
            headers = data[0].keys()
            print("Headers:", headers)
            writer.writerow(headers)

            # Write the data
            for row in data:
                writer.writerow(row.values())

        # Ensure all data is written
        csv_content = buffer.getvalue()
        utf8_bom_csv_content = '\ufeff' + csv_content
        encoded_csv_content = utf8_bom_csv_content.encode('utf-8')
        
        # Prepare the response
        response = HttpResponse(encoded_csv_content, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="data.csv"'

        return response

class get_edittedplan_production(APIView):
    permission_classes = [ScmPerm|PlannerPerm | AdminPerm | ManagerPerm | PISPerm]

    @swagger_auto_schema()
    def get(self, request, *args, **kwargs):
            def calculate_time_difference(start_time, end_time):
                """คำนวณความต่างของเวลาในหน่วยวินาที."""
                time_diff = end_time - start_time
                return time_diff.total_seconds()
        # try:
            date_sql = request.query_params.get('date')
            shift_sql = request.query_params.get('shift')
            sevenday_sql = request.query_params.get('sevenday', 'false').lower() == 'true'

            # Parse the start date
            start_date = datetime.strptime(date_sql, "%Y-%m-%d")

            # Initialize the query filter
            query_filter = Q()

            # Add conditions based on the sevenday parameter
            if sevenday_sql:
                # Calculate the end date if sevenday is True
                seven_days_ago = start_date - timedelta(days=7)
                query_filter |= Q(created_at__gte=seven_days_ago)
                query_filter &= Q(pdplan_shift=shift_sql)
                # query_filter |= Q(receive_date__range=[end_date, start_date])
                # query_filter |= Q(send_date__range=[end_date, start_date])
            else:
                # Apply the original filter if sevenday is False
                seven_days_ago = start_date - timedelta(days=1)
                query_filter |= Q(created_at__gte=seven_days_ago)

            # Apply the filter to the queryset
            queryset = PlanProduction.objects.filter(query_filter).values()


            data = {}
            data["work"] = []
            data["chart_pie"] = {'all': 0, 'inprogress': 0, 'editted': 0,'Eddited':0}
            data["production_report"] = []
            data['plan'] = {}
            for plan in queryset:
                try:
                    if data['plan'][plan['pdplan_operator']]:
                        data['plan'][plan['pdplan_operator']] += 1
                    else:
                        data['plan'][plan['pdplan_operator']] = 1
                except:
                    data['plan'][plan['pdplan_operator']] = 1
                if plan["status_edit"] is not None:
                    data["chart_pie"]["editted"] += 1
                else:
                    # Handle the case where forklift_scan_finish_time is None
                    data["chart_pie"]["inprogress"] += 1

                data["chart_pie"]["all"] += 1

                # Forklift Stat
                if plan["status_edit"]:
                    index_found = None
                    for index, report in enumerate(data["production_report"]):
                        if report["employee_id"] == plan["pdplan_operator"]:
                            index_found = index
                            break

                    if index_found is not None:
                        data["production_report"][index_found]["work_success"] += 1
                    else:
                        data["production_report"].append(
                            {
                                "employee_id":plan["pdplan_operator"],
                                "name":CustomUser.objects.get(employee_id=plan["pdplan_operator"]).first_name,
                                "efficiency": 0,
                                "work_success": 1,
                            }
                        )
                else:
                    index_found = None
                    for index, report in enumerate(data["production_report"]):
                        if report["employee_id"] == plan["pdplan_operator"]:
                            index_found = index
                            break

                    if index_found is None:
                        data["production_report"].append(
                            {
                                "employee_id":plan["pdplan_operator"],
                                "name":CustomUser.objects.get(employee_id=plan["pdplan_operator"]).first_name,
                                "efficiency": 0,
                                "work_success": 0,
                            }
                        )
                # else:
                #     if data["production_report"][plan["pdplan_operator"]]:
                #         pass
                #     else:
                #         data["production_report"].append(
                #             {
                #                 "employee_id":plan["pdplan_operator"],
                #                 "name":CustomUser.objects.get(employee_id=plan["pdplan_operator"]).first_name,
                #                 "efficiency": 0,
                #                 "work_success": 0,
                #             }
                #         )
                data["work"].append(plan)
            
            # for forklift in data["production_report"]:
            #     forklift["efficiency"] = (forklift["total_dis"]/(float(data["para_effciency"])/100*float(data["para_velocity"])*1000/3600)) / (float(forklift["total_time"])) *100
            #     forklift["efficiency"] = round(forklift["efficiency"], 2)
            return Response({'success': True, 'data': data})


class get_userForklift(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | AdminPerm]
    def get(self, request, *args, **kwargs):
        results = CustomUser.objects.filter(role_id=5).values('employee_id','first_name','last_name')
        for result in results:
            result['value'] = result['employee_id']
            result['label'] = f"{result['employee_id']} {result['first_name']} {result['last_name']}"
            del result['first_name']
            del result['last_name']
            del result['employee_id']

        return Response({'success': True, 'data': results})








import random

################################################################
# 
# Flow and CheckWIP
# 
################################################################


class Material:
    def __init__(self, code, field_name, initial_stock, lead_time, is_finished_good=False):
        self.code = code
        self.field_name = field_name
        self.frozen = []
        self.actual = []
        self.dates = []
        self.initial_stock = int(initial_stock)
        self.current_stock = []  # List to store stock changes over time
        self.lead_time = int(lead_time)
        self.children = []
        self.is_finished_good = is_finished_good
        self.level = 0
        self.parent = None  # Initialize parent attribute

    def __str__(self):
        return (f"Material(code={self.code}, frozen={self.frozen}, actual={self.actual}, dates={self.dates}, "
                f"initial_stock={self.initial_stock}, current_stock={self.current_stock[-1]['stock'] if self.current_stock else 'N/A'}, "
                f"lead_time={self.lead_time}, is_finished_good={self.is_finished_good}, children={len(self.children)}, level={self.level})")

    def add_child(self, child):
        child.parent = self  # Set parent attribute for the child
        self.children.append(child)

    def add_frozen_value(self, date, value):
        if date in self.dates:
            index = self.dates.index(date)
            self.actual[index] = value
        else:
            self.dates.append(date)
            self.frozen.append(value)
            self.actual.append(0)  # Initialize corresponding actual value to 0

    def add_actual_value(self, date, value):
        if date in self.dates:
            index = self.dates.index(date)
            self.actual[index] = value
        else:
            self.dates.append(date)
            self.actual.append(value)
            self.frozen.append(0)  # Initialize corresponding frozen value to 0

    def set_hierarchy_level(self, level=0):
        self.level = level
        for child in self.children:
            child.set_hierarchy_level(level + 1)

    def to_hierarchy(self):
        return {
            'code': self.code,
            'field_name': self.field_name,
            'level': self.level,
            'parent': self.parent.code if self.parent else None,
            'stock': self.current_stock,
            'dates': self.dates,
            'actual': self.actual,
            'frozen': self.frozen,
            'children': [child.to_hierarchy() for child in self.children]
        }

    def get_all_event(self):
        all_dates = self.dates[:]
        all_frozen = self.frozen[:]
        all_actual = self.actual[:]

        event_dict = {}
        for date, frozen, actual in zip(all_dates, all_frozen, all_actual):
            if date not in event_dict:
                event_dict[date] = {'frozen': 0, 'actual': 0}
            frozen_value = frozen if frozen is not None else 0
            actual_value = actual if actual is not None else 0
            event_dict[date]['frozen'] += frozen_value
            event_dict[date]['actual'] += actual_value

        sorted_events = sorted(event_dict.items())

        all_dates = [event[0] for event in sorted_events]
        all_frozen = [event[1]['frozen'] for event in sorted_events]
        all_actual = [event[1]['actual'] for event in sorted_events]

        return all_dates, all_frozen, all_actual

    def estimate_stock(self, current_date, start_date, end_date):
        begin = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        stock_changes = []
        previous_stock = self.initial_stock

        all_dates, all_frozen, all_actual = self.get_all_event()

        if self.lead_time:
            all_dates_shift = [date + timedelta(days=self.lead_time) for date in all_dates]
        else:
            all_dates_shift = all_dates

        passing_date = begin
        while passing_date <= end:
            date_str = passing_date.strftime("%Y-%m-%d")

            # Get the parent material's event values
            material_value = 0
            frozen_value = 0
            actual_value = 0
            if passing_date in all_dates_shift:
                index = all_dates_shift.index(passing_date)
                material_value = all_actual[index] if current_date >= all_dates[index] else all_frozen[index]
                frozen_value = all_frozen[index]
                actual_value = all_actual[index]


            # total_child_value = 0
            # for child in self.children:
            #     child_dates, child_frozen, child_actual = child.get_all_event()
            #     if passing_date in child_dates:
            #         child_index = child_dates.index(passing_date)
            #         # print('child_actual',child_actual,'child_frozen',child_frozen,'child_index',child_index)
            #         total_child_value += child_actual[child_index] if current_date >= child_dates[child_index] else child_frozen[child_index]

            total_child_value = 0
            child_records = []
            childs_name=[]
            for child in self.children:
                # print(str(child))
                child_dates, child_frozen, child_actual = child.get_all_event()
                childs_name.append(child.field_name)
                # print(childs_name,'childs_name')
                # print(child.code,child_dates, child_frozen, child_actual,'\n')
                ratio=calculate_ratio(self.field_name,child.field_name)

                if passing_date in child_dates:
                    child_index = child_dates.index(passing_date)

                    child_material_value = (
                        child_actual[child_index] if current_date > passing_date else child_frozen[child_index]
                    )
                    total_child_value += math.floor(child_material_value / ratio)
                    # print(self.code,total_child_value ,'date',current_date,'>=',child_dates[child_index])
                    child_records.append({
                        'child_code': child.code,
                        'child_name': child.field_name,
                        'frozen': child_frozen[child_index],
                        'actual': child_actual[child_index],
                        'ratio':ratio,

                    })
                    # childs_name.append(child.field_name)

            # Calculate new stock, factoring in both parent and child events
            new_stock = previous_stock + material_value - total_child_value
            # print(f'{self.code}\n{new_stock}= {previous_stock}+{material_value}-{total_child_value}')
            # Only store the stock when it changes

            if new_stock != previous_stock:
                if stock_changes:
                    stock_changes[-1]['date_end'] = (passing_date - timedelta(days=1)).strftime("%Y-%m-%d")



                stock_changes.append({
                    'date_start': date_str,
                    'date_end': date_str,
                    'previous_stock': previous_stock,
                    'stock': new_stock,
                    'code': self.code,
                    'name': self.field_name,
                    'frozen': frozen_value,
                    'actual': actual_value,
                    'child': child_records,  
                    'child_usage':total_child_value
                })
                # print(stock_changes)
                # Append the change to current_stock only when stock changes
                self.current_stock.append({
                    'date': date_str,
                    'stock': new_stock
                })

                previous_stock = new_stock

            passing_date += timedelta(days=1)

        if stock_changes:
            stock_changes[-1]['date_end'] = end.strftime("%Y-%m-%d")
        
        return stock_changes





    def calculate_stock(self, current_date, start_date, end_date):
        begin = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        stock_changes = []
        previous_stock = self.initial_stock
        last_recorded_date = begin

        # เอา Event F/S ทั้งหมดมา
        all_dates, all_frozen, all_actual = self.get_all_event() 

        # all_dates_shift คือ date event shift date ทั้งหมด
        if self.lead_time:
            all_dates_shift = [date + timedelta(days=self.lead_time) for date in all_dates]

        else:
            all_dates_shift = all_dates

        for date, frozen_value, actual_value in zip(all_dates, all_frozen, all_actual):
            # เพิ่ม Event F/S เข้า Calendar
            stock_changes.append({
                'date_start': date.strftime('%Y-%m-%d'),
                'date_end': date.strftime('%Y-%m-%d'),  
                'stock': 0,  
                'code': self.code,
                'frozen': frozen_value,
                'actual': actual_value,
                'child': len(self.children)
            })


        passing_date = begin
        while passing_date <= end:
            date_str = passing_date.strftime("%Y-%m-%d")

            # เลิอก F/S เทียบ current_date เทียบ วันของ Event ที่ shift
            material_value = 0
            if passing_date in all_dates_shift:
                index = all_dates_shift.index(passing_date)
                material_value = all_actual[index] if current_date >= all_dates[index] else all_frozen[index]

            # คิด Value event ของลูก
            total_child_value = 0
            for child in self.children:
                child_dates, child_frozen, child_actual = child.get_all_event()
                if passing_date in child_dates:
                    child_index = child_dates.index(passing_date)
                    total_child_value += child_actual[child_index] if current_date >= child_dates[child_index] else child_frozen[child_index]

            # เข้าสูตร
            new_stock = previous_stock + material_value - total_child_value

            # Record stock changes  
            if new_stock != self.current_stock:
                if stock_changes:
                    stock_changes[-1]['date_end'] = (passing_date - timedelta(days=1)).strftime("%Y-%m-%d")
                stock_changes.append({
                    'date_start': date_str,
                    'date_end': date_str,  
                    'stock': new_stock,
                    'code': self.code,
                    'frozen': 0,
                    'actual':0,
                    'child': len(self.children)
                })
                self.current_stock = new_stock
                last_recorded_date = passing_date

            previous_stock = new_stock

            # Move to the next day
            passing_date += timedelta(days=1)

        #สุด Range 
        if stock_changes:
            stock_changes[-1]['date_end'] = end.strftime("%Y-%m-%d")

        return stock_changes

import re

def extract_dimensions(text):
    match = re.search(r'(\d+(\.\d+)?)\s*x\s*(\d+(\.\d+)?)(\s*x\s*(\d+(\.\d+)?))?', text, re.IGNORECASE)

    if match:
        size = float(match.group(1))  
        width = float(match.group(3)) 
        depth = float(match.group(6)) if match.group(6) else None 
        return size, width, depth
    return None, None, None  # Return None if dimensions are not found

def calculate_ratio(parant, child):
    parant_size, parant_width, _ = extract_dimensions(parant)
    child_size, child_width, _ = extract_dimensions(child)
    
    if parant_size and child_size:  
        ratio = math.floor(parant_size / child_size)
        return ratio
    else:
        return 1


class get_machine(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            machines = get_nonduplicate_field_mc()
            print(machines)
            return Response({'success': True, 'data': machines})

        except Exception as e:
            print("ERROR >>>", e)
            return Response({'success': False, 'error': str(e)}, status=500)
        

# Get data when select ZCA
class get_Sche(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            current_date = datetime.now().date()
            zca_code = request.query_params.get('zca', None)
            plan_id = request.query_params.get('planid', None)

            start_date_str = request.query_params.get('start_date', None)
            end_date_str = request.query_params.get('end_date', None)


            if start_date_str and end_date_str:
                start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00')).date()
                end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')).date()
            else:
                start_date = current_date
                end_date = current_date


            # print(start_date,end_date)
            if zca_code:
                pipelines = PipeLineGetter(zca_code,plan_id)
            else:
                pipelines = []
            # print('pipelines',pipelines)
            materials = process_pipelines(pipelines,None,current_date,start_date,end_date)

            if pipelines:
                return self.build_response(materials, pipelines, current_date, start_date, end_date)
            else:
                return Response({'success': False, 'error': 'No pipelines found'}, status=404)

        except Exception as e:
            print("ERROR >>>", e)
            return Response({'success': False, 'error': str(e)}, status=500)

    def build_response(self, materials, pipelines, current_date,start_date,end_date):
        all_stock_changes = []
        if pipelines:
            root_zca = pipelines[0]['field_zca']
            root_material = materials[root_zca][0]  # first node

            root_material.set_hierarchy_level()
            hierarchy = root_material.to_hierarchy()
            machines = get_nonduplicate_field_process(pipelines)

            if root_zca in materials and materials[root_zca]:
                for mat in pipelines:
                    zca=mat['field_zca']
                    zca_material=materials[zca][0]
                    stock_changes = zca_material.calculate_stock(
                        current_date,
                        start_date.strftime('%Y-%m-%d'),
                        end_date.strftime('%Y-%m-%d')
                    )
                    all_stock_changes.extend(stock_changes)


                return Response({
                    'success': True,
                    'data': all_stock_changes,
                    'Chain_mat': hierarchy,
                    'Machines': machines
                })
            else:
                print(f"No root material found for {root_zca}")
                return Response({'success': False, 'error': 'No root material found'}, status=404)
        else:
            return Response({'success': False, 'error': 'No pipelines found'}, status=404)

def get_stock_change(materials, pipelines, current_date, start_date, end_date):
    all_stock_changes = []
    if pipelines:
        root_zca = pipelines[0]['field_zca']
        root_material = materials[root_zca][0]  # first node

        root_material.set_hierarchy_level()


        if root_zca in materials and materials[root_zca]:
            for mat in pipelines:
                zca = mat['field_zca']

                zca_material = materials[zca][0]

                stock_changes = zca_material.estimate_stock(
                    current_date,
                    start_date.strftime('%Y-%m-%d'),
                    end_date.strftime('%Y-%m-%d')
                )
                all_stock_changes.extend(stock_changes)

    # Sort all stock changes by 'date_start'
    all_stock_changes.sort(key=lambda x: x['date_start'])

    # Group stock changes by 'code'
    grouped_stock_changes = defaultdict(list)

    for change in all_stock_changes:
        zca_code = change['code']
        grouped_stock_changes[zca_code].append(change)
    



    return dict(grouped_stock_changes)


def FindPlanID(code):
    first_pid=ProcessLock.objects.filter(field_zca=code).values(
        'planid',
    ).first()

    return first_pid['planid']

# เอา ZCA ไปทำ Pipeline ออกมาเป็น List 
def PipeLineGetter(code, planid):
    # Step 1: Get the initial process entry for the provided ZCA and planid
    search_first = ProcessLock.objects.filter(field_zca=code, planid=planid).values(
        'field_id', 'field_zca', 'field_name', 'field_source', 'field_destination', 'field_mc', 'planid', 'order'
    ).first()


    if not search_first:
        return []  # Return an empty list if no matching records are found

    planid = search_first['planid']
    pipeline_entries = []

    # Step 2: Check if there are other ZCAs in the same planid with field_source as '1*'
    other_zcas_with_1star = ProcessLock.objects.filter(
        planid=planid, field_source='1*', disable=False
    ).values_list('field_zca', flat=True).distinct()
    

    # If there are such ZCAs, we need to search their related plan IDs
    if other_zcas_with_1star.exists():
        related_planids = ProcessLock.objects.filter(
            field_zca__in=other_zcas_with_1star, disable=False
        ).values_list('planid', flat=True).distinct()

        pipeline_entries = ProcessLock.objects.filter(
            planid__in=related_planids, disable=False
        ).values(
            'field_id', 'field_zca', 'field_name', 'field_source', 'field_destination', 'field_mc', 'planid','order'
        ).order_by('order')

    else:
        # If no such ZCAs exist, proceed with the regular pipeline search
        pipeline_entries = ProcessLock.objects.filter(
            planid=planid, disable=False
        ).values(
            'field_id', 'field_zca', 'field_name', 'field_source', 'field_destination', 'field_mc', 'planid','order'
        ).order_by('order')

    # Step 3: Use a dictionary to track unique combinations of field_zca and order to remove duplicates
    unique_entries = {}
    for entry in pipeline_entries:
        key = (entry['field_zca'], entry['order'])
        if key not in unique_entries:
            unique_entries[key] = entry

    # Convert the unique entries dictionary back into a list
    unique_pipeline = list(unique_entries.values())



    return unique_pipeline

# เอา List ไปเพิ่ม Class Material
def process_pipelines(pipelines, initial_stock_map, current_date, start_date, end_date):
    initial_stock_map = initial_stock_map or {}
    zca_list = [mat['field_zca'] for mat in pipelines]  # Get list of all ZCA codes

    # Batch fetch frozen and actual records
    all_frozen_records = get_frozen_records_bulk(zca_list, start_date, end_date)
    all_actual_records = get_actual_records_bulk(zca_list, start_date, end_date)

    # Group frozen and actual records by ZCA
    frozen_map = defaultdict(list)
    actual_map = defaultdict(list)

    for rec in all_frozen_records:
        frozen_map[rec['materialcode']].append(rec)

    for rec in all_actual_records:
        actual_map[rec['zca_on']].append(rec)

    materials = {}
    # print(frozen_map,'frozen_map')
    for mat in pipelines:
        zca = mat['field_zca']
        field_name = mat['field_name']

        init_stock = initial_stock_map.get(zca, 0)

        # Initialize Material instance with the correct stock
        if zca not in materials:
            materials[zca] = [Material(
                code=zca,
                field_name=field_name,
                initial_stock=init_stock,
                lead_time=4 if mat['field_source'] == '1' else 0
            )]

        # Process frozen records
        for frozen_record in frozen_map[zca]:
            # print(zca,frozen_record)
            starttime = frozen_record['starttime'].date() if isinstance(frozen_record['starttime'], datetime) else frozen_record['starttime']
            frozen_value = int(frozen_record['plancount'])

            # If the date already exists, sum the value
            if starttime in materials[zca][0].dates:
                index = materials[zca][0].dates.index(starttime)
                materials[zca][0].frozen[index] += frozen_value
            else:
                materials[zca][0].add_frozen_value(starttime, frozen_value)

        # Process actual records
        for actual_record in actual_map[zca]:
            if actual_record['qty_good'] is None:
                continue  # Skip this record if 'goodcount' is None

            starttime = actual_record['starttime'].date() if isinstance(actual_record['product_date'], datetime) else actual_record['product_date']
            actual_value = int(actual_record['qty_good'])

            # If the date already exists, sum the value
            if starttime in materials[zca][0].dates:
                index = materials[zca][0].dates.index(starttime)
                materials[zca][0].actual[index] += actual_value
            else:
                materials[zca][0].add_actual_value(starttime, actual_value)

    # Build material hierarchy
    build_material_hierarchy(materials, pipelines)

    return materials


# เอา zca ไปหา [frozen , date]
def get_frozen_records_bulk(zca_list, start_date, end_date):
    return Tempactiveplan.objects.filter(
        materialcode__in=zca_list,
        starttime__gte=start_date,
        starttime__lte=end_date
    ).values('planname', 'materialcode', 'plancount', 'starttime', 'shift')

def get_actual_records_bulk(zca_list, start_date, end_date):
    return ViewWmsListfillplanproduction.objects.filter(
        zca_on__in=zca_list,
        product_date__gte=start_date,
        product_date__lte=end_date
    ).exclude(qty_good__isnull=True).values(
        'zca_on', 'qty_good', 'product_date'
    )

# ปั้น Tree
def build_material_hierarchy(materials, pipelines):
    # {zca : Mat instance}
    material_map = {pipeline['field_zca']: materials[pipeline['field_zca']] for pipeline in pipelines if pipeline['field_zca'] in materials}
    
    # Find the first duplicate order, which indicates where branching should occur
    branching = find_first_duplicate_order(pipelines)

    parent_material = None

    remaining_pipelines = pipelines.copy()  

    prev_planid = None 
    node_branch = None  
    leaf_branch = None  

    for pipeline in pipelines:
        current_material_code = pipeline['field_zca']
        current_material_list = material_map.get(current_material_code, [])
        
        if not current_material_list:
            print(f"No materials found for {current_material_code}")
            continue
        
        current_material = current_material_list[0]

        if pipeline['order'] + 1 == branching:
            node_branch = current_material

        if pipeline['order'] >= branching:  # Branching point
            curr_pid = pipeline['planid']
            
            # Check if planid has changed
            if prev_planid is not None and curr_pid != prev_planid:
                leaf_branch = node_branch  # Update leaf_branch to node_branch on planid change

            prev_planid = curr_pid  
            
            # Process pipelines with the same planid
            for bline in remaining_pipelines[:]:  # Iterate over a copy of the list to avoid modifying the list while iterating
                if bline['planid'] == curr_pid and bline['order'] >= branching:
                    bline_material_code = bline['field_zca']
                    bline_material_list = material_map.get(bline_material_code, [])

                    
                    if not bline_material_list:
                        print(f"No materials found for {bline_material_code}")
                        continue
                    
                    bline_material = bline_material_list[0]
                    if bline['field_destination'] == 'NULL' or bline['field_destination'] == '':
                        if leaf_branch is not None:
                            # print(leaf_branch.code,'add',bline_material.code)
                            leaf_branch.add_child(bline_material)

                        leaf_branch = bline_material
                    elif bline['field_destination'] == "*":
                        if leaf_branch is not None:
                            # print(leaf_branch.code,'add',bline_material.code)
                            leaf_branch.add_child(bline_material)

                    else:
                        leaf_branch = bline_material

                    # Remove processed bline from remaining_pipelines
                    remaining_pipelines.remove(bline)

        else:  # Normal processing before branching
            if pipeline['field_destination'] == 'NULL' or pipeline['field_destination'] == '':
                if parent_material is not None:
                    # print(parent_material.code,'add',current_material.code)
                    parent_material.add_child(current_material)
                parent_material = current_material
                leaf_branch = current_material  # Update leaf_branch
            elif pipeline['field_destination'] == "*":
                if parent_material is not None:
                    # print(parent_material.code,'add',current_material.code)
                    parent_material.add_child(current_material)
            else:
                parent_material = current_material
                leaf_branch = current_material  # Update leaf_branch
    
    return parent_material

# กรณี Branching 
def find_first_duplicate_order(pipelines):
    order_set = set()
    for pipeline in pipelines:
        order = pipeline.get("order")
        if order in order_set:
            return order
        order_set.add(order)
    return 9999




# แสดง Option ZCA ทั้งหมดที่มีอยู่ใน SKUs
class GetOptions(APIView):
    permission_classes = [ ScmPerm|ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            # Fetch field_zca, field_name, and planid from the ProcessLock model
            options = ProcessLock.objects.all().values('field_zca', 'field_name', 'planid')

            # Construct the response data with value, label, and planid
            response_data = [
                {
                    "value": option["field_zca"], 
                    "label": f"{option['field_zca']} - {option['field_name']}",
                    "planid": option["planid"]
                } 
                for option in options
            ]

            return Response({'success': True, 'options': response_data})
        except Exception as e:
            print("ERROR >>>", e)
            return Response({'success': False, 'error': str(e)}, status=500)


# เอา Machines ที่ไม่ซ้ำไปทำ TABS เลือก

def get_nonduplicate_field_process(data):
    # Extract 'field_zca' values from the input data
    field_zca_values = [item['field_zca'] for item in data]
    
    # Query the Masterwipstk model to filter by 'zca' values in the extracted list
    master_zca_records = Masterwipstk.objects.filter(zca__in=field_zca_values).values('id', 'zca', 'hs', 'name_th', 'stk_p_hr')
    
    # Create a dictionary to hold the results
    zca_hs_dict = {}
    
    # Populate the dictionary with 'zca' as keys and lists of unique 'hs' values as values
    for record in master_zca_records:
        zca = record['zca']
        hs = record['hs']
        if zca not in zca_hs_dict:
            zca_hs_dict[zca] = []
        if hs not in zca_hs_dict[zca]:
            zca_hs_dict[zca].append(hs)
    
    return zca_hs_dict



def get_nonduplicate_field_mc():
    master_zca_records = Masterwipstk.objects.values('id', 'zca', 'hs', 'name_th', 'stk_p_hr')
    zca_hs_dict = {}
    for record in master_zca_records:
        zca = record['zca']
        hs = record['hs']
        if zca not in zca_hs_dict:
            zca_hs_dict[zca] = []
        if hs not in zca_hs_dict[zca]:
            zca_hs_dict[zca].append(hs)
    
    return zca_hs_dict




#  <======================= Stock <=======================>

def get_unique_lookup(data):
    lookup = {}
    unique_data = {}

    for item in data:
        key = f"{item['mapid']}-{item['level']}-{item['sub_column']}"
        # print("Unique Key Generated:", key)  # Debugging log

        if key not in unique_data:
            unique_data[key] = True # That key exist
            lookup[key] = item

    return lookup

def get_today_stock(zca_on_list, unique_lookup, PlanId):

    total_qty_sum = 0  
    zca_on_data = defaultdict(lambda: defaultdict(lambda: {'total_qty': 0, 'levels': defaultdict(int)})) 

    # Filter unique_lookup by zca_on_list
    for item in unique_lookup.values():
        zca_on = item['zca_on']
        
        # Filter the items by checking if zca_on is in zca_on_list
        if zca_on not in zca_on_list:
            continue  # Skip if zca_on is not in the list
        
        key = (item['warehouse_id'], item['zone'], item['row'], item['column'], item['sub_column'])

        qty_value = item.get('qty', 0)
        if qty_value is None:
            qty_value = 0
        else:
            qty_value = int(qty_value)
        
        if item['level'] is None:
            pass
            # print(f"Found item with level None: {item}")
        
        # Accumulate total quantity and levels
        zca_on_data[zca_on][key]['total_qty'] += qty_value
        zca_on_data[zca_on][key]['levels'][item['level']] += qty_value
        
        total_qty_sum += qty_value

    response_data = []

    for zca_on, zone_row_column_data in zca_on_data.items():
        # Handle None values by using a placeholder for sorting
        sorted_keys = sorted(
            zone_row_column_data.keys(),
            key=lambda k: (
                str(k[0]) if k[0] is not None else '',  
                int(k[1]) if k[1] is not None and isinstance(k[1], int) else -1,  
                int(k[2]) if k[2] is not None and isinstance(k[2], int) else -1,
                int(k[3]) if k[3] is not None and isinstance(k[3], int) else -1
            )
        )

        zca_response = {
            'zca_on': zca_on,
            'PlanId': PlanId,
            'data_by_zone_row_column': [
                {
                    'warehouse': key[0],
                    'zone': key[1],
                    'row': key[2],
                    'column': key[3],
                    'total_qty': zone_row_column_data[key]['total_qty'],
                    'levels': dict(zone_row_column_data[key]['levels'])  
                }
                for key in sorted_keys  
            ],
            'total_qty': sum([zone_row_column_data[key]['total_qty'] for key in sorted_keys])  
        }

        response_data.append(zca_response)

    return response_data


def get_all_today_stock():
    total_qty_sum = 0
    queryset = ViewWmsMapManagement.objects.values().order_by('-created_at')
    unique_lookup = get_unique_lookup(queryset)


    zca_on_data = defaultdict(lambda: defaultdict(lambda: {'total_qty': 0, 'levels': defaultdict(int)}))
    zca_on_names = {}  
    
    for item in unique_lookup.values():
        zca_on = item['zca_on']
        name_th = item['name_th']
        

        if zca_on not in zca_on_names:
            zca_on_names[zca_on] = name_th
        
        key = (item['warehouse_id'], item['zone'], item['row'], item['column'], item['sub_column'])
        
        qty_value = item.get('qty', 0)
        qty_value = int(qty_value) if qty_value is not None else 0


        pcsperpallet = item.get('pcsperpallet', 1)
        pcsperpallet = int(pcsperpallet) if pcsperpallet is not None else 1

        if item['level'] is None:
            pass
            # print(f"Found item with level None: {item}")

        zca_on_data[zca_on][key]['total_qty'] += qty_value
        zca_on_data[zca_on][key]['levels'][item['level']] += qty_value
        
        total_qty_sum += qty_value

    response_data = []

    for zca_on, zone_row_column_data in zca_on_data.items():

        sorted_keys = sorted(
            zone_row_column_data.keys(),
            key=lambda k: (
                str(k[0]) if k[0] is not None else '',
                int(k[1]) if k[1] is not None and isinstance(k[1], int) else -1,
                int(k[2]) if k[2] is not None and isinstance(k[2], int) else -1,
                int(k[3]) if k[3] is not None and isinstance(k[3], int) else -1
            )
        )

        # Get the stored name_th for this zca_on
        name_th = zca_on_names.get(zca_on, 'N/A')  # Fallback in case name_th is missing

        # Calculate total_qty for each zca_on
        total_qty = sum([zone_row_column_data[key]['total_qty'] for key in sorted_keys])


        zca_response = {
            'zca_on': zca_on,
            'name': name_th,  # Correct name_th is now retrieved
            'data_by_zone_row_column': [
                {
                    'warehouse': key[0],
                    'zone': key[1],
                    'row': key[2],
                    'column': key[3],
                    'total_qty': zone_row_column_data[key]['total_qty'],
                    'levels': dict(zone_row_column_data[key]['levels'])
                }
                for key in sorted_keys
            ],
            'total_qty': total_qty,
            'pallet':total_qty/pcsperpallet
        }

        response_data.append(zca_response)
    
    return response_data




def GetWipProcess(process):
    Wip_list = []
    for mat in process:
        if mat['field_destination'] != '*':
            Wip_list.append(mat['field_zca'])
    return Wip_list

def GetFirstZca(planid):
    first_zca = ProcessLock.objects.filter(planid=planid).values('field_zca').first()
    return first_zca['field_zca']

def GetDistinctPlanID():
    # Get distinct planid values from the ProcessLock model, ordered by planid
    distinct_plan_ids = ProcessLock.objects.values_list('planid', flat=True).distinct().order_by('planid')
    
    # Convert the queryset to a list
    planid_list = list(distinct_plan_ids)
    
    return planid_list


class get_stockestimate(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            # To store the total time for each step (initialize with 0)
            timings = {
                'GetDistinctPlanID': 0,
                'GetFirstZca': 0,
                'PipeLineGetter': 0,
                'GetWipProcess': 0,
                'get_today_stock': 0,
                'process_pipelines': 0,
                'get_stock_change': 0,
                'total_execution_time': 0
            }

            # Capture the total start time
            start_time = timer.time()

            # Step 1: Get PlanId and process list
            forcast = []
            temp=[]
            step_start = timer.time()  # Start timing for GetDistinctPlanID
            planid_list = GetDistinctPlanID()
            timings['GetDistinctPlanID'] += timer.time() - step_start  # Accumulate time

            queryset = ViewWmsMapManagement.objects.values().order_by('-created_at')
            unique_lookup = get_unique_lookup(queryset)
            # Simulate the current date
            # current_date = datetime(2024, 8, 20).date()
            current_date = datetime.now().date()
            # current_date = date(2024, 10, 22)
            print('current_date',current_date)

            # Iterate over each PlanId
            for PlanId in planid_list:

                # Step 2: Get the first ZCA for the PlanId
                step_start = timer.time()
                first_zca = GetFirstZca(PlanId)
                timings['GetFirstZca'] += timer.time() - step_start  # Accumulate time

                # Step 3: Get ZCA on list through the pipeline getter
                step_start = timer.time()
                zca_on_list = PipeLineGetter(first_zca, PlanId)
                timings['PipeLineGetter'] += timer.time() - step_start  # Accumulate time

                # Step 4: Get WIP process information
                step_start = timer.time()
                zca_on_wip = GetWipProcess(zca_on_list)
                timings['GetWipProcess'] += timer.time() - step_start  # Accumulate time

                # Step 5: Get the initial stock data for the ZCAs
                step_start = timer.time()
                initial_stock_data = get_today_stock(zca_on_wip,unique_lookup, PlanId)
                timings['get_today_stock'] += timer.time() - step_start  # Accumulate time

                # Map stock data

                initial_stock_map = {item['zca_on']: item['total_qty'] for item in initial_stock_data}
                location_stock = {item['zca_on']: item['data_by_zone_row_column'] for item in initial_stock_data}
                temp.append(initial_stock_map)
                # Step 6: Process the pipelines
                step_start = timer.time()
                shift_week = current_date + timedelta(weeks=1)
                materials = process_pipelines(zca_on_list, initial_stock_map, current_date, current_date, shift_week)
                timings['process_pipelines'] += timer.time() - step_start  # Accumulate time

                # Step 7: Get stock change estimates
                step_start = timer.time()
                stock_estimate = get_stock_change(materials, zca_on_list, current_date, current_date, shift_week)
                timings['get_stock_change'] += timer.time() - step_start  # Accumulate time

                # Step 8: Build hierarchy from materials
                if len(zca_on_list) > 0:
                    root_zca = zca_on_list[0]['field_zca']
                    root_material = materials[root_zca][0]  # First node
                    hierarchy = root_material.to_hierarchy()

                # Store forecast data
                if stock_estimate:
                    final_response = {
                        'PlanId': PlanId,
                        'data': stock_estimate,
                        'hierarchy': hierarchy,
                        'location': location_stock
                    }
                    forcast.append(final_response)

            # Calculate total execution time for the entire function
            total_execution_time = timer.time() - start_time
            timings['total_execution_time'] = total_execution_time 

            # Calculate the percentage of time spent on each function
            timing_percentages = {}
            for key, value in timings.items():
                if key != 'total_execution_time': 
                    timing_percentages[key] = (value / total_execution_time) * 100 if total_execution_time > 0 else 0


            # stock=get_stock()
            return Response({
                'forecast': forcast,
                'timings': timings,
                'percentages': timing_percentages,
                'initial_stock_map':temp,
            })

        except Exception as e:
            traceback.print_exc()  # Print full stack trace in server logs for debugging
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from django.utils import timezone


class get_stock_location(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            today = timezone.now().date()
            all_stock_location = get_all_today_stock()
            today = datetime.today()
            start_of_30_days = today - timedelta(days=30)


            history_stock_queryset = StockHistory.objects.filter(
                    stock_date__gte=start_of_30_days,
                    stock_date__lt=today
                ).exclude(cur_stock=0)


            formatted_history_stock = defaultdict(list)
            for record in history_stock_queryset.values('name_th', 'stock_date', 'cur_stock', 'zca'):
                key = f"{record['zca']} - {record['name_th']}"
                print(key)
                if key not in formatted_history_stock:
                    formatted_history_stock[key] = []
                
                formatted_history_stock[key].append({
                    "x": record['stock_date'].strftime("%Y-%m-%d"),
                    "y": record['cur_stock']
                })



            history_stock = [{"name": name, "data": data} for name, data in formatted_history_stock.items()]
            # print(history_stock)
            return Response({
                'history_stock': history_stock,
                'all_stock_location': all_stock_location,
            })

        except Exception as e:
            traceback.print_exc()  # Print full stack trace in server logs for debugging
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



################################################################
# 
# MACHINE PLAN
# 
################################################################

def calculate_stk(phr,hr):
    result=(int(phr))*(int(hr))
    return result


# Post Machine Plan เข้า Data
class post_machplan(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            request_input = request.data
            machine = request_input.get('priorityId')
            zca_name = request_input.get('text')
            start_str = request_input.get('startDate')
            end_str = request_input.get('endDate')
            type_str = request_input.get('type')
            user_str=request_input.get('username')
            if not machine or not zca_name or not start_str or not end_str:
                return Response({'success': False, 'error': 'Missing required fields'}, status=status.HTTP_400_BAD_REQUEST)

            start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
            duration = (end - start).total_seconds() / 3600


            if 'zca' in zca_name.lower():
                # Filter Masterwipstk objects by zca_name and machine
                filtered_object = Masterwipstk.objects.filter(
                    zca__icontains=zca_name, hs=machine
                ).values('zca', 'stk_p_hr', 'name_th', 'name_en', 'ton_p_hr').first()

                if filtered_object:
                    ton_p_hr = filtered_object['ton_p_hr']
                    stk_p_hr = filtered_object['stk_p_hr']
                    zca = filtered_object['zca']
                    name_th = filtered_object['name_th']
                    total = round(float(stk_p_hr)) * int(duration)
                    total_ton = round(float(ton_p_hr) * int(duration))

                    save_data = FrozenMat(
                        zca=zca,
                        th_name=name_th,
                        stk_frozen=total,
                        ton=total_ton,
                        date_start=start,
                        date_end=end,
                        machine=machine,
                        type=type_str,
                        planner=user_str,
                    )
                    save_data.save()


                    return Response({'success': True, 'calculated_stock': total})
            
            elif zca_name == 'Cleaning':
                save_data = FrozenMat(
                    zca=zca_name,
                    th_name=zca_name,
                    stk_frozen=0,
                    ton=0,
                    date_start=start,
                    date_end=end,
                    machine=machine,
                    type='constrain',
                )
                save_data.save()

                return Response({'success': True, 'calculated_stock': None})  # Adjusted to None or other suitable value
            
            else:
                return Response({'success': False, 'error': 'Incorrect ZCA or Machine'}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            print(f"Error: {e}")  # Print the error to the console
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Get Machine Plan จาก Database        
class GetAppointments(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            machines = request.query_params.getlist('machines[]')
            if machines:
                appointments = Tempactiveplan.objects.filter(
                    machine__in=machines, 
                ).values(
                    'planname', 'materialcode', 'materialname', 'plancount', 'planweight', 'starttime', 'duration', 'machine', 'shift','versionno','created_by'
                )

                appointment_list = []
                for appointment in appointments:
                    zca_data=get_zca_details(appointment.get('materialcode'))

                    shift = appointment.get('shift')
                    duration = float(appointment.get('duration', 0)) / 60  
                    if not appointment['materialname']:
                        appointment['materialname']=zca_data.get('name_th')

                    if isinstance(appointment['starttime'], datetime):
                        start_date = appointment['starttime'].date()  
                    else:
                        start_date = appointment['starttime']  
                    
                    # Set start time based on the shift
                    if shift == 'A':
                        start_time = datetime.combine(start_date, datetime.strptime("08:00", "%H:%M").time())
                    elif shift == 'B':
                        start_time = datetime.combine(start_date, datetime.strptime("16:00", "%H:%M").time())
                    elif shift == 'C':
                        start_time = datetime.combine(start_date, datetime.strptime("00:00", "%H:%M").time())
                    else:
                        start_time = appointment['starttime']

                    # Calculate the end time by adding the duration
                    date_end = start_time + timedelta(hours=duration)


                    appointment['date_start'] = start_time.isoformat()  
                    appointment['date_end'] = date_end.isoformat()      
                    appointment['planner'] = appointment['created_by']
                    appointment['versionno'] = appointment['versionno']
                    appointment['type'] = None

                    appointment_list.append(appointment)
                print(appointment_list)
                return Response({'success': True, 'appointments': appointment_list})
            else:
                return Response({'success': True, 'appointments': []})
        except Exception as e:
            traceback.print_exc()  # This will print the full stack trace in the server logs
            return Response({'success': False, 'error': str(e)}, status=500)
        




        
# Del Machine Plan จาก Database  
class DeleteAppointment(APIView):
    permission_classes = [PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        appointment_id = request_input.get('id')
        if not appointment_id:
            return Response({'success': False, 'error': 'Appointment ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            appointment = FrozenMat.objects.get(id_frozen=appointment_id)
            appointment.delete()
            return Response({'success': True}, status=status.HTTP_200_OK)
        except FrozenMat.DoesNotExist:
            return Response({'success': False, 'error': 'Appointment not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# Edit Machine Plan เข้า Data
class EditAppointment(APIView):
    permission_classes = [PlannerPerm|PlantCoPerm| AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        request_input = request.data
        print(request_input,'request_input')
        appointment_id = request_input.get('id')
        if not appointment_id:
            return Response({'success': False, 'error': 'Appointment ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            id_zca=request_input['id']
            zca=request_input['text']
            start_str=request_input['startDate']
            end_str=request_input['endDate']
            machine=request_input['priorityId']
            th_name=request_input['th_name']
            planner=request_input['username']
            start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
            end = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
            duration = (end - start).total_seconds() / 3600
            filtered_object = Masterwipstk.objects.filter(zca=zca,hs=machine).values('zca', 'stk_p_hr','ton_p_hr').first()
            stk_p_hr = filtered_object['stk_p_hr']
            ton_p_hr=filtered_object['ton_p_hr']
            if stk_p_hr:
                total_stk = int(stk_p_hr) * int(duration)
                total_ton = float(ton_p_hr) * int(duration)
            else:
                total_stk=0
                total_ton=0

            save_data = FrozenMat(
                id_frozen=id_zca,
                zca=zca,
                th_name=th_name,
                stk_frozen=total_stk, 
                ton=total_ton,
                date_start=start,
                date_end=end,
                machine=machine,
                planner=planner,
            )
            save_data.save()
            return Response({'success': True}, status=status.HTTP_200_OK)
        except CheckWip.DoesNotExist:
            return Response({'success': False, 'error': 'Appointment not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print('error',e)
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class GetDemand(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            demand = Demand.objects.all().values('class_field', 'materialcode', 'nameth', 'groupdetail', 'kg_per_pcs', 'pcs_pal', 'sp_short')
            return Response({'success': True, 'demand': demand})
        except Exception as e:
            print("ERROR >>>", e)
            return Response({'success': False, 'error': str(e)}, status=500)
        






class GetWorkload(APIView):
    permission_classes = [AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            plans = MapListWithdrawPlan.objects.filter(withdraw_success=1).values()
            plans2 = ViewWmsListfillplanproduction.objects.filter(approve_fill='success').values()
            print(len(plans))
            return Response({'success': True, 'workload': plans , 'workload2':plans2})
        except Exception as e:
            print("ERROR >>>", e)
            return Response({'success': False, 'error': str(e)}, status=500)


class GetZcasView(APIView):
    permission_classes = [AdminPerm]
    def get(self, request, *args, **kwargs):
        materials = MaterialModel.objects.all()
        serializer = MaterialSerializer(materials, many=True)

        return Response({'success': True, 'zcas': serializer.data})
    
class MaterialCreateView(APIView):
    permission_classes = [AdminPerm]

    def post(self, request, *args, **kwargs):
        print(f"Received data: {request.data}")

        parent_id = request.data.get('parent_id')
        child_id = request.data.get('child_id')

        # Validate the parent and child IDs exist in the MaterialModel
        try:
            parent = MaterialModel.objects.get(materialid=parent_id)
            child = MaterialModel.objects.get(materialid=child_id)
        except MaterialModel.DoesNotExist:
            return Response({'success': False, 'error': 'Parent or Child Material not found'}, status=status.HTTP_400_BAD_REQUEST)

        # Prepare the data for the serializer
        data = {
            'parentid': parent.materialid,
            'childid': child.materialid
        }
        material_process_serializer = MaterialProcessSerializer(data=data)

        if material_process_serializer.is_valid():
            material_process_serializer.save()
            return Response({'success': True, 'material_process': material_process_serializer.data}, status=status.HTTP_201_CREATED)
        else:
            print(f"MaterialProcess Serializer errors: {material_process_serializer.errors}")
            return Response({'success': False, 'errors': material_process_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class GetZCAByName(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        zca_name = request.query_params.get('zca', None)
        if zca_name:
            try:
                zca = MaterialModel.objects.get(zca=zca_name)
                return Response({'success': True, 'zca': {'materialid': zca.materialid, 'zca': zca.zca}})
            except MaterialModel.DoesNotExist:
                return Response({'success': False, 'error': 'ZCA not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': False, 'error': 'ZCA name not provided'}, status=status.HTTP_400_BAD_REQUEST)



#  <======================= PIS =================================>


def get_week_range(week,year):

    quart = int(week) / 4
    month = math.ceil(quart) 

    current_year = year

    if quart % 1 == 0.25:  
        week_start_date = datetime(current_year, month, 1)
        week_end_date = datetime(current_year, month, 7)

    elif quart % 1 == 0.5:  
        week_start_date = datetime(current_year, month, 8)
        week_end_date = datetime(current_year, month, 14)

    elif quart % 1 == 0.75:  
        week_start_date = datetime(current_year, month, 15)
        week_end_date = datetime(current_year, month, 21)

    elif quart % 1 == 0: 
        week_start_date = datetime(current_year, month, 22)
        last_day_of_month = calendar.monthrange(current_year, month)[1]
        week_end_date = datetime(current_year, month, last_day_of_month)

    return (week_start_date, week_end_date)



def get_week_number(input_date):
    week_count = 0

    otherweeks = (input_date.month -1)*4
    week_count+=otherweeks
    if input_date.day < 8:
        week_count += 1

    elif input_date.day < 15:
        week_count += 2

    elif input_date.day < 22:
        week_count += 3

    elif input_date.day >= 22:
        week_count += 4

    return week_count

class GetWeeks(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            # Get the current date or a date from request
            today = date.today()

            week_num = get_week_number(today)
            print(week_num)
            return Response({
                'success': True, 
                'week_number': week_num,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetWeeksRange(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            # Get the current date or a date from request
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')

            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_date += timedelta(days=1)
            end_date += timedelta(days=1)
            week_num_start = get_week_number(start_date)
            week_num_end = get_week_number(end_date)

            print(end_date,week_num_end)
            print(start_date,week_num_start)
            return Response({
                'success': True, 
                'week_number': [week_num_start,week_num_end],
            }, status=status.HTTP_200_OK)

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def get_zca_details(zca):
    pcsperpallet = 1
    name_th = ''
    kgpcs =  1
    if "ZCAW" in zca:
        product = ViewItemmasterproductwip.objects.filter(field_zca=zca).first()
        if product:
            pcsperpallet = int(product.pcsperpallet)
            name_th = product.field_name
            kgpcs =  product.field_kgpcs
    elif "ZCA" in zca and "ZCAW" not in zca:
        product = ViewItemmasterproductfg.objects.filter(zca=zca).first()
        if product:
            pcsperpallet =  int(product.pcpallet)
            name_th =  product.name
            kgpcs =  product.kg

        

    return {'name_th':name_th,'pcsperpallet':pcsperpallet,'kgpcs':kgpcs}

class get_remain_plan(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            
            # Retrieve start_date and end_date from query parameters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            if start_date:
                start_year = datetime.strptime(start_date, '%Y-%m-%d').year
            if end_date:
                end_year = datetime.strptime(end_date, '%Y-%m-%d').year
            count_finish = 0
            count_unfinish = 0
            ton_done = ton_undone = 0
            total_pallet_done = total_pallet_undone = 0
            machine_list = []

            machine_done = {}
            machine_undone = {}




            if not start_date or not end_date:
                return Response({
                    'success': False,
                    'message': 'Both start_date and end_date are required.'
                }, status=status.HTTP_400_BAD_REQUEST)


            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')


            start_date += timedelta(days=1)
            end_date += timedelta(days=2, seconds=-1)
            start_week =get_week_number(start_date)
            end_week =get_week_number(end_date)
            print(type(start_year),'start_year')
            start_yw = f"{start_year}_{start_week}"
            end_yw = f"{end_year}_{end_week}"

            weeklist=[]



            #ดู maxversion ของแต่ละ week        
            plan_data_wk = Tempactiveplan.objects.filter(
                planweek__gte=start_yw,
                planweek__lte=end_yw,
            ).values('planweek').annotate(latest_version=Max('versionno'))


            #maxversion
            latest_versions = []
            for wk in plan_data_wk:
                latest_versions.append(wk['latest_version'])

            # เอาแผนที่มีอยู่ใน week 
            plan_data = Tempactiveplan.objects.filter(
                planweek__gte=start_yw,
                planweek__lte=end_yw,
                versionno__in=latest_versions
            ).values('materialcode', 'machine', 'planweek' ,'versionno').annotate(
                total_plancount=Sum('plancount')
            )

            #maxversion {y_w:version}
            latest_versions_map = {wk['planweek']: wk['latest_version'] for wk in plan_data_wk}


            #รายการส่งยอดทั้งหมด
            done_data = ViewWmsListfillplanproduction.objects.filter(
                product_date__gte=start_date,
                product_date__lte=end_date,
                approve_fill='success'
            ).values(
                'zca_on', 'machine', 'pcsperpallet', 'name_th', 'kgpcs', 'product_date', 'qty_good', 'qty_loss', 'qty_lab'
            )

            #รวมยอดทั้งหมด Group by sku , machine , year_week
            grouped_done_data = defaultdict(lambda: {'total_goodcount': 0, 'total_loss': 0, 'total_lab': 0})
            for item in done_data:
                zca = item['zca_on']
                machine = item['machine']
                product_year=item['product_date'].year
                week_num = get_week_number(item['product_date'])
                week_number=f"{product_year}_{week_num}"
                key = (zca, machine, week_number)

                grouped_done_data[key]['total_goodcount'] += item['qty_good'] or 0
                grouped_done_data[key]['total_loss'] += item['qty_loss'] or 0
                grouped_done_data[key]['total_lab'] += item['qty_lab'] or 0
                grouped_done_data[key]['pcsperpallet'] = item['pcsperpallet']
                grouped_done_data[key]['name_th'] = item['name_th']
                grouped_done_data[key]['kgpcs'] = item['kgpcs']

 
            done_dict = {(key[0], key[1],key[2]): {
                'total_goodcount': value['total_goodcount'],
                'total_loss': value['total_loss'],
                'pcsperpallet': value['pcsperpallet'],
                'name_th': value['name_th'],
                'kgpcs': value['kgpcs'],
            } for key, value in grouped_done_data.items()}


            # PVP
            result_data = []
            # สำหนรับการเปลี่ยนหน่วย
            machine_done_ton = {}
            machine_undone_ton = {}
            machine_done_pallet = {}
            machine_undone_pallet = {}



            # PVP
            for plan in plan_data:
                planweek = plan['planweek']
                versionno = plan['versionno']

                if planweek in latest_versions_map and versionno == latest_versions_map[planweek]:
                    listmc_done=[]
                    zca = plan['materialcode']
                    plancount = plan['total_plancount'] or 0  
                    machine_plan = plan['machine']
                    listmc_done.append(machine_plan)
                    done_entry = done_dict.pop((zca, machine_plan,planweek), {})

                    goodcount = done_entry.get('total_goodcount', 0)
                    losscount = done_entry.get('total_loss', 0)
                    pcsperpallet = done_entry.get('pcsperpallet')
                    name_th = done_entry.get('name_th')
                    kgpcs = done_entry.get('kgpcs', 0)  

                    # หา ชื่อไทย / pcs ถ้า ไม่มีใน query ส่งยอด
                    if not pcsperpallet or not name_th:
                        if "ZCAW" in zca:
                            product = ViewItemmasterproductwip.objects.filter(field_zca=zca).first()
                            if product:
                                pcsperpallet = pcsperpallet or int(product.pcsperpallet)
                                name_th = name_th or product.field_name
                                kgpcs = kgpcs or product.field_kgpcs
                        elif "ZCA" in zca and "ZCAW" not in zca:
                            product = ViewItemmasterproductfg.objects.filter(zca=zca).first()
                            if product:
                                pcsperpallet = pcsperpallet or int(product.pcpallet)
                                name_th = name_th or product.name
                                kgpcs = kgpcs or product.kg


                    kgpcs = float(kgpcs) if kgpcs else 0

                    # Calculate remaining counts and pallets
                    remain_ton_t = remain = plancount - goodcount
                    remain_pallet_t=remain_pallet = round(remain / pcsperpallet, 2) if pcsperpallet else 1
                    plan_pallet = round(plancount / pcsperpallet, 2) if pcsperpallet else 1
                    done_pallet = round(goodcount / pcsperpallet, 2) if pcsperpallet else 1
                    losscount_plt= round(losscount / pcsperpallet, 2) if pcsperpallet else 1
                    if remain_ton_t <0 or remain_pallet_t <0:
                        remain_ton_t=0
                        remain_pallet_t=0

                    remain_ton = round(((int(remain_ton_t) * kgpcs) / 1000), 2)
                    plan_ton = round(((int(plancount) * kgpcs) / 1000), 2)
                    done_ton = round(((int(goodcount) * kgpcs) / 1000), 2)
                    losscount_ton= round(((int(losscount) * kgpcs) / 1000), 2)

                    ton_done += done_ton
                    ton_undone += remain_ton

                    total_pallet_done += done_pallet
                    total_pallet_undone = float(total_pallet_undone) + float(remain_pallet_t)

                    machine_undone_pallet[machine_plan] = float(machine_undone_pallet.get(machine_plan, 0)) + float(remain_pallet_t)
                    machine_done_pallet[machine_plan] = machine_done_pallet.get(machine_plan, 0) + done_pallet

                    machine_undone_ton[machine_plan] = machine_undone_ton.get(machine_plan, 0) + remain_ton
                    machine_done_ton[machine_plan] = machine_done_ton.get(machine_plan, 0) + done_ton

                    machine_list.append(machine_plan)

                    if remain <= 0:
                        machine_done.setdefault(machine_plan, []).append(zca)
                        count_finish += 1
                    else:
                        machine_undone.setdefault(machine_plan, []).append(zca)
                        count_unfinish += 1

                    # รวมยอด PVP แบบไม่นับแยกเครื่อง
                    result_data.append({
                        'materialcode': zca,
                        'name_th': name_th,

                        'planweek':planweek,
                        'pcsperpallet': pcsperpallet,
                        'kgpcs':kgpcs,
                        'listmc_done':listmc_done,

                        'plan': plancount,
                        'done': goodcount,
                        'remain': remain,


                        'remain_pallet': remain_pallet,
                        'done_pallet': done_pallet,
                        'machine_plan': machine_plan,
                        'plan_pallet': plan_pallet,

                        
                        'remain_ton':remain_ton,
                        'plan_ton':plan_ton,
                        'done_ton':done_ton,

                        'losscount':losscount,
                        'losscount_plt':losscount_plt,
                        'losscount_ton':losscount_ton,

                        'change':0
                    })


            # ดูของเครื่องอื่น
            for item in result_data:
                zca = item['materialcode']
                week = item['planweek']
                matching_entries = {
                    k: v for k, v in done_dict.items() if k[0] == zca and k[2] == week
                }


                if matching_entries:
                    for key, done_entry in matching_entries.items():
                        if item['done']==0:
                            item['listmc_done'].pop()
                            item['listmc_done'].append(key[1])
                            item['change']=1
                        elif item['done']!=0:
                            item['listmc_done'].append(key[1])
                            item['change']=1
                        done_data_out=done_entry.get('total_goodcount',0)
                        total_plan=item['plan']
                        item['done']+=done_data_out
                        item['remain']=total_plan-item['done']
                        print(item['remain'])
                        item['remain_pallet'] = round(float(item['remain']) / item['pcsperpallet'], 2)
                        item['remain_ton'] = float(item['remain']) * (kgpcs / 1000)
                        remove=done_dict.pop(key,{})
            print(done_dict)

            # Return the result
            return Response({
                'success': True,
                'data': result_data,
                'count_finish': count_finish,
                'count_unfinish': count_unfinish,
                'ton_done': ton_done,
                'ton_undone': ton_undone,
                'total_pallet_undone': total_pallet_undone,
                'total_pallet_done': total_pallet_done,
                'machine_list': list(set(machine_list)),
                'machine_done': machine_done,
                'machine_undone': machine_undone,
                'machine_undone_pallet': machine_undone_pallet,
                'machine_done_pallet': machine_done_pallet,
                'machine_undone_ton': machine_undone_ton,
                'machine_done_ton': machine_done_ton,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class get_sent_approve(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | PlantCoPerm | AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            zca = request.query_params.get('zca')
            machines = request.query_params.getlist('listmc_done[]')
            week_request = request.query_params.get('week')
            date = request.query_params.get('startdate')
            week_str = week_request.split('_')
            week = week_str[1]
            year = week_str[0]

            parsed_date = datetime.fromisoformat(date.replace('Z', '+00:00'))
            year = parsed_date.year

            if not all([zca, machines, week]):
                return Response({'success': False, 'error': 'Missing parameters'}, status=status.HTTP_400_BAD_REQUEST)

            date_range = get_week_range(week, year)

            plandata = Tempactiveplan.objects.filter(
                materialcode=zca,
                machine__in=machines, 
                starttime__gte=date_range[0],
                starttime__lte=date_range[1],
            ).values(
                'materialcode','starttime', 'machine', 'plancount', 'duration','shift'
            )


            done_data = ViewWmsListfillplanproduction.objects.filter(
                zca_on=zca,
                machine__in=machines, 
                product_date__gte=date_range[0],
                product_date__lte=date_range[1],
                approve_fill='success',
            ).values(
                'zca_on', 'machine', 'pcsperpallet', 'name_th', 'qty_good', 'qty_loss', 'qty_lab',
                'send_date', 'send_shift', 'kgpcs', 'product_date', 'operator_keyin'
            )

            done_data_log = list(done_data)


            operator_ids = {entry['operator_keyin'] for entry in done_data_log if entry['operator_keyin']}
            operator_names = {
                user.employee_id: f"{user.first_name} {user.last_name}"
                for user in ViewWmsCustomuser.objects.filter(employee_id__in=operator_ids)
            }
            print(operator_names)

            for entry in done_data_log:
                operator_id = entry.get('operator_keyin')
                if operator_id:
                    operator_id = operator_id
                entry['operator_name'] = operator_names.get(operator_id, "Unknown")

                if entry['send_date']:
                    entry['send_date'] = entry['send_date'].strftime('%d-%m-%Y')
                else:
                    entry['send_date'] = "No Date"

                if entry['product_date']:
                    entry['product_date'] = entry['product_date'].strftime('%d-%m-%Y')
                else:
                    entry['product_date'] = "No Date"

            return Response({
                'success': True,
                'event_log': done_data_log,
                'plan': plandata,
            }, status=status.HTTP_200_OK)

        except ValueError as ve:
            return Response({'success': False, 'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class get_plan_pef(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            
            # Retrieve start_date and end_date from query parameters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            count_finish = 0
            count_unfinish = 0
            ton_done = ton_undone = 0
            total_pallet_done = total_pallet_undone = 0
            machine_list = []

            machine_done = {}
            machine_undone = {}




            if not start_date or not end_date:
                return Response({
                    'success': False,
                    'message': 'Both start_date and end_date are required.'
                }, status=status.HTTP_400_BAD_REQUEST)


            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')


            start_date += timedelta(days=1)
            end_date += timedelta(days=2, seconds=-1)
            start_week =get_week_number(start_date)
            end_week =get_week_number(end_date)
            weeklist=[]
            year = start_date.year

            for week in range(start_week, end_week + 1):
                weeklist.append(f"{year}_{week}")

            print(weeklist)

            plan_data_wk = Tempactiveplan.objects.filter(
                planweek__gte=weeklist[0],
                planweek__lte=weeklist[-1]
            ).values('planweek').annotate(latest_version=Max('versionno'))



            latest_versions = []
            for wk in plan_data_wk:
                latest_versions.append(wk['latest_version'])

            plan_data = Tempactiveplan.objects.filter(
                planweek__gte=weeklist[0],
                planweek__lte=weeklist[-1],
                versionno__in=latest_versions
            ).values('materialcode', 'machine', 'planweek' ,'versionno','starttime','shift').annotate(
                total_plancount=Sum('plancount')
            )



            latest_versions_map = {wk['planweek']: wk['latest_version'] for wk in plan_data_wk}


            done_data = ViewWmsListfillplanproduction.objects.filter(
                product_date__gte=start_date,
                product_date__lte=end_date,
                approve_fill='success'
            ).values('zca_on', 'machine', 'pcsperpallet' ,'name_th','product_date','product_shift').annotate(
                total_product=Sum('qty_good'))


            done_data_map = {
                (item['zca_on'], item['machine'], item['product_date'], item['product_shift']): item['total_product']
                for item in done_data
            }



            result_data = []
            machine_done_ton = {}
            machine_undone_ton = {}
            machine_done_pallet = {}
            machine_undone_pallet = {}


            for plan in plan_data:

                lookup_key = (
                plan['materialcode'],
                plan['machine'],
                plan['starttime'],
                plan['shift']
    )
                planweek = plan['planweek']
                versionno = plan['versionno']

                if planweek in latest_versions_map and versionno == latest_versions_map[planweek]: #ver ล่าสุด
                    listmc_done=[]
                    zca = plan['materialcode']

                    plancount = plan['total_plancount'] or 0  
                    plandate=plan['starttime']
                    planshift=plan['shift']
                    machine_plan = plan['machine']
                    listmc_done.append(machine_plan)

                    total_product = done_data_map.get(lookup_key, 0)
                    name_th = done_data_map.get('name_th')
                    product_date=done_data_map.get('product_date')
                    product_shift=done_data_map.get('product_shift')

                    if  not name_th:
                        if "ZCAW" in zca:
                            product = ViewItemmasterproductwip.objects.filter(field_zca=zca).first()
                            if product:
                                name_th = name_th or product.field_name

                        elif "ZCA" in zca and "ZCAW" not in zca:
                            product = ViewItemmasterproductfg.objects.filter(zca=zca).first()
                            if product:
                                name_th = name_th or product.name

                    remain_ton_t = remain = plancount - total_product


                    machine_list.append(machine_plan)

                    if remain <= 0:
                        machine_done.setdefault(machine_plan, []).append(zca)
                        count_finish += 1
                    else:
                        machine_undone.setdefault(machine_plan, []).append(zca)
                        count_unfinish += 1

                    # Append results to the list
                    result_data.append({
                        'materialcode': zca,
                        'name_th': name_th,
                        'planshift':planshift,
                        'plandate':plandate,
                        'product_date':product_date,
                        'product_shift':product_shift,
                        'planweek':planweek,
                        'listmc_done':listmc_done,

                        'plan': plancount,
                        'done': total_product,
                        'remain': remain,
                        'machine_plan': machine_plan,

                        'change':0
                    })
                else:
                    print(plan)



            # print(result_data)
            return Response({
                'success': True,
                'data': result_data,
                'count_finish': count_finish,
                'count_unfinish': count_unfinish,
                'ton_done': ton_done,
                'ton_undone': ton_undone,
                'total_pallet_undone': total_pallet_undone,
                'total_pallet_done': total_pallet_done,
                'machine_list': list(set(machine_list)),
                'machine_done': machine_done,
                'machine_undone': machine_undone,
                'machine_undone_pallet': machine_undone_pallet,
                'machine_done_pallet': machine_done_pallet,
                'machine_undone_ton': machine_undone_ton,
                'machine_done_ton': machine_done_ton,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)






def generate_empty_shifts_for_date(date, machine, start_sequence):
    """Generate empty shifts for a given date and machine."""
    shifts = []
    sequence = start_sequence
    for shift in ['A', 'B', 'C']:
        shifts.append({
            'PlanID': f"{date.strftime('%y%m%d')}{shift}_{machine}_S{sequence:04d}",
            'Material': None,
            'Machine': machine,
            'Quantity': 0,
            'StartDate': date.strftime('%d/%m/%Y'),
            'Shift': shift,
            'Duration': 480,  # Default to the full shift duration
            'OperationCode': 'SD',
        })
        sequence += 1
    return shifts, sequence

def determine_shift(date_time):
    """Determine shift based on the time of day."""
    hour = date_time.hour
    if 8 <= hour < 16:
        return 'A'
    elif 16 <= hour < 24:
        return 'B'
    else:
        return 'C'

def manage_shift_events(events, max_duration=480):
    """Split events to fit within the shift's maximum duration."""
    split_events = []
    current_duration = 0
    
    for event in events:
        event_duration = int((event.date_end - event.date_start).total_seconds() / 60)
        event_shift = determine_shift(event.date_start)
        
        # If event fits entirely within the remaining shift duration
        if current_duration + event_duration <= max_duration:
            split_events.append({
                'PlanID': f"{event.date_start.strftime('%y%m%d')}{event_shift}_{event.machine}_{event.zca}",
                'Material': event.zca,
                'Machine': event.machine,
                'Quantity': event.stk_frozen,
                'StartDate': event.date_start.strftime('%d/%m/%Y'),
                'Shift': event_shift,
                'Duration': event_duration,
                'OperationCode': 'a',
            })
            current_duration += event_duration
        else:
            # Calculate how much of the event can fit in the remaining shift time
            remaining_time = max_duration - current_duration
            if remaining_time > 0:
                split_events.append({
                    'PlanID': f"{event.date_start.strftime('%y%m%d')}{event_shift}_{event.machine}_{event.zca}",
                    'Material': event.zca,
                    'Machine': event.machine,
                    'Quantity': event.stk_frozen,
                    'StartDate': event.date_start.strftime('%d/%m/%Y'),
                    'Shift': event_shift,
                    'Duration': remaining_time,
                    'OperationCode': 'a',
                })
                current_duration += remaining_time
            
            # Mark the event as continued in the next shift, or handle overflow
            overflow_duration = event_duration - remaining_time
            if overflow_duration > 0:
                # Handle the overflow (could be handled by splitting into multiple shifts)
                pass
    
    return split_events

class get_pis_csv(APIView):
    permission_classes = [AdminPerm | PlannerPerm]

    def get(self, request, *args, **kwargs):
        week_start_date, week_end_date = get_current_week_range()

        machines = [
            "HS3", "HS4", "HS5", "HS6", "HS7", "HS8", "HS9",
            "CM5", "CM6", "CM7", "CM8",
            "OS1", "PK5", "SEG1", "OC2", "RT1", "XY1",
            "MS1", "CT1", "CT4", "SD1", "DP1", "AS1", "DET1",
            "PL1", "PK2", "PK3", "PK4", "DP2", "RT2", "PK6",
        ]

        data = []
        sequence = 1
        for machine in machines:
            frozen_mat_queryset = FrozenMat.objects.filter(
                Q(date_start__date__gte=week_start_date) & 
                Q(date_end__date__lte=week_end_date) &
                Q(machine=machine)
            )
            
            # print(week_start_date,'week_start_date')
            today = datetime.now().date()
            year = today.year
            month = today.month
            start_date = date(year, month, 1)
            days_in_month = date(year,month,calendar.monthrange(year, month)[1])
            while start_date <= days_in_month:
                empty_shifts, sequence = generate_empty_shifts_for_date(start_date, machine, sequence)
                
                day_events = [mat for mat in frozen_mat_queryset if mat.date_start.date() == start_date]

                # Track the total duration per shift
                shift_duration = {'A': 0, 'B': 0, 'C': 0}
                shift_events = {'A': [], 'B': [], 'C': []}

                for event in day_events:
                    event_shift = determine_shift(event.date_start)
                    shift_events[event_shift].append(event)
                
                # Process each shift to ensure durations fit within limits
                for shift_key in ['A', 'B', 'C']:
                    processed_events = manage_shift_events(shift_events[shift_key])
                    data.extend(processed_events)
                
                # Add empty shifts if no events are present
                for shift in empty_shifts:
                    if not shift_events[shift['Shift']]:
                        data.append(shift)
                
                start_date += timedelta(days=1)
        weeknum=get_week_number()
        return JsonResponse({'success': True, 'data': data, 'week': weeknum})
    
    
def convert_date_to_DMY(date_string):
    try:
        new_date_str = date_string.strftime("%d-%m-%Y")
    except:
        date_obj = datetime.strptime(date_string, "%Y-%m-%d")
        new_date_str = date_obj.strftime("%d-%m-%Y")
    return new_date_str


# ================================================ Automate line noti ===============================





def filter_and_generate_negative_stock_text(data):
    text_list = []
    
    for plan in data:
        plan_id = plan["PlanId"]
        plan_text = f"PlanID {plan_id}\n"
        plan_records = []
        
        for code, records in plan["data"].items():
            for record in records:
                if record["stock"] < 0:
                    # Fetch the first child name if available
                    child_name = record["child"][0]["child_name"] if record["child"] else "No Child"
                    record_text = f"-{record['name']}, ของขาด {record['stock']} แผ่น ในวันที่ {record['date_start']} Material ที่ใช้ {child_name}\n  "
                    plan_records.append(record_text)
        
        if plan_records:
            text_list.append(plan_text + ''.join(plan_records))
    
    return text_list



#  <==============================================================>  





#  <======================= UploadExcel <=======================>
import decimal

def convert_date_format(date_input):
    """Convert either an Excel serial date (float) or a 'DD/MM/YYYY' date string to 'YYYY-MM-DD'."""
    if isinstance(date_input, float):
        # Handle Excel serial date (Excel's date system starts on 1899-12-30)
        base_date = datetime(1899, 12, 30)
        try:
            # Add the number of days from the base date
            converted_date = base_date + timedelta(days=date_input)
            return converted_date.strftime('%Y-%m-%d')
        except (TypeError, ValueError):
            return None
    elif isinstance(date_input, str):
        # Handle 'DD/MM/YYYY' date strings
        try:
            date_obj = datetime.strptime(date_input, '%d/%m/%Y')
            return date_obj.strftime('%Y-%m-%d')
        except ValueError:
            return None
    else:
        return None
    
class upload_plan(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def post(self, request, *args, **kwargs):
        try:
            # Extract the uploaded file
            today = timezone.now()

            full_name = request.data.get('full_name')
            file = request.FILES['file']
            if not file.name.endswith('.xlsb'):
                return Response({'success': False, 'error': 'Invalid file format. Please upload an .xlsb file.'}, status=400)

            workbook = open_workbook(file)
            file_name = file.name  
            file_base_name = os.path.splitext(file_name)[0] 
            parts = file_base_name.split('-')
            week_file = parts[-2]
            week_file_num=int(week_file[1:])

            revision = parts[-1]  
            version=int(revision[-2:])
            
            # Check if the 'PIS' sheet exists
            if 'PIS' not in workbook.sheets:
                return Response({'success': False, 'error': "Sheet 'PIS' not found."}, status=400)
            sheet = workbook.get_sheet('PIS')
            
            data_inserted = 0


            with transaction.atomic():
                for i, row in enumerate(sheet.rows()):
                    if i == 0:  # Skip header row
                        continue
                    
    
                    row_data = [item.v for item in row]
                    

                    if row_data[9] == 'a' and row_data[3]:  
                        def safe_decimal(value, default=0):
                            try:
                                # Ensure the decimal has two places of precision (adjust as needed)
                                return decimal.Decimal(value).quantize(decimal.Decimal('0.00'), rounding=decimal.ROUND_DOWN)
                            except (TypeError, decimal.InvalidOperation):
                                # Set default precision for zero or any invalid values
                                return decimal.Decimal(default).quantize(decimal.Decimal('0.00'))

                        start_time_str = convert_date_format(row_data[6])
                        start_time = datetime.strptime(start_time_str, '%Y-%m-%d')
                        year_start = start_time.year
                        plan_week = get_week_number(start_time)  # Calculate week number
                        year_week = str(year_start)+'_'+str(plan_week)
                        plan_id=row_data[2]
                        prod_no=row_data[1]
                        prod_no = prod_no.replace('.', '_')
                        planname_comp=plan_id+'_'+prod_no
                        machine=row_data[4]
                        if machine == 'SEG1':
                            machine = 'SEG'

                        elif machine == 'DET1':
                            machine ='DET'
                        if week_file_num == plan_week:
                            obj, created = Tempactiveplan.objects.update_or_create(
                                starttime=start_time_str,
                                machine=machine,
                                productnum=prod_no,
                                shift=row_data[7],
                                defaults={
                                    'company': 'SFCG',
                                    'plant': 'C221',
                                    'shift': row_data[7],
                                    'planweek': year_week,  
                                    'versionno': version,
                                    'planname':planname_comp,
                                    'machine': machine,
                                    'materialcode': row_data[3],
                                    'materialname': '',
                                    'planweight': 0,
                                    'plancount': safe_decimal(row_data[5], 0),
                                    'starttime': start_time_str,
                                    'duration': safe_decimal(row_data[8], 0),
                                    'operationcode': row_data[9],
                                    'setupduration': safe_decimal(row_data[10], 0),
                                    'cleaningduration': safe_decimal(row_data[11], 0),
                                    'startupduration': safe_decimal(row_data[12], 0),
                                    'ms': safe_decimal(row_data[19], 0),
                                    'es': safe_decimal(row_data[20], 0),
                                    'os': safe_decimal(row_data[21], 0),
                                    'mc': safe_decimal(row_data[22], 0),
                                    'ec': safe_decimal(row_data[23], 0),
                                    'oc': safe_decimal(row_data[24], 0),
                                    'msu': safe_decimal(row_data[25], 0),
                                    'esu': safe_decimal(row_data[26], 0),
                                    'osu': safe_decimal(row_data[27], 0),
                                    'contractor': '',
                                    'buname': '',
                                    'description': '',
                                    'created_by':full_name,
                                    'create_at':today,
                                }
                            )
                            print(created ,'cre \n')
                            if created:
                                data_inserted += 1  # Count only newly inserted records
            
            return Response({'success': True, 'message': f'{data_inserted} new rows inserted.'}, status=200)
        
        except Exception as e:
            traceback.print_exc()  # Print full stack trace in server logs for debugging
            return Response({'success': False, 'error': str(e)}, status=500)




import os
import pandas as pd 
from django.conf import settings 
from openpyxl import load_workbook  # pip install openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.styles import Font
from openpyxl.styles import Alignment

class export_to_excel(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def post(self, request, *args, **kwargs):
        try:

            template_path = os.path.join(settings.BASE_DIR, 'wms/static/assets/data/FG-GG-AD-001 (ใบกำหนดแจ้งพื้นที่กองเก็บ)_Excel_JengPreme2.xlsx')
            print(f"File exists: {os.path.exists(template_path)}") 
            

            wb = load_workbook(template_path)
            ws = wb.active  


            data = request.data.get('data')
            if not data:
                return Response({'success': False, 'error': 'No data provided'}, status=400)

            # Convert data to a pandas DataFrame

            df = pd.DataFrame(data)
            rearranged_columns = ['materialcode', 'name_th', 'machine_plan','listmc_done', 'pcsperpallet', 'plan', 'done', 'remain', 'remain_pallet', 'done_pallet', 'plan_pallet',]
            df['listmc_done'] = df['listmc_done'].apply(lambda x: ' / '.join(map(str, x)) if isinstance(x, list) else str(x))
            df = df[rearranged_columns]
            
            alignment_center = Alignment(horizontal='center', vertical='center')  
            black_font = Font(color="000000")
            merged_cells = ws.merged_cells.ranges  

            for row_num, row_data in enumerate(df.values, start=8):  # Start inserting from row 8
                for col_num, cell_value in enumerate(row_data, start=2):  # Start inserting from column 2
                    is_merged = False
                    for merged_cell in merged_cells:
                        min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
                        if min_row <= row_num <= max_row and min_col <= col_num <= max_col:
                            if row_num == min_row and col_num == min_col:
                                # print(row_num,col_num,cell_value)
                                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                                cell.font = black_font  # Set font color to black
                                cell.alignment = alignment_center
                            is_merged = True
                            break
                    if not is_merged:
                        # print('m',row_num,col_num,cell_value)
                        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                        cell.font = black_font  # Set font color to black
                        cell.alignment = alignment_center


            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="updated_template.xlsx"'

            # Save the updated workbook to the response
            wb.save(response)

            return response

        except Exception as e:
            traceback.print_exc()  # Print full stack trace for debugging
            return Response({'success': False, 'error': str(e)}, status=500)
        


class get_plans(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            start_date += timedelta(days=1)
            end_date += timedelta(days=1)

            print(start_date,end_date)


            plan_data = Tempactiveplan.objects.filter(
                starttime__gte=start_date,
                starttime__lte=end_date
            )

            if not plan_data.exists():
                return Response({'success': False, 'error': 'No active plans found'}, status=status.HTTP_404_NOT_FOUND)

            # Prepare the list of plans with additional materialname lookup
            plan_data_list = []
            for plan in plan_data:
                zca = plan.materialcode  
                materialname = None
                pcsperpallet = None
                kgpcs = None

                # Lookup based on zca
                if "ZCAW" in zca:
                    product = ViewItemmasterproductwip.objects.filter(field_zca=zca).first()
                    if product:
                        pcsperpallet = product.pcsperpallet if pcsperpallet is None or pcsperpallet == 1 else pcsperpallet
                        kgpcs = product.field_kgpcs if kgpcs is None or kgpcs == 1 else kgpcs
                        materialname = product.field_name

                elif "ZCA" in zca and "ZCAW" not in zca:
                    product = ViewItemmasterproductfg.objects.filter(zca=zca).first()
                    if product:
                        pcsperpallet = product.pcpallet if pcsperpallet is None or pcsperpallet == 1 else pcsperpallet
                        kgpcs = product.kg if kgpcs is None or kgpcs == 1 else kgpcs
                        materialname = product.name


                # Append to the result list
                plan_data_list.append({
                    'company': plan.company,
                    'plant': plan.plant,
                    'shift': plan.shift,
                    'planweek': plan.planweek,
                    'versionno': plan.versionno,
                    'planname': plan.planname,  # Primary key
                    'machine': plan.machine,
                    'materialcode': plan.materialcode,
                    'materialname': materialname,  # Material name based on zca lookup
                    'planweight': round(float(plan.plancount) * float(kgpcs), 2) if plan.plancount and kgpcs is not None else None,
                    'planpallet': round(float(plan.plancount) / float(pcsperpallet), 2) if plan.plancount and pcsperpallet is not None else None,
                    'plancount': float(plan.plancount) if plan.plancount else None,




                    'starttime': plan.starttime.strftime('%Y-%m-%d') if plan.starttime else None,
                    'duration': float(plan.duration) if plan.duration else None,
                    'operationcode': plan.operationcode,
                    'setupduration': float(plan.setupduration) if plan.setupduration else None,
                    'cleaningduration': float(plan.cleaningduration) if plan.cleaningduration else None,
                    'startupduration': float(plan.startupduration) if plan.startupduration else None,
                    'ms': float(plan.ms) if plan.ms else None,
                    'es': float(plan.es) if plan.es else None,
                    'os': float(plan.os) if plan.os else None,
                    'mc': float(plan.mc) if plan.mc else None,
                    'ec': float(plan.ec) if plan.ec else None,
                    'oc': float(plan.oc) if plan.oc else None,
                    'msu': float(plan.msu) if plan.msu else None,
                    'esu': float(plan.esu) if plan.esu else None,
                    'osu': float(plan.osu) if plan.osu else None,
                    'contractor': plan.contractor,
                    'buname': plan.buname,
                    'description': plan.description,
                    'created_by': plan.created_by,
                })

            return Response({'success': True, 'plans': plan_data_list})

        except Exception as e:
            traceback.print_exc() 
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class get_version(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')

            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_date += timedelta(days=1)
            end_date += timedelta(days=1)
            start_year=start_date.year
            end_year=end_date.year

            print('wk',start_date,end_date)
            week_start=get_week_number(start_date)
            week_end=get_week_number(end_date)
            year_week_start=str(start_year)+'_'+str(week_start)
            year_week_end=str(end_year)+'_'+str(week_end)
            # Filter plans within the date range
            plan_data = Tempactiveplan.objects.filter(
                planweek__gte=year_week_start,
                planweek__lte=year_week_end
            ).values('planweek').annotate(latest_version=Max('versionno'))

            if not plan_data.exists():
                return Response({'success': False, 'message': 'No plans found for the given date range'}, status=status.HTTP_404_NOT_FOUND)

            # Prepare the result dictionary for each week and its latest version
            week_version_data = {plan['planweek']: plan['latest_version'] for plan in plan_data}

            return Response({'success': True, 'week_versions': week_version_data})


        except Exception as e:
            traceback.print_exc() 
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class get_reject(APIView):
    permission_classes = [ProductionPerm |PlannerPerm|PlantCoPerm| AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            # Extract and validate date parameters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            machines_list = request.query_params.getlist('machines[]')
            print(machines_list)
            if not start_date or not end_date:
                return Response({'success': False, 'error': 'Both start_date and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return Response({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

            # Adjust dates by adding 1 day
            start_date += timedelta(days=1)
            end_date += timedelta(days=1)

            # Fetch raw reject data within the date range
            raw_reject = ViewWmsListfillplanproduction.objects.filter(
                send_date__gte=start_date,
                send_date__lte=end_date,
                machine__in=machines_list,
            ).values('machine', 'qty_good', 'qty_loss', 'kgpcs', 'pcsperpallet', 'zca_on', 'name_th')

            # Dictionary to hold aggregated data by machine
            reject_data_by_machine = defaultdict(lambda: {
                'total_good_qty': 0,
                'total_reject_qty': 0,
                'sum_ton_reject': 0,
                'sum_ton_good': 0
            })

            # Dictionary to hold aggregated data by zca and machine
            reject_data_by_zca = defaultdict(lambda: {
                'total_good_qty': 0,
                'total_reject_qty': 0,
                'sum_ton_reject': 0,
                'sum_ton_good': 0
            })

            # Iterate and aggregate data for each machine and zca
            for item in raw_reject:
                machine = item['machine'] if item['machine'] else ''
                name_th = item['name_th'] if item['name_th'] else ''
                qty_good = item['qty_good'] if item['qty_good'] is not None else 0
                qty_loss = item['qty_loss'] if item['qty_loss'] is not None else 0
                kgpcs = item['kgpcs'] if item['kgpcs'] is not None else 0

                # Update the aggregated values for machine
                reject_data_by_machine[machine]['total_good_qty'] += qty_good
                reject_data_by_machine[machine]['total_reject_qty'] += qty_loss
                reject_data_by_machine[machine]['sum_ton_reject'] += qty_loss * kgpcs / 1000 if kgpcs else 0
                reject_data_by_machine[machine]['sum_ton_good'] += qty_good * kgpcs / 1000 if kgpcs else 0

                # Update the aggregated values for zca and machine
                reject_data_by_zca[(name_th, machine)]['total_good_qty'] += qty_good
                reject_data_by_zca[(name_th, machine)]['total_reject_qty'] += qty_loss
                reject_data_by_zca[(name_th, machine)]['sum_ton_reject'] += qty_loss * kgpcs / 1000 if kgpcs else 0
                reject_data_by_zca[(name_th, machine)]['sum_ton_good'] += qty_good * kgpcs / 1000 if kgpcs else 0

            # Convert machine data dictionary to list format and calculate reject percentage
            reject_data_mc_total = []
            for machine, data in reject_data_by_machine.items():
                total_good_qty = data['total_good_qty']
                total_reject_qty = data['total_reject_qty']
                
                reject_data_mc_total.append({
                    'machine': machine,
                    'total_reject_qty': total_reject_qty,
                    'total_good_qty': total_good_qty,
                    'reject_percentage_qty': round((total_reject_qty / total_good_qty * 100), 2) if total_good_qty else 0,
                    'good_percentage_qty': round((total_good_qty / (total_good_qty+total_reject_qty) * 100), 2) if total_good_qty else 0,
                    'sum_ton_reject': round(data['sum_ton_reject'], 2),
                    'sum_ton_good': round(data['sum_ton_good'], 2)
                })

            # Convert zca and machine data to a list format
            formatted_reject_data_zca_mc_total = []
            for (name_th, machine), data in reject_data_by_zca.items():
                total_good_qty = data['total_good_qty']
                total_reject_qty = data['total_reject_qty']
                if total_reject_qty > 0:
                    print(name_th,total_reject_qty)
                    formatted_reject_data_zca_mc_total.append({
                        'name_th': name_th,
                        'machine': machine,
                        'total_good_qty': total_good_qty,
                        'total_reject_qty': total_reject_qty,
                        'sum_ton_reject': round(data['sum_ton_reject'], 2),
                        'sum_ton_good': round(data['sum_ton_good'], 2),
                        'reject_percentage_qty': round((total_reject_qty / total_good_qty * 100), 2) if total_good_qty else 0
                    })

            formatted_reject_data_zca_mc_total = sorted(
                formatted_reject_data_zca_mc_total, 
                key=lambda x: x['reject_percentage_qty'], 
                reverse=True
            )
            print(formatted_reject_data_zca_mc_total)
            return Response({
                'success': True,
                'reject_data_mc_total': reject_data_mc_total,
                'reject_data_zca_mc_total': formatted_reject_data_zca_mc_total
            })

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        



class get_operator_pef(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | PlantCoPerm | AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            # Extract and validate date parameters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            machines_list = request.query_params.getlist('machines[]')
            # print('selected machine',machines_list)
            if not start_date or not end_date:
                return Response({'success': False, 'error': 'Both start_date and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return Response({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
            # print('before',start_date,end_date)
            start_date += timedelta(days=1)
            end_date += timedelta(days=1)
            print('after',start_date,end_date)



            # Fetch rawdata จากการส่งยอด
            performance_data = ViewWmsListfillplanproduction.objects.filter(
                send_date__gte=start_date,
                send_date__lte=end_date,
                machine__in=machines_list,
                approve_fill='success',
            ).values(
                'operator_keyin', 'machine', 'zca_on', 'name_th', 'send_date'
            ).annotate(
                total_qty_good=Sum('qty_good'),
                total_qty_loss=Sum('qty_loss')

            ).order_by('send_date')


            
            # Fetch operator names 
            operator_ids = {item['operator_keyin'] for item in performance_data if item['operator_keyin']}
            # print('operator all',operator_ids)
            # print(ViewWmsCustomuser.objects.filter(employee_id__in=operator_ids))
            operator_names = {
                user.employee_id: f"{user.first_name} {user.last_name}"
                for user in ViewWmsCustomuser.objects.filter(employee_id__in=operator_ids)
            }
            # print('operator_names',operator_names)

            #{'รหัสพนังงาน':data[]}
            operator_performance = {}
            for item in performance_data:
                print(item,'row')
                operator = item['operator_keyin']
                machine = item['machine']
                zca_on = item['zca_on']
                name = item['name_th']
                send_date = item['send_date']
                qty_good = item['total_qty_good'] or 0
                qty_loss = item['total_qty_loss'] or 0
                reject_percent = qty_loss / qty_good if qty_good else 0

                # ใน operator_performance ยังไม่เจอพนังงาน
                if operator not in operator_performance:

                    operator_performance[operator] = {
                        'operator_name': operator_names.get(operator, "Unknown"),
                        'total_qty_good': 0,
                        'total_qty_loss': 0,
                        'reject_percent': 0,
                        'details': []
                    }

                operator_performance[operator]['total_qty_good'] += qty_good
                operator_performance[operator]['total_qty_loss'] += qty_loss


                operator_performance[operator]['details'].append({
                    'machine': machine,
                    'zca_on': zca_on,
                    'name': name,
                    'send_date': send_date,
                    'total_qty_good': qty_good,
                    'total_qty_loss': qty_loss,
                    'reject_percent': reject_percent
                })


            for operator, data in operator_performance.items():
                total_good = data['total_qty_good']
                total_loss = data['total_qty_loss']
                data['reject_percent'] = total_loss / total_good if total_good else 0

            print(operator_performance)
            
            return Response({
                'success': True,
                'data': operator_performance,
            })

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




from django.db.models.functions import ExtractYear, ExtractMonth

from dateutil.relativedelta import relativedelta

class get_prod_timeserie(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | PlantCoPerm | AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            # Extract and validate date parameters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            zca_list = request.query_params.getlist('zca[]')
            
            if not start_date or not end_date:
                return Response({'success': False, 'error': 'Both start_date and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return Response({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

            start_date += timedelta(days=1)
            end_date += timedelta(days=1)

            filter_params = {
                'prodorderdate__gte': start_date,
                'prodorderdate__lte': end_date,
            }
            if zca_list:
                filter_params['materialcode__in'] = zca_list

            Pastproduct = ViewPisMergereportProduction.objects.filter(**filter_params).annotate(
                year=ExtractYear('prodorderdate'),
                month=ExtractMonth('prodorderdate')
            ).values(
                'year', 'month', 'materialcode'
            ).annotate(
                total_goodcount=Sum('goodcount'),
                total_goodweight=Sum('goodweight')
            ).order_by('year', 'month', 'materialcode')

            # Generate all months between start_date and end_date
            all_periods = []
            current_date = datetime(start_date.year, start_date.month, 1)
            end_period = datetime(end_date.year, end_date.month, 1)
            while current_date <= end_period:
                all_periods.append(current_date.strftime("%b-%Y"))
                current_date += relativedelta(months=1)

            # Initialize formatted data with all zca codes and zero counts for each period
            formatted_data = {zca: {period: 0 for period in all_periods} for zca in zca_list}
            formatted_data_ton = {zca: {period: 0 for period in all_periods} for zca in zca_list}

            # Fill formatted_data with actual data from Pastproduct query
            for item in Pastproduct:
                materialcode = item['materialcode']
                period = f"{datetime(year=item['year'], month=item['month'], day=1):%b-%Y}"
                total_goodcount = item['total_goodcount'] if item['total_goodcount'] is not None else 0
                total_goodweight = item['total_goodweight'] if item['total_goodweight'] is not None else 0
                formatted_data[materialcode][period] = total_goodcount
                formatted_data_ton[materialcode][period] = total_goodweight
            # Convert to the required list format for the response
            final_data = {
                materialcode: [{'x': period, 'y': count} for period, count in periods.items()]
                for materialcode, periods in formatted_data.items()
            }
            final_data_ton = {
                materialcode: [{'x': period, 'y': count} for period, count in periods.items()]
                for materialcode, periods in formatted_data_ton.items()
            }


            return Response({
                'success': True,
                'Pastproduct': final_data,
                'Pastproductton':final_data_ton
            })

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class get_mc_timeserie(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | PlantCoPerm | AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            # Extract and validate date parameters
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            machinelist = request.query_params.getlist('machineFilter[]')
            print(machinelist)
            if not start_date or not end_date:
                return Response({'success': False, 'error': 'Both start_date and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return Response({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

            start_date += timedelta(days=1)
            end_date += timedelta(days=1)

            filter_params = {
                'prodorderdate__gte': start_date,
                'prodorderdate__lte': end_date,
            }
            if machinelist:
                filter_params['machine__in'] = machinelist

            Pastproduct = ViewPisMergereportProduction.objects.filter(**filter_params).annotate(
                year=ExtractYear('prodorderdate'),
                month=ExtractMonth('prodorderdate')
            ).values(
                'year', 'month', 'machine'
            ).annotate(
                total_goodcount=Sum('goodcount'),
                total_reject=Sum('goodcount'),
                total_goodweight=Sum('goodweight')
            ).order_by('year', 'month', 'machine')

            print(Pastproduct)
            # Generate all months between start_date and end_date
            all_periods = []
            current_date = datetime(start_date.year, start_date.month, 1)
            end_period = datetime(end_date.year, end_date.month, 1)
            while current_date <= end_period:
                all_periods.append(current_date.strftime("%b-%Y"))
                current_date += relativedelta(months=1)

            # Initialize formatted data with all zca codes and zero counts for each period
            formatted_data = {zca: {period: 0 for period in all_periods} for zca in machinelist}
            formatted_data_ton = {zca: {period: 0 for period in all_periods} for zca in machinelist}

            # Fill formatted_data with actual data from Pastproduct query
            for item in Pastproduct:
                machine = item['machine']
                period = f"{datetime(year=item['year'], month=item['month'], day=1):%b-%Y}"
                total_goodcount = item['total_goodcount'] if item['total_goodcount'] is not None else 0
                total_goodweight = item['total_goodweight'] if item['total_goodweight'] is not None else 0
                formatted_data[machine][period] = total_goodcount
                formatted_data_ton[machine][period] = total_goodweight
            # Convert to the required list format for the response
            final_data = {
                machine: [{'x': period, 'y': count} for period, count in periods.items()]
                for machine, periods in formatted_data.items()
            }
            final_data_ton = {
                machine: [{'x': period, 'y': count} for period, count in periods.items()]
                for machine, periods in formatted_data_ton.items()
            }


            return Response({
                'success': True,
                'Pastproduct': final_data,
                'Pastproductton':final_data_ton
            })

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class get_zca_option(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | PlantCoPerm | AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            # Retrieve unique pairs of 'zca' and 'name_th' from each model
            fg_zca_name = ViewItemmasterproductfg.objects.values('zca', 'name').distinct()
            wip_zca_name = ViewItemmasterproductwip.objects.values('field_zca', 'field_name').distinct()

            # Normalize field names to ensure consistency in the response
            fg_zca_name = [{'zca': item['zca'], 'name_th': item['name']} for item in fg_zca_name]
            wip_zca_name = [{'zca': item['field_zca'], 'name_th': item['field_name']} for item in wip_zca_name]

            # Combine both lists and remove duplicates
            combined_zca_name = {tuple(item.items()): item for item in (fg_zca_name + wip_zca_name)}.values()

            return Response({
                'success': True,
                'unique_zca_name': list(combined_zca_name)
            })

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        



def insert_result_data_into_remain_plan(result_data):
    print('inserted')
    for entry in result_data:
        planname = entry.get('plan_name')
        planweek = entry.get('planweek')
        machine = entry.get('machine')
        zca = entry.get('materialcode')
        planamount = str(entry.get('plan', 0))
        pcsperpallet = entry.get('pcsperpallet', 1)

        field_ids = entry.get('field_ids', [])
        done_list = entry.get('done', [])
        remain_list = entry.get('remain', [])

        max_length = max(len(done_list), len(remain_list), len(field_ids))

        for i in range(max_length):
            total_submit = str(done_list[i]) if i < len(done_list) else '0'
            remain_qty = str(remain_list[i]) if i < len(remain_list) else '0'
            list_fillplan_id = field_ids[i] if i < len(field_ids) else None

            if pcsperpallet and float(pcsperpallet) != 0:
                remain_qty_float = float(remain_qty)
                pcsperpallet_float = float(pcsperpallet)
                whole_pallets = int(remain_qty_float // pcsperpallet_float)
                remainder_qty = int(remain_qty_float % pcsperpallet_float)
                remain_pallet = f"{whole_pallets}({remainder_qty})"
            else:
                remain_pallet = '0(0)'

            obj, created = RemainPlan.objects.update_or_create(
                planname=planname,
                planweek=planweek,
                machine=machine,
                zca=zca,
                list_fillplan_link=list_fillplan_id,
                defaults={
                    'planname': planname,
                    'planweek': planweek,
                    'machine': machine,
                    'zca': zca,
                    'planamount': planamount,
                    'total_submit': total_submit,
                    'remain_pallet': remain_pallet,
                    'remain_qty': remain_qty,
                    'list_fillplan_link': list_fillplan_id,
                    'created': datetime.now(),
                }
            )






class Initilize_schedual(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | AdminPerm]

    @swagger_auto_schema()
    def post(self, request, *args, **kwargs):
        try:
            current_date = datetime.now().date()
            current_year = current_date.year

            # Call the store_remain function
            store_remain(current_date)

            # Return success response
            return Response({'success': True, 'message': 'Schedule initialized successfully'}, status=status.HTTP_200_OK)

        except Exception as e:
            # Log the error and return a 500 response
            print(f"Error: {e}")
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




def store_remain(day):
    year = day.year
    week = get_week_number(day)
    date_range = get_week_range(week, day.year)
    
    plan_data = Tempactiveplan.objects.filter(
        starttime__gte=date_range[0],
        starttime__lte=date_range[1]
    ).values('materialcode', 'machine', 'planweek' ,'versionno').annotate(total_plancount=Sum('plancount'))

    plan_data_wk = Tempactiveplan.objects.filter(
        starttime__gte=date_range[0],
        starttime__lte=date_range[1]
    ).values('planweek').annotate(latest_version=Max('versionno'))

    latest_versions_map = {wk['planweek']: wk['latest_version'] for wk in plan_data_wk}


    done_data = ViewWmsListfillplanproduction.objects.filter(
        product_date__gte=date_range[0],
        product_date__lte=date_range[1],
        approve_fill='success'
    ).values(
        'id', 'zca_on', 'machine', 'pcsperpallet', 'name_th', 'kgpcs', 'product_date', 'qty_good', 'qty_loss', 'qty_lab'
    )

    grouped_done_data = defaultdict(lambda: {
        'total_goodcount': [],  # Store good counts as a list
        'total_loss': 0,
        'total_lab': 0,
        'field_ids': [],
    })
    for item in done_data:
        field_id = item['id']
        product_year=item['product_date'].year
        zca = item['zca_on']
        machine = item['machine']
        week_num = get_week_number(item['product_date'])
        week_number=f"{product_year}_{week_num}"
        key = (zca, machine, week_number)

        grouped_done_data[key]['total_goodcount'].append(item['qty_good'] or 0)
        grouped_done_data[key]['total_loss'] += item['qty_loss'] or 0
        grouped_done_data[key]['total_lab'] += item['qty_lab'] or 0
        grouped_done_data[key]['pcsperpallet'] = item['pcsperpallet']
        grouped_done_data[key]['name_th'] = item['name_th']
        grouped_done_data[key]['kgpcs'] = item['kgpcs']
        grouped_done_data[key]['field_ids'].append(field_id)

    done_dict = {
        (key[0], key[1], key[2]): {
            'total_goodcount': value['total_goodcount'],
            'pcsperpallet': value['pcsperpallet'],
            'name_th': value['name_th'],
            'kgpcs': value['kgpcs'],
            'field_ids': value['field_ids']
        }
        for key, value in grouped_done_data.items()
    }

    result_data = []

    for plan in plan_data:
        planweek = plan['planweek']
        versionno = plan['versionno']

        if planweek in latest_versions_map and versionno == latest_versions_map[planweek]:
            zca = plan['materialcode']
            plancount = plan['total_plancount'] or 0  
            machine_plan = plan['machine']
            planweek = plan['planweek']
            done_entry = done_dict.get((zca, machine_plan, planweek), {})
            
            field_ids = done_entry.get('field_ids', [])
            goodcount_list = done_entry.get('total_goodcount', [])
            pcsperpallet = done_entry.get('pcsperpallet')
            name_th = done_entry.get('name_th')
            kgpcs = done_entry.get('kgpcs', 0)

            if not pcsperpallet or not name_th:
                if "ZCAW" in zca:
                    product = ViewItemmasterproductwip.objects.filter(field_zca=zca).first()
                    if product:
                        pcsperpallet = product.pcsperpallet if pcsperpallet is None or pcsperpallet == 1 else pcsperpallet
                        kgpcs = product.field_kgpcs if kgpcs is None or kgpcs == 1 else kgpcs
                        name_th = product.field_name
                elif "ZCA" in zca and "ZCAW" not in zca:
                    product = ViewItemmasterproductfg.objects.filter(zca=zca).first()
                    if product:
                        pcsperpallet = product.pcpallet if pcsperpallet is None or pcsperpallet == 1 else pcsperpallet
                        kgpcs = product.kg if kgpcs is None or kgpcs == 1 else kgpcs
                        name_th = product.name

            kgpcs = float(kgpcs) if kgpcs else 0
            pcsperpallet = float(pcsperpallet) if isinstance(pcsperpallet, str) else pcsperpallet

            # Calculate cumulative remains for each goodcount value
            remain_list = []
            remaining = plancount  # Start with the total plan count
            for goodcount in goodcount_list:
                remaining -= goodcount  # Subtract each done value from the remaining
                remain_list.append(remaining)  # Append the updated remaining value

            # Ensure that if there are no done counts, remain is set to the original plan count
            if not goodcount_list:
                remain_list = [plancount]

            result_data.append({
                'field_ids': field_ids,
                'materialcode': zca,
                'machine': machine_plan,
                'planweek': planweek,
                'pcsperpallet': pcsperpallet,
                'plan': plancount,
                'done': goodcount_list, 
                'remain': remain_list, 
                'kgpcs': kgpcs,
                'plan_name': f"{year}_{planweek}_{machine_plan}_{zca}"
            })
        else:
            print(plan)
    # print('insertdata')
    insert_result_data_into_remain_plan(result_data)













def calculate_history_mc_perf():
    machine_data = (
        ViewPisMergereportProduction.objects
        .values('materialcode','materialname','machine')  
        .annotate(
            total_qty=Sum('goodcount'),  
            avg_qty=Avg('goodcount'),   
            total_loss=Sum('rejectcount'),  
            avg_loss=Avg('rejectcount'),   
            total_run=Count('materialcode'),  
            total_run_time=Sum('runningduration'),  
            avg_run=Avg('runningduration'),  
            total_breakdown=Sum('downtimeduration'),  
            avg_breakdown=Avg('downtimeduration'),    
            reject_percentage=ExpressionWrapper(
                F('total_loss') / (F('total_qty') + F('total_loss')) * 100,
                output_field=FloatField()
            ),  
            rate=ExpressionWrapper(
                F('total_qty') / F('total_run_time'),
                output_field=FloatField()
            )  
        )
        .order_by('materialname', 'machine')  
    )
    return machine_data





class get_zca_machinerate(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | PlantCoPerm | AdminPerm]

    def get(self, request, *args, **kwargs):
        try:
            zca_list = [zca.strip() for zca in request.query_params.get('zca', '').split(',')]
            zca_rate = Masterwipstk.objects.filter(zca__in=zca_list).values()

            return Response({
                'success': True,
                'data': list(zca_rate),
            })

        except Exception as e:
            traceback.print_exc()
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class get_simulate(APIView):
    permission_classes = [ProductionPerm | PlannerPerm | PlantCoPerm | AdminPerm]

    def get(self, request, *args, **kwargs):
        try:

            zca_rate=calculate_history_mc_perf()



            return Response({
                'success': True,
                'data': zca_rate,
            })

        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

